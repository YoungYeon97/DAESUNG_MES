#-*- coding:utf-8 -*-
import os
import re
import sys
import time
import atexit
import ftplib
import serial
import shutil #file copy
import socket
import logging
import pymysql
import keyboard
import webbrowser
import configparser
import pandas as pd
import multiprocessing as mp

from tqdm import *
from os import environ
from pymodbus.client import *
from openpyxl import load_workbook
from logging.config import dictConfig
from multiprocessing import Process, Queue, freeze_support

from PyQt5 import QtCore, QtWidgets, uic, QtGui
from PyQt5.QtWidgets import * #QDialog

from PyQt5.uic import * #loadUi
from PyQt5.QtCore import * #QThread
from PyQt5.QtGui import *  #font

from DAESUNG_Query import *
from DAESUNG_Fuctions import *

RECONNECT_FLAG, JAKUP_TIME, DATA_COUNT, PRT_INDEX, PRT_FLAG, PRT_NAME = 0, 0, 0, -1, '%', '전체'
#----------------------------------------------------------------------------
s_label, ss_label = 'DAESUNG_MES_S.txt', 'DAESUNG_MES_S(S).txt'
c_label, cs_label = 'DAESUNG_MES_C.txt', 'DAESUNG_MES_C(S).txt'
z_label, zs_label, z300_label = 'DAESUNG_MES_Z.txt', 'DAESUNG_MES_Z(S).txt', 'DAESUNG_MES_Z(300).txt'
m_label, f_label = 'DAESUNG_MES_M.txt', 'DAESUNG_MES_F.txt'
if_label = 'DAESUNG_MES_IF.txt'
#----------------------------------------------------------------------------
S_PROC, SS_PROC = '0101', '0201'
C_PROC, CS_PROC = '0103', '0202'
Z_PROC, ZS_PROC = '0117', '0206'
#----------------------------------------------------------------------------
COMPANY = '4C 53 49 53 2D 58 47 54 00 00 00 00 00 33 00 00' #기본
POSTION = '00 00' #기본
BLOCK = '00 00 01 00' #기본
STX, EOT, CR = '\x02'.encode(), '\x04'.encode(), '\x0D'.encode()

#프로그램 종료 시 재시작
def main():
    executable = sys.executable
    args = sys.argv[:]
    args.insert(0, sys.executable)
    os.execvp(executable, args)

#해상도
def suppress_qt_warnings():
    environ["QT_DEVICE_PIXEL_RATIO"] = "0"
    environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    environ["QT_SCREEN_SCALE_FACTORS"] = "1"
    environ["QT_SCALE_FACTOR"] = "1"
    
def handleExit():
    try: closeDB() #DB연결 해제
    except: logging.debug("handleExit : failed")

#FTP파일 다운로드
def downLoadFiles(filename):
    try:
        ftp = ftplib.FTP()
        ftp.connect("ehandax.com", 2012)
        ftp.login("woodnsoft!", "woodnsoft.com")
        ftp.cwd('upgrade/WDNS_MES/DAESUNG_MES')
        fd = open("./" + filename, 'wb')
        ftp.retrbinary("RETR " + filename, fd.write)
        fd.close()
        logging.debug("FTP 파일 다운로드 성공(%s)"%filename)
    except:
        logging.debug("업데이트 서버 연결 실패")
        MessageWindow("업데이트 서버 연결 실패").showModal()

#프로그램 버전 체크
def checkVersion():
    config_version = configparser.ConfigParser()
    #---------------------------------------------------
    config_version.read('DAESUNG_VERSION.ini')
    current_ver = config_version['VERSION']['VERSION']
    #---------------------------------------------------
    downLoadFiles("DAESUNG_NEW_VERSION.ini")
    config_version.read('DAESUNG_NEW_VERSION.ini')
    new_ver = config_version['VERSION']['VERSION']
    #---------------------------------------------------
    if current_ver != new_ver: flag = 0
    elif current_ver == new_ver: flag = 1
    
    return flag
    
################################################################################################################
class ConnectDBThread(QThread):
    sig_data = pyqtSignal(int)
    
    def run(self):
        while True:
            try:
                test_db = pymysql.connect(host=host, port=port, user=user, password="doorerp1!", db = name)
                logging.debug('ConnectDBThread : test db 연결성공')
                test_db.close()
                logging.debug('ConnectDBThread : test db 닫기성공')
                self.sig_data.emit(1)
            except: self.sig_data.emit(0)
            time.sleep(1)

class FormatThread(QThread):
    sig_data = pyqtSignal(int)
    
    def __init__(self, flag):
        super().__init__()
        self.flag = flag
    
    def run(self):
        if self.flag == "focus":
            while True:
                self.sig_data.emit(2)
                time.sleep(1)
        elif self.flag == "format":
            time.sleep(3)
            self.sig_data.emit(1)
        elif self.flag == 'light':
            time.sleep(10)
            self.sig_data.emit(3)
        elif self.flag == 'dorna':
            time.sleep(3)
            self.sig_data.emit(4)
        elif self.flag == 'clear':
            time.sleep(1.5)
            self.sig_data.emit(5)
        elif self.flag == 'load':
            while True:
                time.sleep(20)
                self.sig_data.emit(6)

class SocketThread(QThread):
    sig_data = pyqtSignal(str)
    
    def __init__(self, ip, port):
        super().__init__()
        self.ip = ip
        self.port = port
        
    def run(self):
        while True:
            try:
                test_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                test_socket.settimeout(0.5)
                test_socket.connect((self.ip, self.port))
                self.sig_data.emit("success")
            except: self.sig_data.emit("failed")
            time.sleep(30)
            
class SerialThread(QThread):
    sig_data = pyqtSignal(str)
    
    def __init__(self, port, rate):
        super().__init__()
        self.port = port
        self.rate = rate
        
    def run(self):
        while True:
            try:
                test_ser = serial.Serial(self.port, self.rate, timeout = 0.5)
                test_ser.close()
                self.sig_data.emit("success")
            except: self.sig_data.emit("failed")
            time.sleep(30)

################################################################################################################
class LoginWindow(QDialog):
    def __init__(self):
        super(LoginWindow, self).__init__()
        loadUi("ui\DAESUNG_MES_Login.ui", self)
        
        DaesungFunctions.setStyle(self, "login", '', '') #기본 셋팅
        DaesungFunctions.clickLogin(self, self.pwd_input)
        
        global PRT_FLAG, PRT_NAME, SORT_ARRAY
        PRT_FLAG, PRT_NAME, SORT_ARRAY = '%', '전체', []
        
        self.connectDB() #DB연결
        
        self.pwd_input.setFocus()
        if pwd_data != '': self.pwd_input.setText(pwd_data)
        if pwd_flag == 't': self.pwd_check.setChecked(True)
        
        if LINE_FLAG == '1': self.one_radio.setChecked(True)
        else: self.two_radio.setChecked(True)
        
        self.btn_0.clicked.connect(lambda state, button = self.btn_0 : DaesungFunctions.NumClicked(self, state, button))
        self.btn_1.clicked.connect(lambda state, button = self.btn_1 : DaesungFunctions.NumClicked(self, state, button))
        self.btn_2.clicked.connect(lambda state, button = self.btn_2 : DaesungFunctions.NumClicked(self, state, button))
        self.btn_3.clicked.connect(lambda state, button = self.btn_3 : DaesungFunctions.NumClicked(self, state, button))
        self.btn_4.clicked.connect(lambda state, button = self.btn_4 : DaesungFunctions.NumClicked(self, state, button))
        self.btn_5.clicked.connect(lambda state, button = self.btn_5 : DaesungFunctions.NumClicked(self, state, button))
        self.btn_6.clicked.connect(lambda state, button = self.btn_6 : DaesungFunctions.NumClicked(self, state, button))
        self.btn_7.clicked.connect(lambda state, button = self.btn_7 : DaesungFunctions.NumClicked(self, state, button))
        self.btn_8.clicked.connect(lambda state, button = self.btn_8 : DaesungFunctions.NumClicked(self, state, button))
        self.btn_9.clicked.connect(lambda state, button = self.btn_9 : DaesungFunctions.NumClicked(self, state, button))
        self.backspace.clicked.connect(lambda: DaesungFunctions.NumDeleted(self))
        
        self.empl_combo.currentIndexChanged.connect(lambda : self.changeData(0))
        self.wc_combo.currentIndexChanged.connect(lambda : self.changeData(1))
        self.proc_combo.currentIndexChanged.connect(lambda : self.changeData(2))
        
        # --------------------------------------------------
        self.update_btn.clicked.connect(self.update)
        self.as_btn.clicked.connect(lambda: DaesungFunctions.openUrl(self))
        self.login_btn.clicked.connect(self.clickedLogin)
        self.exit_btn.clicked.connect(self.exitWindow)
    
    def connectDB(self):
        DB_config = configparser.ConfigParser()
        DB_config.read('DB_Manager.ini')
        global company, host, port, user, name
        company = DB_config['LOGIN']['COMPANY']
        host = DB_config[company]['HOST']
        port = int(DB_config[company]['PORT'])
        user = DB_config[company]['USER']
        name = DB_config[company]['DBNAME']
        #----------------------------------------------------------------------------
        result = DaesungQuery.connectDB(self, host, port, user, name)
        #----------------------------------------------------------------------------
        if result == "success": self.selectEmpl()
        else: self.connectDBThread()
    
    #작업자 조회
    def selectEmpl(self):
        #----------------------------------------------------------------------------
        self.E_rows = DaesungQuery.selectEmpl(self)
        #----------------------------------------------------------------------------
        if self.E_rows != () and self.E_rows != 'failed':
            for i in range(len(self.E_rows)):
                self.empl_combo.addItem(self.E_rows[i]['EMPL_NAME'])
                if EMPL_CODE == self.E_rows[i]['EMPL_CODE']:
                    self.empl_combo.setCurrentText(self.E_rows[i]['EMPL_NAME'])
            self.changeData(0)
        elif self.E_rows == 'failed': self.connectDBThread()
    
    #작업장 조회
    def selectWc(self):
        self.wc_combo.clear()
        #----------------------------------------------------------------------------
        self.W_rows = DaesungQuery.selectWc(self, EMPL_CODE)
        #----------------------------------------------------------------------------
        if self.W_rows != () and self.W_rows != 'failed':
            for i in range(len(self.W_rows)):
                self.wc_combo.addItem(self.W_rows[i]['WC_NAME'])
                if WC_CODE == self.W_rows[i]['WC_CODE']:
                    self.wc_combo.setCurrentText(self.W_rows[i]['WC_NAME'])
            self.changeData(1)
        elif self.W_rows == 'failed': self.connectDBThread()
    
    #공정 조회
    def selectProc(self):
        self.proc_combo.clear()
        #----------------------------------------------------------------------------
        self.P_rows = DaesungQuery.selectProc(self, EMPL_CODE, WC_CODE, '')
        #----------------------------------------------------------------------------
        if self.P_rows != () and self.P_rows != 'failed':
            for i in range(len(self.P_rows)):
                self.proc_combo.addItem(self.P_rows[i]['PROC_NAME'])
                if PROC_CODE == self.P_rows[i]['PROC_CODE']:
                    self.proc_combo.setCurrentText(self.P_rows[i]['PROC_NAME'])
            self.changeData(2)
        elif self.P_rows == 'failed': self.connectDBThread()
    
    def changeData(self, flag):
        if flag == 0 and self.E_rows != ():
            global EMPL_CODE, EMPL_NAME
            index = self.empl_combo.currentIndex()
            EMPL_CODE = self.E_rows[index]['EMPL_CODE']
            EMPL_NAME = self.E_rows[index]['EMPL_NAME']
            self.selectWc()
        elif flag == 1 and self.W_rows != ():
            global WC_CODE, WC_NAME
            index = self.wc_combo.currentIndex()
            WC_CODE = self.W_rows[index]['WC_CODE']
            WC_NAME = self.W_rows[index]['WC_NAME']
            if WC_CODE == '01' or WC_CODE == '09': self.line_radio.show()
            else: self.line_radio.hide()
            self.selectProc()
        elif flag == 2 and self.P_rows != ():
            global PROC_NAME, PROC_CODE
            index = self.proc_combo.currentIndex()
            PROC_NAME = self.P_rows[index]['PROC_NAME']
            PROC_CODE = self.P_rows[index]['PROC_CODE']
    
    def clickedLogin(self):
        pwd = self.pwd_input.text()
        if self.wc_combo.currentText() == "": MessageWindow(self, "작업장을 선택해주세요.").showModal()
        elif self.proc_combo.currentText() == "": MessageWindow(self, "생산공정을 선택해주세요.").showModal()
        elif pwd == "": MessageWindow(self, "비밀번호를 입력해주세요.").showModal()
        else:
            #----------------------------------------------------------------------------
            PWD_rows = DaesungQuery.checkPassword(self, EMPL_CODE, pwd)
            #----------------------------------------------------------------------------
            if PWD_rows != ():
                global RECONNECT_FLAG
                RECONNECT_FLAG = 0
                #------------------------------------------------------------------------
                self.saveLoginData()
                #------------------------------------------------------------------------
                if PROC_CODE == '0110' or PROC_CODE == '0115': main_win = MesWindow(TodayData) #접착, 테노너
                elif PROC_CODE == '0117': main_win = MesEdgeWindow(TodayData) #엣지
                elif PROC_CODE == '0120': main_win = MesPackWindow(TodayData, 0) #포장검수
                elif PROC_CODE == '0903': main_win = MesBogangWindow(TodayData) #보강재 부착
                elif WC_CODE == '04': main_win = MesMoldingLotWindow(TodayData) #몰딩부
                elif WC_CODE == '05' or WC_CODE == '08' or WC_CODE == '09' or WC_CODE == '16': main_win = MesFrameWindow(TodayData) #문틀부
                else: main_win = MesLotWindow(TodayData)
                widget.setWindowTitle("DAESUNG MES ♣ [{0}] {1} > {2} [{3}라인]".format(EMPL_NAME, WC_NAME, PROC_NAME, LINE_FLAG))
                widget.addWidget(main_win)
                widget.setCurrentIndex(widget.currentIndex() + 1)
                self.deleteLater()
            elif PWD_rows == 'failed': self.connectDBThread()
            else: MessageWindow(self, "비밀번호가 올바르지 않습니다.").showModal()
    
    def saveLoginData(self):
        try:
            global pwd_flag, pwd_data, LINE_FLAG
            if self.pwd_check.isChecked() == True: pwd_flag, pwd_data = 't', self.pwd_input.text()
            else: pwd_flag = 'f'
            if self.one_radio.isChecked() == True: LINE_FLAG = '1'
            else: LINE_FLAG = '2'
            #----------------------------------------------------------------------------
            config['PROCCODE']['EMPL'] = EMPL_CODE
            config['PROCCODE']['WC'] = WC_CODE
            config['PROCCODE']['PROC'] = PROC_CODE
            config['PROCCODE']['LINE'] = LINE_FLAG
            config['PROCCODE']['PWD/FLAG'] = pwd_flag
            config['PROCCODE']['PWD/DATA'] = pwd_data
            with open('DAESUNG_ADDRESS.ini', 'w') as configfile: config.write(configfile) # save
            logging.debug("saveLoginData : 데이터 저장 성공")
        except: logging.debug("saveLoginData : 데이터 저장 실패")
    
    def update(self):
        try:
            flag = checkVersion()
            widget.hide()
            UpgradeWindow(1, flag).showModal()
        except: logging.debug("upgrade : 연결 실패")
    
    #실시간 DB연결
    def connectDBThread(self):
        if self.DB_flag == 0:
            self.DB_flag = 1
            logging.debug("connectDBThread : DB연결 실패")
            MessageWindow(self, "DB연결 실패").showModal()
            #----------------------------------------------------------------------------
            self.DB_th = ConnectDBThread()
            self.DB_th.sig_data.connect(self.DbThreadSlot)
            self.DB_th.start()
    
    @pyqtSlot(int)       
    def DbThreadSlot(self, con):
        if con == 1:
            self.DB_th.terminate()
            self.DB_flag = 0
            self.connectDB() #DB연결
    
    def exitWindow(self):
        self.saveLoginData()
        os._exit(1)

################################################################################################################
#Settings 화면
class SetWindow(QDialog):
    def __init__(self):
        super(SetWindow, self).__init__()
        loadUi("ui\Settings.ui", self)
        
        DaesungFunctions.setStyle(self, "set", WC_CODE, PROC_CODE)
        
        self.setData()
        self.connectLight()
        
        DaesungFunctions.clickable(self, self.scanner_ip_input, self.scanner_ip_input)
        DaesungFunctions.clickable(self, self.scanner_port_input, self.scanner_port_input)
        DaesungFunctions.clickable(self, self.scanner_ip_input2, self.scanner_ip_input2)
        DaesungFunctions.clickable(self, self.scanner_port_input2, self.scanner_port_input2)
        #----------------------------------------------------------------------------
        DaesungFunctions.clickable(self, self.plc_ip_input, self.plc_ip_input)
        DaesungFunctions.clickable(self, self.plc_port_input, self.plc_port_input)
        DaesungFunctions.clickable(self, self.plc_lenx_input, self.plc_lenx_input)
        DaesungFunctions.clickable(self, self.plc_widx_input, self.plc_widx_input)
        DaesungFunctions.clickable(self, self.plc_tikx_input, self.plc_tikx_input)
        DaesungFunctions.clickable(self, self.plc_edgeC_input, self.plc_edgeC_input)
        DaesungFunctions.clickable(self, self.plc_edge1_input, self.plc_edge1_input)
        DaesungFunctions.clickable(self, self.plc_edge2_input, self.plc_edge2_input)
        DaesungFunctions.clickable(self, self.plc_holeP_input, self.plc_holeP_input)
        DaesungFunctions.clickable(self, self.plc_holeF_input, self.plc_holeF_input)
        DaesungFunctions.clickable(self, self.plc_comp_input, self.plc_comp_input)
        DaesungFunctions.clickable(self, self.plc_no_input, self.plc_no_input)
        DaesungFunctions.clickable(self, self.plc_edgeR1_input, self.plc_edgeR1_input)
        DaesungFunctions.clickable(self, self.plc_edgeR2_input, self.plc_edgeR2_input)
        DaesungFunctions.clickable(self, self.plc_pack_input, self.plc_pack_input)
        DaesungFunctions.clickable(self, self.plc_frame_ck_input, self.plc_frame_ck_input)
        DaesungFunctions.clickable(self, self.plc_frame_lotid_input, self.plc_frame_lotid_input)
        DaesungFunctions.clickable(self, self.plc_frame_lenx_input, self.plc_frame_lenx_input)
        DaesungFunctions.clickable(self, self.plc_frame_widx_input, self.plc_frame_widx_input)
        #----------------------------------------------------------------------------
        DaesungFunctions.clickable(self, self.printer_ip_input, self.printer_ip_input)
        DaesungFunctions.clickable(self, self.printer_port_input, self.printer_port_input)
        DaesungFunctions.clickable(self, self.printer_mode_input, self.printer_mode_input)
        DaesungFunctions.clickable(self, self.printer2_ip_input, self.printer2_ip_input)
        DaesungFunctions.clickable(self, self.printer2_port_input, self.printer2_port_input)
        DaesungFunctions.clickable(self, self.printer2_mode_input, self.printer2_mode_input)
        DaesungFunctions.clickable(self, self.printer3_ip_input, self.printer3_ip_input)
        DaesungFunctions.clickable(self, self.printer3_port_input, self.printer3_port_input)
        DaesungFunctions.clickable(self, self.printer3_mode_input, self.printer3_mode_input)
        #----------------------------------------------------------------------------
        DaesungFunctions.clickable(self, self.sensor_port_input, self.sensor_port_input)
        DaesungFunctions.clickable(self, self.sensor_rate_input, self.sensor_rate_input)
        #----------------------------------------------------------------------------
        DaesungFunctions.clickable(self, self.light_port_input, self.light_port_input)
        DaesungFunctions.clickable(self, self.light_rate_input, self.light_rate_input)
        
        self.one.clicked.connect(lambda state, button = self.one : DaesungFunctions.NumClicked(self, state, button))
        self.two.clicked.connect(lambda state, button = self.two : DaesungFunctions.NumClicked(self, state, button))
        self.three.clicked.connect(lambda state, button = self.three : DaesungFunctions.NumClicked(self, state, button))
        self.four.clicked.connect(lambda state, button = self.four : DaesungFunctions.NumClicked(self, state, button))
        self.five.clicked.connect(lambda state, button = self.five : DaesungFunctions.NumClicked(self, state, button))
        self.six.clicked.connect(lambda state, button = self.six : DaesungFunctions.NumClicked(self, state, button))
        self.seven.clicked.connect(lambda state, button = self.seven : DaesungFunctions.NumClicked(self, state, button))
        self.eight.clicked.connect(lambda state, button = self.eight : DaesungFunctions.NumClicked(self, state, button))
        self.nine.clicked.connect(lambda state, button = self.nine : DaesungFunctions.NumClicked(self, state, button))
        self.zero.clicked.connect(lambda state, button = self.zero : DaesungFunctions.NumClicked(self, state, button))
        self.dot.clicked.connect(lambda state, button = self.dot : DaesungFunctions.NumClicked(self, state, button))
        self.backspace.clicked.connect(lambda: DaesungFunctions.NumDeleted(self))
        
        self.save_btn.clicked.connect(self.saveData)
        self.cancel_btn.clicked.connect(lambda: DaesungFunctions.closeWindow(self))
        
    def setData(self):
        try:
            self.scanner_ip_input.setText(config['SETTINGS']['SCANNER/IP'])
            self.scanner_port_input.setText(config['SETTINGS']['SCANNER/PORT'])
            self.scanner_ip_input2.setText(config['SETTINGS']['SCANNER2/IP'])
            self.scanner_port_input2.setText(config['SETTINGS']['SCANNER2/PORT'])
            #----------------------------------------------------------------------------
            self.plc_ip_input.setText(config['SETTINGS']['PLC/IP'])
            self.plc_port_input.setText(config['SETTINGS']['PLC/PORT'])
            self.plc_lenx_input.setText(config['SETTINGS']['PLC/LENX'])
            self.plc_widx_input.setText(config['SETTINGS']['PLC/WIDX'])
            self.plc_tikx_input.setText(config['SETTINGS']['PLC/TIKX'])
            self.plc_edgeC_input.setText(config['SETTINGS']['PLC/EDGE'])
            self.plc_edge1_input.setText(config['SETTINGS']['PLC/EDGE1'])
            self.plc_edge2_input.setText(config['SETTINGS']['PLC/EDGE2'])
            self.plc_holeP_input.setText(config['SETTINGS']['PLC/HOLE/POS'])
            self.plc_holeF_input.setText(config['SETTINGS']['PLC/HOLE/FLAG'])
            self.plc_comp_input.setText(config['SETTINGS']['PLC/COMP'])
            self.plc_no_input.setText(config['SETTINGS']['PLC/NO'])
            self.plc_edgeR1_input.setText(config['SETTINGS']['PLC/EDGE1/R'])
            self.plc_edgeR2_input.setText(config['SETTINGS']['PLC/EDGE2/R'])
            self.plc_pack_input.setText(config['SETTINGS']['PLC/PACK'])
            try: 
                self.plc_frame_ck_input.setText(config['SETTINGS']['PLC/CK'])
                self.plc_frame_lotid_input.setText(config['SETTINGS']['PLC/LOTID'])
                self.plc_frame_lenx_input.setText(config['SETTINGS']['PLC/BLENX'])
                self.plc_frame_widx_input.setText(config['SETTINGS']['PLC/BWIDX'])
            except: pass
            #----------------------------------------------------------------------------
            self.printer_ip_input.setText(config['SETTINGS']['PRINTER/IP'])
            self.printer_port_input.setText(config['SETTINGS']['PRINTER/PORT'])
            try:
                self.printer_mode_input.setText(config['SETTINGS']['PRINTER/MODE'])
                self.printer2_ip_input.setText(config['SETTINGS']['PRINTER2/IP'])
                self.printer2_port_input.setText(config['SETTINGS']['PRINTER2/PORT'])
                self.printer2_mode_input.setText(config['SETTINGS']['PRINTER2/MODE'])
                self.printer3_ip_input.setText(config['SETTINGS']['PRINTER3/IP'])
                self.printer3_port_input.setText(config['SETTINGS']['PRINTER3/PORT'])
                self.printer3_mode_input.setText(config['SETTINGS']['PRINTER3/MODE'])
            except: pass
            #----------------------------------------------------------------------------
            self.sensor_port_input.setText(config['SETTINGS']['SENSOR/PORT'])
            self.sensor_rate_input.setText(config['SETTINGS']['SENSOR/RATE'])
            #----------------------------------------------------------------------------
            self.light_port_input.setText(config['SETTINGS']['LIGHT/PORT'])
            self.light_rate_input.setText(config['SETTINGS']['LIGHT/RATE'])
            try:
                if config['HWFLAG']['SCANNER'] == "true": self.scanner_check.setChecked(True)
                if config['HWFLAG']['SCANNER2'] == "true": self.scanner_check2.setChecked(True)
                #----------------------------------------------------------------------------
                if config['HWFLAG']['PLC'] == "true": self.plc_check.setChecked(True)
                #----------------------------------------------------------------------------
                if config['HWFLAG']['PRINTER'] == "true": self.printer_check.setChecked(True)
                if config['HWFLAG']['MODE'] == "true": self.printer_mode_check.setChecked(True)
                if config['HWFLAG']['PO'] == "true": self.printer_po_check.setChecked(True)
                try:
                    if config['HWFLAG']['PRINTER2'] == "true": self.printer2_check.setChecked(True)
                    if config['HWFLAG']['MODE2'] == "true": self.printer2_mode_check.setChecked(True)
                    if config['HWFLAG']['PO2'] == "true": self.printer2_po_check.setChecked(True)
                    if config['HWFLAG']['PRINTER3'] == "true": self.printer3_check.setChecked(True)
                    if config['HWFLAG']['MODE3'] == "true": self.printer3_mode_check.setChecked(True)
                    if config['HWFLAG']['PO3'] == "true": self.printer3_po_check.setChecked(True)
                except: pass
                #----------------------------------------------------------------------------
                if config['HWFLAG']['SENSOR'] == "true": self.sensor_check.setChecked(True)
                if config['HWFLAG']['LIGHT'] == "true": self.light_check.setChecked(True)
            except: logging.debug("setData : H/W체크 가져오기 실패")
        except: logging.debug("setData : H/W정보 가져오기 실패")
    
    def saveData(self):
        try:
            config['SETTINGS']['SCANNER/IP'] = self.scanner_ip_input.text()
            config['SETTINGS']['SCANNER/PORT'] = self.scanner_port_input.text()
            config['SETTINGS']['SCANNER2/IP'] = self.scanner_ip_input2.text()
            config['SETTINGS']['SCANNER2/PORT'] = self.scanner_port_input2.text()
            #----------------------------------------------------------------------------
            config['SETTINGS']['PLC/IP'] = self.plc_ip_input.text()
            config['SETTINGS']['PLC/PORT'] = self.plc_port_input.text()
            config['SETTINGS']['PLC/LENX'] = self.plc_lenx_input.text()
            config['SETTINGS']['PLC/WIDX'] = self.plc_widx_input.text()
            config['SETTINGS']['PLC/TIKX'] = self.plc_tikx_input.text()
            config['SETTINGS']['PLC/EDGE'] = self.plc_edgeC_input.text()
            config['SETTINGS']['PLC/EDGE1'] = self.plc_edge1_input.text()
            config['SETTINGS']['PLC/EDGE2'] = self.plc_edge2_input.text()
            config['SETTINGS']['PLC/HOLE/POS'] = self.plc_holeP_input.text()
            config['SETTINGS']['PLC/HOLE/FLAG'] = self.plc_holeF_input.text()
            config['SETTINGS']['PLC/COMP'] = self.plc_comp_input.text()
            config['SETTINGS']['PLC/NO'] = self.plc_no_input.text()
            config['SETTINGS']['PLC/EDGE1/R'] = self.plc_edgeR1_input.text()
            config['SETTINGS']['PLC/EDGE2/R'] = self.plc_edgeR2_input.text()
            config['SETTINGS']['PLC/PACK'] = self.plc_pack_input.text()
            config['SETTINGS']['PLC/CK'] = self.plc_frame_ck_input.text()
            config['SETTINGS']['PLC/LOTID'] = self.plc_frame_lotid_input.text()
            config['SETTINGS']['PLC/BLENX'] = self.plc_frame_lenx_input.text()
            config['SETTINGS']['PLC/BWIDX'] = self.plc_frame_widx_input.text()
            #----------------------------------------------------------------------------
            config['SETTINGS']['PRINTER/IP'] = self.printer_ip_input.text()
            config['SETTINGS']['PRINTER/PORT'] = self.printer_port_input.text()
            config['SETTINGS']['PRINTER/MODE'] = self.printer_mode_input.text()
            config['SETTINGS']['PRINTER2/IP'] = self.printer2_ip_input.text()
            config['SETTINGS']['PRINTER2/PORT'] = self.printer2_port_input.text()
            config['SETTINGS']['PRINTER2/MODE'] = self.printer2_mode_input.text()
            config['SETTINGS']['PRINTER3/IP'] = self.printer3_ip_input.text()
            config['SETTINGS']['PRINTER3/PORT'] = self.printer3_port_input.text()
            config['SETTINGS']['PRINTER3/MODE'] = self.printer3_mode_input.text()
            #----------------------------------------------------------------------------
            config['SETTINGS']['SENSOR/PORT'] = self.sensor_port_input.text()
            config['SETTINGS']['SENSOR/RATE'] = self.sensor_rate_input.text()
            #----------------------------------------------------------------------------
            if config['HWFLAG']['LIGHT'] == 'true' and self.light_check.isChecked() == False:
                if self.light_flag == 'success':
                    try:
                        light_ser.write('RY 1 0\r'.encode()) #green light
                        logging.debug("saveData : RY 1 0 성공")
                        light_ser.close()
                    except: logging.debug("saveData : RY 1 0 실패")
            config['SETTINGS']['LIGHT/PORT'] = self.light_port_input.text()
            config['SETTINGS']['LIGHT/RATE'] = self.light_rate_input.text()
            if self.light_check.isChecked() == True: config['HWFLAG']['LIGHT'] = "true"
            else: config['HWFLAG']['LIGHT'] = "false"
            #----------------------------------------------------------------------------
            if self.scanner_check.isChecked() == True: config['HWFLAG']['SCANNER'] = "true"
            else: config['HWFLAG']['SCANNER'] = "false"
            if self.scanner_check2.isChecked() == True: config['HWFLAG']['SCANNER2'] = "true"
            else: config['HWFLAG']['SCANNER2'] = "false"
            #----------------------------------------------------------------------------
            if self.plc_check.isChecked() == True: config['HWFLAG']['PLC'] = "true"
            else: config['HWFLAG']['PLC'] = "false"
            #----------------------------------------------------------------------------
            if self.printer_check.isChecked() == True: config['HWFLAG']['PRINTER'] = "true"
            else: config['HWFLAG']['PRINTER'] = "false"
            if self.printer_mode_check.isChecked() == True: config['HWFLAG']['MODE'] = "true"
            else: config['HWFLAG']['MODE'] = "false"
            if self.printer_po_check.isChecked() == True: config['HWFLAG']['PO'] = "true"
            else: config['HWFLAG']['PO'] = "false"
            #----------------------------------------------------------------------------
            if self.printer2_check.isChecked() == True: config['HWFLAG']['PRINTER2'] = "true"
            else: config['HWFLAG']['PRINTER2'] = "false"
            if self.printer2_mode_check.isChecked() == True: config['HWFLAG']['MODE2'] = "true"
            else: config['HWFLAG']['MODE2'] = "false"
            if self.printer2_po_check.isChecked() == True: config['HWFLAG']['PO2'] = "true"
            else: config['HWFLAG']['PO2'] = "false"
            #----------------------------------------------------------------------------
            if self.printer3_check.isChecked() == True: config['HWFLAG']['PRINTER3'] = "true"
            else: config['HWFLAG']['PRINTER3'] = "false"
            if self.printer3_mode_check.isChecked() == True: config['HWFLAG']['MODE3'] = "true"
            else: config['HWFLAG']['MODE3'] = "false"
            if self.printer3_po_check.isChecked() == True: config['HWFLAG']['PO3'] = "true"
            else: config['HWFLAG']['PO3'] = "false"
            #----------------------------------------------------------------------------
            if self.sensor_check.isChecked() == True: config['HWFLAG']['SENSOR'] = "true"
            else: config['HWFLAG']['SENSOR'] = "false"
            #----------------------------------------------------------------------------
            with open('DAESUNG_ADDRESS.ini', 'w') as configfile: config.write(configfile) # save
            l_text, m_text = "saveData : 데이터 저장 성공", "데이터가 저장되었습니다."
        except: l_text, m_text = "saveData : 데이터 저장 실패", "데이터를 저장할 수 없습니다."
        logging.debug(l_text)
        self.close()
        MessageWindow(self, m_text).showModal()
    
    def connectLight(self):
        global light_ser
        #--------------------------------------------------
        try: light_ser.close()
        except: pass
        #--------------------------------------------------
        if self.light_check.isChecked():
            light_port = self.light_port_input.text()
            light_rate = int(self.light_rate_input.text())
            try:
                light_ser = serial.Serial(light_port, light_rate, timeout = 0.5)
                self.light_flag = "success"
            except: self.light_flag = "failed"
        else: self.light_flag = "unable"
    
    def showModal(self):
        return super().exec_()

################################################################################################################
#실시간 DB연결 및 rowCount 확인
class lotCountThread(QThread):
    sig_data = pyqtSignal(int, str)
    
    def __init__(self, s_date, proc_data, appr_flag, w_data):
        super().__init__()
        self.s_date, self.proc_data, self.appr_flag, self.w_data = s_date, proc_data, appr_flag, w_data

    def run(self):
        while True:
            try:
                count = DaesungQuery.selectLotCount(self, self.proc_data, self.s_date, PRT_FLAG, self.appr_flag, self.w_data)
                self.sig_data.emit(len(count), count[0]['JAKUP_APPR_TIME'])
            except: self.sig_data.emit(100000, 'f')
            time.sleep(5)

#실시간 DB연결 및 rowCount 확인
class lotLoadThread(QThread):
    sig_data = pyqtSignal(int)
    
    def __init__(self, time_delay):
        super().__init__()
        self.time_delay = time_delay

    def run(self):
        while True:
            time.sleep(self.time_delay)
            self.sig_data.emit(1)

#---------------------------------------------------------------------------------------------------------------
#LOT Window
class MesLotWindow(QDialog):
    def __init__(self, date):
        super(MesLotWindow, self).__init__()
        loadUi("ui\DAESUNG_MES_L.ui", self)
        
        #콤보박스 공정 조회
        if WC_CODE == '02':
            PROC = DaesungQuery.selectProc(self, EMPL_CODE, WC_CODE, '')
            PROC.insert(0, {'PROC_CODE': '0117', 'PROC_NAME': '엣지', 'EMPL_CODE': '%s'%EMPL_CODE})
            PROC.insert(0, {'PROC_CODE': '0103', 'PROC_NAME': '판재재단', 'EMPL_CODE': '%s'%EMPL_CODE})
            PROC.insert(0, {'PROC_CODE': '0101', 'PROC_NAME': '심재재단조립', 'EMPL_CODE': '%s'%EMPL_CODE})
            self.tableWidget.setSortingEnabled(True)
            self.tableWidget.setSelectionMode(QAbstractItemView.ExtendedSelection)
        elif PROC_CODE == '0101' or PROC_CODE == '0102' or PROC_CODE == '0103':
            PROC = DaesungQuery.selectProc(self, EMPL_CODE, WC_CODE, "AND PROC.PROC_CODE IN ('0101', '0102', '0103', '0120')")
        else: PROC = ''
        
        DaesungFunctions.setComboStyle(self, date, WC_CODE, PROC_CODE, PROC_NAME, PROC) #기본 셋팅
        self.flag_combo.setCurrentText(PRT_NAME)
        self.sort_num, self.key_flag = [], 0
        self.JAKUP_APPR_FLAG, self.PROC_CODE, self.W_DATA = '2', "MPJAKUP.PROC_CODE = '{0}'".format(PROC_CODE), ''
        self.jackup_set_btn.hide()
        
        self.hwConnect() #HW연결
        DaesungFunctions.replaceDate(self) #DB로드
        
        self.aproc_combo.currentIndexChanged.connect(self.procChanged) #공정 변경
        self.flag_combo.currentIndexChanged.connect(self.flagChanged) #상태값 변경
        
        self.print_btn.clicked.connect(self.printLabel) #라벨 발행
        self.jackup_btn.clicked.connect(lambda: self.jackupPrint(1)) #작업지시서 인쇄
        self.jackup_set_btn.clicked.connect(lambda: self.jackupPrint(2)) #세트 인쇄
        self.reload_btn.clicked.connect(self.DBload) #DB로드
        self.tableWidget.horizontalHeader().sectionClicked.connect(self.selectedAll)
        self.tableWidget.clicked.connect(self.click_row)
        
        keyboard.on_press_key("ctrl", lambda _:self.clickedCtrl(1))
        keyboard.on_release_key("ctrl", lambda _:self.clickedCtrl(0))
        
        # 화살표버튼 --------------------------------------------------
        self.top.clicked.connect(lambda: DaesungFunctions.topData(self))
        self.prev.clicked.connect(lambda: DaesungFunctions.prevData(self))
        self.next.clicked.connect(lambda: DaesungFunctions.nextData(self))
        self.bottom.clicked.connect(lambda: DaesungFunctions.bottomData(self))
        
        # 상단버튼 --------------------------------------------------
        self.DB_btn.clicked.connect(self.resetLight)
        self.date_btn.clicked.connect(lambda: DaesungFunctions.calendar(self))
        self.set_btn.clicked.connect(self.setData)
        self.log_btn.clicked.connect(self.logWindow)
        self.as_btn.clicked.connect(lambda: DaesungFunctions.openUrl(self))
        self.exit_btn.clicked.connect(self.exitWindow)
    
    def hwConnect(self):
        try:
            self.set_win = SetWindow()
            self.connectLight() #경광등 연결
            if WC_CODE == '02': self.connectPrint() #바코드 프린터 연결
        except: logging.debug("hwConnect : 실패")
    
    def connectLight(self):
        global light_ser
        #--------------------------------------------------
        try: light_ser.close()
        except: pass
        #--------------------------------------------------
        light_port = self.set_win.light_port_input.text()
        if self.set_win.light_check.isChecked():
            light_rate = int(self.set_win.light_rate_input.text())
            try:
                light_ser = serial.Serial(light_port, light_rate, timeout = 0.5)
                self.light_flag = "success"
                logging.debug("connectLight : 경광등(%s) 연결 성공"%light_port)
            except:
                self.light_flag = "failed"
                logging.debug("connectLight : 경광등(%s) 연결 실패"%light_port)
        else:
            self.light_flag = "unable"
            logging.debug("connectLight : 경광등(%s) 비활성"%light_port)
    
    def connectPrint(self):
        try:
            flag, self.count = 1, 0
            if (PROC_CODE == C_PROC or PROC_CODE == CS_PROC) and self.set_win.printer2_check.isChecked(): 
                self.ip = self.set_win.printer2_ip_input.text()
                self.port = int(self.set_win.printer2_port_input.text())
                self.mode_edit = self.set_win.printer2_mode_check
                self.mode = '^' + self.set_win.printer2_mode_input.text().replace(' ', '')
                try: self.count = int(self.set_win.printer2_count_input.text())
                except: pass
            elif (PROC_CODE == Z_PROC or PROC_CODE == ZS_PROC) and self.set_win.printer3_check.isChecked(): 
                self.ip = self.set_win.printer3_ip_input.text()
                self.port = int(self.set_win.printer3_port_input.text())
                self.mode_edit = self.set_win.printer3_mode_check
                self.mode = '^' + self.set_win.printer3_mode_input.text().replace(' ', '')
                try: self.count = int(self.set_win.printer3_count_input.text())
                except: pass
            elif (PROC_CODE == S_PROC or PROC_CODE == SS_PROC) and self.set_win.printer_check.isChecked():
                self.ip = self.set_win.printer_ip_input.text()
                self.port = int(self.set_win.printer_port_input.text())
                self.mode_edit = self.set_win.printer_mode_check
                self.mode = '^' + self.set_win.printer_mode_input.text().replace(' ', '')
                try: self.count = int(self.set_win.printer_count_input.text())
                except: pass
            else:
                self.printer_flag, flag = "unable", 0
                logging.debug("connectPrint : 프린트 비활성")
                self.print_status.setStyleSheet("background-color: #CDCDCD;") #gray
            #---------------------------------------------------------------------    
            if flag == 1:
                print_socket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
                print_socket.settimeout(0.5)
                print_socket.connect((self.ip, self.port))
                print_socket.close()
                #---------------------------------------------------------------------
                self.printer_flag = "success"
                logging.debug("connectPrint : 프린트 연결 성공")
                self.print_status.setStyleSheet("background-color: #55cba7;") #green
        except:
            self.printer_flag = "failed"
            logging.debug("connectPrint : 프린트 연결 실패")
            self.print_status.setStyleSheet("background-color: #fd97a5;") #red
    
    #---------------------------------------------------------------------------------------------------
    def setCheckFlag(self):
        try:
            self.set_check = []
            for count, checkbox in enumerate(self.checkBoxList):
                if checkbox.isChecked() == True: self.set_check.append(count)
        except: pass
    
    #미출력 LOT 알림
    def printAlarm(self, index):
        try:
            global PRT_INDEX
            if index != PRT_INDEX:
                PRT_INDEX = index
                if self.light_flag == 'success':
                    try:
                        light_ser.write('RY 1 1\r'.encode()) #red light
                        logging.debug("printAlarm : RY 1 1 성공")
                    except: logging.debug("printAlarm : RY 1 1 실패")
                    MessageWindow(self, '미출력 로트 : [ %s ]'%self.tableWidget.item(index, 5).text()).showModal()
                    try:
                        light_ser.write('RY 1 0\r'.encode()) #green light
                        logging.debug("printAlarm : RY 1 0 성공")
                    except: logging.debug("printAlarm : RY 1 0 실패")
                else: MessageWindow(self, '미출력 로트 : [ %s ]'%self.tableWidget.item(index, 5).text()).showModal()
        except: pass
    
    def DBload(self):
        self.setCheckFlag()
        self.tableWidget.setSortingEnabled(False)
        self.tableWidget.setRowCount(0)
        try:
            global DATA_COUNT, JAKUP_TIME, PRT_INDEX
            DATA_COUNT, JAKUP_TIME = 0, 0
            #----------------------------------------------------------------------------
            S_rows = DaesungQuery.selectLotList(self, self.PROC_CODE, self.s_date, PRT_FLAG, self.JAKUP_APPR_FLAG, 'MJAKUP.REG_NO', self.W_DATA)
            #----------------------------------------------------------------------------
            if S_rows == 'failed': self.connectDBThread()
            elif S_rows == ():
                self.tableWidget.setRowCount(1)
                self.tableWidget.setRowHeight(0, 85)
                self.tableWidget.setSpan(0, 0, 1, 11)
                item_data = QTableWidgetItem("해당일자의 데이터가 없습니다.")
                item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                self.tableWidget.setItem(0, 0, item_data)
            else:
                self.checkBoxList, qty, prt_index = [], 0, -1
                self.tableWidget.setRowCount(len(S_rows))
                for i in range(len(S_rows)):
                    self.tableWidget.setRowHeight(i, 85)
                    ckbox = QCheckBox()
                    ckbox.setStyleSheet(t_checkStyle)
                    self.checkBoxList.append(ckbox)
                    if JAKUP_TIME < int(S_rows[i]['JAKUP_APPR_TIME']): JAKUP_TIME = int(S_rows[i]['JAKUP_APPR_TIME'])
                    #----------------------------------------------------------------------------
                    for count, j in enumerate(['', '', '', 'CHECK', 'JAKUP_FLAG', 'LOT_NUMB', 'ITEM_TEXT', 'QTY', 'LK_PRT_FLAG', 'REG_NO', 'MES_PRT_FLAG']):
                        if j == 'CHECK':
                            cellWidget = QWidget()
                            layoutCB = QHBoxLayout(cellWidget)
                            layoutCB.addWidget(self.checkBoxList[i])
                            layoutCB.setAlignment(QtCore.Qt.AlignCenter)
                            layoutCB.setContentsMargins(0, 0, 0, 0)
                            cellWidget.setLayout(layoutCB)
                            self.tableWidget.setCellWidget(i, count, cellWidget)
                            if i in self.set_check: self.checkBoxList[i].setChecked(True)
                        elif j == '': self.tableWidget.setItem(i, count, QTableWidgetItem(''))
                        else:
                            print_data = S_rows[i][j]
                            if j == 'LK_PRT_FLAG':
                                states = QPushButton()
                                if print_data == '0' or print_data == '00': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/대기.png);")
                                elif print_data == '2' or print_data == '01': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/진행.png);")
                                elif print_data == '1' or print_data == '02': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/완료.png);")
                                self.state_group.addButton(states, i)
                                self.tableWidget.setCellWidget(i, count, states)
                            elif j == 'QTY':
                                qty = qty + int(print_data)
                                item_data = QTableWidgetItem()
                                item_data.setData(Qt.DisplayRole, int(print_data))
                                item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                                self.tableWidget.setItem(i, count, item_data)
                            else:
                                if j == 'MES_PRT_FLAG' and (print_data == None or print_data != '1'):
                                    print_data = ''
                                    if prt_index == -1: prt_index = i
                                elif print_data == None: print_data = ''
                                elif j == 'JAKUP_FLAG' and print_data == '1': print_data = "시판"
                                elif j == 'JAKUP_FLAG' and print_data == '2': print_data = "LX"
                                item_data = QTableWidgetItem(str(print_data))
                                if j == 'ITEM_TEXT': item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                                else: item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                                self.tableWidget.setItem(i, count, item_data)
                DATA_COUNT = len(S_rows)
                self.qty_label.setText(str(qty))
                if WC_CODE == '01':
                    if prt_index > -1: self.printAlarm(prt_index) #미출력 LOT 알림
                    else: PRT_INDEX = -1
                if self.reload_num == 0:
                    try:
                        self.th_rowCount = lotCountThread(self.s_date, self.PROC_CODE, self.JAKUP_APPR_FLAG, self.W_DATA)
                        self.th_rowCount.sig_data.connect(self.newData)
                        self.th_rowCount.start()
                        self.reload_num = 1
                    except: logging.debug("DBload : th_rowCount 실패")
            DaesungFunctions.tableWidth(self, 'LOT', WC_CODE, len(S_rows))
            if WC_CODE == '02':
                self.tableWidget.setSortingEnabled(True)
                c_sort, c_count = -1, 0
                if SORT_ARRAY != []:
                    for s in SORT_ARRAY:
                        if c_sort != s:
                            self.tableWidget.sortByColumn(s, Qt.AscendingOrder)
                            c_count = 0
                        else:
                            if c_count%2 == 0:
                                self.tableWidget.sortByColumn(s, Qt.DescendingOrder)
                                c_count = 1
                            else:
                                self.tableWidget.sortByColumn(s, Qt.AscendingOrder)
                                c_count = 0
                        c_sort = s
        except: logging.debug("DBload : DB로드 실패")
    
    #라벨 발행
    def printLabel(self):
        checkArray, t_count, l_count = [], 0, 1
        if self.printer_flag == "success":
            for count, checkbox in enumerate(self.checkBoxList):
                if checkbox.isChecked() == True:
                    t_count = t_count + int(self.tableWidget.item(count, 7).text())
                    checkArray.append(count)
            if checkArray != []:
                try:
                    self.mysocket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
                    self.mysocket.settimeout(0.5)
                    self.mysocket.connect((self.ip, self.port))
                    self.print_status.setStyleSheet("background-color: #55cba7;") #green
                    try: self.th_rowCount.terminate()
                    except: pass
                    #----------------------------------------------------------------------------
                    if PROC_CODE == '0103' or PROC_CODE == '0202': self.fileName, hole_flag = cs_label, 0
                    elif PROC_CODE == '0117' or PROC_CODE == '0206': self.fileName, hole_flag = zs_label, 1
                    elif WC_CODE == '19': self.fileName, hole_flag = if_label, 0
                    else: self.fileName, hole_flag = ss_label, 0
                    try:
                        for i in checkArray:
                            REG_NO, C_REG_SEQ = self.tableWidget.item(i, 9).text(), ''
                            #----------------------------------------------------------------------------
                            P_rows = DaesungQuery.selectDetailList(self, REG_NO, '%', '%', self.s_date, PROC_CODE, '')
                            #----------------------------------------------------------------------------
                            if P_rows == 'failed': self.connectDBThread()
                            elif P_rows != ():
                                for j in range(len(P_rows)):
                                    try:
                                        f_name = open(self.fileName, 'r', encoding = 'utf-8')
                                        textData = f_name.read()
                                        f_name.close()
                                        if self.mode == '^MMC' and self.mode_edit.isChecked() == True:
                                            if l_count == t_count or l_count == self.count:
                                                textData = textData.replace("^CI28^MMT", "^CI28^MMC")
                                                t_count = t_count - l_count
                                                l_count = 0
                                        elif self.mode != '^MMC' and self.mode_edit.isChecked() == True: textData = textData.replace("^CI28", "^CI28" + self.mode)
                                        if self.set_win.printer_po_check.isChecked() == True: textData = textData.replace("^LS0", "^LS0^POI")
                                        elif self.set_win.printer_po_check.isChecked() == False: textData = textData.replace("^LS0", "^LS0^PON")
                                        try:
                                            REG_SEQ = P_rows[j]['REG_SEQ']
                                            for t in ['REG_NO', 'LOT_NUMB', 'REG_SEQ', 'REG_DATE', 'HOPE_DATE', 'LENX', 'WIDX', 'TIKX', 'LW', 'W', 'L', 'CAL_HOLE_VALUE', 'ITEM_MA_NAME', 'ITEM_NAME', 'SPCL_NAME', 'EDGE_NAME', 'GLAS_NAME', 'CONN_CPROC_NAME', 'QTY', 'QTY_NO_ALL', 'BUYER_NAME', 'TRANS_FLAG_NAME', 'CPROC_BIGO', 'LABEL_BIGO', 'BAR_CODE', 'FSET_FLAG_NAME', 'CONN_CPROC_NAME_BIGO']:
                                                if t == 'CONN_CPROC_NAME_BIGO':
                                                    #----------------------------------------------------------------------------
                                                    CONN_CPROC_NAME_BIGO = DaesungQuery.selectConnBigo(self, REG_NO, REG_SEQ)
                                                    #----------------------------------------------------------------------------
                                                    print_data = CONN_CPROC_NAME_BIGO[0]['CONN_CPROC_NAME_BIGO']
                                                elif t == 'BAR_CODE' and (PROC_CODE == '0101' or PROC_CODE == '0201'):
                                                    row, KYU = P_rows[j], P_rows[j]['KYU'].split('*')
                                                    if row['DR3_YN'] == 'Y' or row['DR6_YN'] =='Y' or row['DR16_YN'] =='Y' or row['DR17_YN'] == 'Y':
                                                        print_data = str(KYU[1]).zfill(4) + str(int(KYU[0]) - 70).zfill(4) + '1'
                                                    elif row['DR1_YN'] == 'Y' or row['DR7_YN'] =='Y' or row['DR9_YN'] =='Y' or row['DR10_YN'] =='Y' or row['DR11_YN'] =='Y' or row['DR12_YN'] =='Y' or row['DR13_YN'] == 'Y':
                                                        print_data = str(KYU[1]).zfill(4) + str(KYU[0]).zfill(4) + '1'
                                                    elif row['DR14_YN'] =='Y' or row['DR15_YN'] =='Y' or row['DR18_YN'] =='Y' or row['DR19_YN'] =='Y' or row['DR20_YN'] == 'Y' or row['TK_YN'] == 'N':
                                                        print_data = str(KYU[1]).zfill(4) + str(KYU[0]).zfill(4) + '1'
                                                    elif row['DR4_YN'] == 'Y': print_data = str(KYU[1]).zfill(4) + str(int(KYU[0]) - 35).zfill(4) + 'Y'
                                                    else: print_data = str(KYU[1]).zfill(4) + str(KYU[0]).zfill(4) + '0'
                                                    print_data = print_data + P_rows[j]['REG_NO'] + P_rows[j]['REG_SEQ'] + str(P_rows[j]['SEQ_QTY'])
                                                elif t == 'LW':
                                                    if P_rows[j]['WIDX'] == None: widx = '-'
                                                    else: widx = int(P_rows[j]['WIDX'])
                                                    if P_rows[j]['LENX'] == None: lenx = '-'
                                                    else: lenx = int(P_rows[j]['LENX'])
                                                    print_data = str(widx)[:-2].zfill(2) + str(lenx)[:-2].zfill(2)
                                                elif t == 'W':
                                                    if P_rows[j]['WIDX'] == None: print_data = ''
                                                    else: print_data = int(P_rows[j]['WIDX']) - 10
                                                elif t == 'L':
                                                    if P_rows[j]['LENX'] == None: print_data = ''
                                                    else: print_data = int(P_rows[j]['LENX']) + 10
                                                else: 
                                                    print_data = P_rows[j][t]
                                                    if print_data == None: print_data = ""
                                                    elif t == 'REG_DATE' or t == 'HOPE_DATE': print_data = "{0}.{1}.{2}".format(print_data[2:4], print_data[4:6], print_data[6:8]) 
                                                    elif t == 'LENX' or t == 'WIDX' or t == 'TIKX': print_data = int(print_data)
                                                    elif t == 'CAL_HOLE_VALUE' and hole_flag == 1: print_data = '(%d)'%int(print_data)
                                                    elif t == 'CAL_HOLE_VALUE' and hole_flag == 0: print_data = str(int(print_data))
                                                    elif t == 'QTY': print_data = '{0}/{1}'.format(P_rows[j]['SEQ_QTY'], int(print_data))
                                                textData = textData.replace("{%s}"%t, str(print_data))
                                            self.mysocket.send(textData.encode())
                                            l_count += 1
                                            #----------------------------------------------------------------------------
                                            if P_rows[j]['PRT_FLAG'] != '1': DaesungQuery.LABEL_UPDATE_SQL(self, REG_NO, P_rows[j]['REG_SEQ'], P_rows[j]['SEQ_QTY'])
                                            #----------------------------------------------------------------------------
                                            if C_REG_SEQ != REG_SEQ:
                                                C_REG_SEQ = REG_SEQ
                                                try:
                                                    if PROC_CODE == '0101': M_rows = DaesungQuery.selectDetailItem(self, self.s_date, P_rows[j]['LOT_NUMB'], REG_NO, C_REG_SEQ)
                                                    else: M_rows = DaesungQuery.selectMakeData(self, PROC_CODE, P_rows[j]['BAR_CODE'])
                                                    if M_rows == (): DaesungQuery.PR_SAVE_MAKE(self, 'insert', '0', EMPL_CODE, REG_NO, C_REG_SEQ, P_rows[j]['SORT_KEY'], self.c_date, int(P_rows[j]['QTY']), 0) #실적등록
                                                except: pass
                                                #----------------------------------------------------------------------------
                                            time.sleep(0.3)
                                        except: logging.debug("printLabel : selectDetailList 실패")
                                    except: logging.debug("printLabel : txt read 실패")
                            else: logging.debug("printLabel : 등록된 바코드 없음")
                        #----------------------------------------------------------------------------
                        try: DaesungQuery.SELECT_PR_PASS_JAKUP_MAKE(self)
                        except: pass
                        #----------------------------------------------------------------------------
                    except: logging.debug("printLabel : select 실패")
                    self.mysocket.close()
                    self.DBload() #DB로드
                    self.th_rowCount.start()
                except: self.print_status.setStyleSheet("background-color: #fd97a5;") #red
            else: MessageWindow(self, "출력할 LOT를 선택해주세요.").showModal()
        else: MessageWindow(self, "프린터를 연결해주세요.").showModal()
    
    @pyqtSlot(int, str)
    def newData(self, count, time):
        global DATA_COUNT, JAKUP_TIME
        if time == 'f': self.connectDBThread()
        elif time == 'None': self.DB_btn.setStyleSheet("background-color: #55cba7;") #green
        elif count == DATA_COUNT and int(time) == JAKUP_TIME: self.DB_btn.setStyleSheet("background-color: #55cba7;") #green
        elif count != DATA_COUNT or int(time) != JAKUP_TIME:
            self.DB_btn.setStyleSheet("background-color: #fd97a5;") #red
            if self.light_flag == 'success':
                try:
                    light_ser.write('RY 1 1\r'.encode()) #red light
                    logging.debug("newData : RY 1 1 성공")
                except: logging.debug("newData : RY 1 1 실패")
            self.DBload() #DB로드
    
    #---------------------------------------------------------------------------------------------------
    def clickedCtrl(self, flag):
        self.key_flag = flag
    
    def click_row(self, index):
        try:
            JAKUP_FLAG = self.tableWidget.item(index.row(), 4).text()
            LOT_NO = self.tableWidget.item(index.row(), 5).text()
            REG_NO = self.tableWidget.item(index.row(), 9).text()
            s_date = self.date_btn.text()
            print_window = MesDetailWindow(JAKUP_FLAG, LOT_NO, REG_NO, s_date) #DETAIL 화면 연결
            try: self.th_rowCount.terminate()
            except: pass
            widget.addWidget(print_window)
            widget.setCurrentIndex(widget.currentIndex() + 1)
            self.deleteLater()
        except: logging.debug("clickRow : 상세페이지 연결 실패")
    
    def selectedAll(self, num):
        global SORT_ARRAY
        if len(SORT_ARRAY) == 5: SORT_ARRAY.pop(0)
        SORT_ARRAY.append(num)
        #-----------------------------------------------------------------
        if num == 3 and self.check_flag == False:
            for checkbox in self.checkBoxList: checkbox.setChecked(True)
            self.check_flag = True
        elif num == 3 and self.check_flag == True:
            for checkbox in self.checkBoxList: checkbox.setChecked(False)
            self.check_flag = False
    
    #---------------------------------------------------------------------------------------------------
    def jackupPrint(self, flag):
        DaesungFunctions.jackupPrint(self, WC_CODE, PROC_CODE, '', EMPL_CODE, '') #작업지시서 인쇄
    
    #공정 변경
    def procChanged(self):
        global PROC_CODE, PROC_NAME
        self.tableWidget.clearSelection()
        if PROC_CODE[:2] != self.aproc_combo.currentText()[:2]: self.checkBoxList = []
        PROC_CODE = self.aproc_combo.currentText()[:4]
        PROC_NAME = self.aproc_combo.currentText()[4:]
        if PROC_CODE == '0120':
            self.threadTerminate()
            widget.addWidget(MesPackWindow(TodayData, 1))
            widget.setCurrentIndex(widget.currentIndex() + 1)
            self.deleteLater()
        else:
            self.DB_flag, self.reload_num = 0, 0
            DaesungFunctions.replaceDate(self) #DB로드
        self.hwConnect() #HW연결
    
    def flagChanged(self):
        global PRT_FLAG, PRT_NAME
        self.tableWidget.clearSelection()
        self.checkBoxList = []
        PRT_NAME = self.flag_combo.currentText()
        if PRT_NAME == '전체': PRT_FLAG = '%'
        elif PRT_NAME == '대기': PRT_FLAG = '0'
        elif PRT_NAME == '부분': PRT_FLAG = '2'
        elif PRT_NAME == '완료': PRT_FLAG = '1'
        DaesungFunctions.replaceDate(self) #DB로드
    
    def resetLight(self):
        if self.light_flag == 'success':
            try:
                light_ser.write('RY 1 0\r'.encode()) #green light
                logging.debug("resetLight : RY 1 0 성공")
            except: logging.debug("resetLight : RY 1 0 실패")
        self.DB_btn.setStyleSheet("background-color: #55cba7;") #green
    
    def showDate(self, date):
        self.calender.hide()
        a = date.toString().split()
        if len(a[1]) == 1: a[1] = "0" + a[1]
        if len(a[2]) == 1: a[2] = "0" + a[2]
        self.date_btn.setText("%s-%s-%s "%(a[3],a[1],a[2])) # 월, 일자가 1자 일때
        self.calendar_flag = False
        global PRT_INDEX
        PRT_INDEX = -1
        DaesungFunctions.replaceDate(self) #DB로드
    
    def mousePressEvent(self, event):
        if event.buttons() == Qt.LeftButton and self.calendar_flag == True:
            self.calender.hide()
            self.calendar_flag = False
    
    def setData(self):
        try:
            SetWindow().showModal()
            self.hwConnect() #HW연결
        except: logging.debug("setData : 상세페이지 연결 실패")
    
    def logWindow(self):
        try:
            try: self.th_rowCount.terminate()
            except: pass
            MesLogWindow(TodayData).showModal()
            self.th_rowCount.start()
        except: logging.debug("logWindow : 로그 화면 연결 실패")
    
    #실시간 DB연결
    def connectDBThread(self):
        if self.DB_flag == 0:
            self.DB_flag = 1
            MessageWindow(self, "DB연결 실패").showModal()
            self.DB_th = ConnectDBThread()
            self.DB_th.sig_data.connect(self.DbThreadSlot)
            self.DB_th.start()
    
    @pyqtSlot(int)       
    def DbThreadSlot(self, con):
        if con == 1:
            self.DB_th.terminate()
            self.DB_flag = 0
            #----------------------------------------------------------------------------
            result = DaesungQuery.connectDB(self, host, port, user, name)
            #----------------------------------------------------------------------------
            if result == 'success': DaesungFunctions.replaceDate(self) #DB로드
    
    def threadTerminate(self):
        try: self.th_rowCount.terminate()
        except: pass
        #--------------------------------------------------
        try: light_ser.close()
        except: pass
    
    def exitWindow(self):
        MessageWindow(self, 2).showModal()
        #----------------------------------
        if RECONNECT_FLAG == 1: self.threadTerminate()

################################################################################################################
#DETAIL Window
class MesDetailWindow(QDialog):
    def __init__(self, gubun, lot, reg_no, date):
        super(MesDetailWindow, self).__init__()
        loadUi("ui\DAESUNG_MES_D.ui", self)
        
        DaesungFunctions.setDetailStyle(self, date, WC_CODE, PROC_CODE, gubun, lot, reg_no) #기본 셋팅
        
        if PROC_CODE == C_PROC or PROC_CODE == CS_PROC: self.label_combo.setCurrentIndex(1)
        elif PROC_CODE == Z_PROC or PROC_CODE == ZS_PROC: self.label_combo.setCurrentIndex(2)
        else: self.label_combo.setCurrentIndex(0)
        
        if WC_CODE == '02':
            self.ORDER = 'DJAKUP.WIDX, DJAKUP.LENX,'
            self.tableWidget.setSortingEnabled(True)
            self.tableWidget.setSelectionMode(QAbstractItemView.ExtendedSelection)
        else: self.ORDER = 'DJAKUP.WIDX, DJAKUP.LENX,'
        
        self.JAKUP_APPR_FLAG = '2'
        self.PROC_CODE, self.W_DATA = "MPJAKUP.PROC_CODE = '{0}'".format(PROC_CODE), ''
        
        self.setPrintIp()
        self.DBload() #DB로드
        
        self.tableWidget.horizontalHeader().sectionClicked.connect(self.selectedAll)
        self.select_all.clicked.connect(lambda: self.selectedAll(3))
        
        self.label_combo.currentIndexChanged.connect(self.connectPrint) #바코드 프린터 연결
        self.print_btn.clicked.connect(self.printLabel) #라벨 발행
        self.jackup_btn.clicked.connect(lambda: DaesungFunctions.jackupPrint(self, WC_CODE, PROC_CODE, lot, EMPL_CODE, '')) #작업지시서 인쇄
       
        self.tableWidget.currentCellChanged.connect(self.connectTable)
        self.tableWidget.cellPressed.connect(self.clickedRow)
        self.tableWidget.cellClicked.connect(self.clickedRow)
        
        # 화살표버튼 --------------------------------------------------
        self.top.clicked.connect(lambda: DaesungFunctions.topData(self))
        self.prev.clicked.connect(lambda: DaesungFunctions.prevData(self))
        self.next.clicked.connect(lambda: DaesungFunctions.nextData(self))
        self.bottom.clicked.connect(lambda: DaesungFunctions.bottomData(self))
        
        # 상단버튼 --------------------------------------------------
        self.DB_btn.clicked.connect(self.resetLight)
        self.set_btn.clicked.connect(self.setData)
        self.log_btn.clicked.connect(self.logWindow)
        self.as_btn.clicked.connect(lambda: DaesungFunctions.openUrl(self))
        self.back_btn.clicked.connect(self.back)
    
    def setPrintIp(self):
        try:
            self.label_flag = self.label_combo.currentIndex()
            self.set_win = SetWindow()
            self.label_combo.clear()
            if WC_CODE == '02':
                for c_text in ['심재라벨(수동)', '판재라벨(수동)', '엣지라벨(수동)']: self.label_combo.addItem(c_text)
            elif PROC_CODE == '0101':
                self.s_ip = str(self.set_win.printer_ip_input.text()).split('.')
                #---------------------------------------------------------------------
                self.label_combo.addItem('심재라벨({0})'.format(self.s_ip[3]))
                self.label_combo.addItem('판재라벨({0})'.format(int(self.s_ip[3]) + 1))
                self.label_combo.addItem('엣지라벨({0})'.format(int(self.s_ip[3]) + 1))
            self.label_combo.setCurrentIndex(self.label_flag)
            self.connectPrint()
        except: logging.debug("setPrintIp : failed")
    
    def connectPrint(self):
        try:
            self.count = 0
            if WC_CODE == '02':
                flag = 1
                if self.label_combo.currentIndex() == 0 and self.set_win.printer_check.isChecked():
                    self.ip = self.set_win.printer_ip_input.text()
                    self.port = int(self.set_win.printer_port_input.text())
                    self.mode_edit = self.set_win.printer_mode_check
                    self.mode = '^' + self.set_win.printer_mode_input.text().replace(' ', '')
                    try: self.count = int(self.set_win.printer_count_input.text())
                    except: pass
                elif self.label_combo.currentIndex() == 1 and self.set_win.printer2_check.isChecked(): 
                    self.ip = self.set_win.printer2_ip_input.text()
                    self.port = int(self.set_win.printer2_port_input.text())
                    self.mode_edit = self.set_win.printer2_mode_check
                    self.mode = '^' + self.set_win.printer2_mode_input.text().replace(' ', '')
                    try: self.count = int(self.set_win.printer_count_input.text())
                    except: pass
                elif self.label_combo.currentIndex() == 2 and self.set_win.printer3_check.isChecked(): 
                    self.ip = self.set_win.printer3_ip_input.text()
                    self.port = int(self.set_win.printer3_port_input.text())
                    self.mode_edit = self.set_win.printer3_mode_check
                    self.mode = '^' + self.set_win.printer3_mode_input.text().replace(' ', '')
                    try: self.count = int(self.set_win.printer3_count_input.text())
                    except: pass
                else:
                    self.printer_flag, flag = "unable", 0
                    logging.debug("connectPrint : 프린트 비활성")
                    self.print_status.setStyleSheet("background-color: #CDCDCD;") #gray
                #---------------------------------------------------------------------
                if flag == 1:
                    print_socket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
                    print_socket.settimeout(0.5)
                    print_socket.connect((self.ip, self.port))
                    print_socket.close()
                    #---------------------------------------------------------------------
                    self.printer_flag = "success"
                    logging.debug("connectPrint : 프린트 연결 성공")
                    self.print_status.setStyleSheet("background-color: #55cba7;") #green        
            else:
                if self.set_win.printer_check.isChecked():
                    if PROC_CODE == '0101':
                        ip_last = re.findall(r'[0-9]+', self.label_combo.currentText())[0]
                        self.ip = '{0}.{1}.{2}.{3}'.format(self.s_ip[0], self.s_ip[1], self.s_ip[2], ip_last)
                    else: self.ip = self.set_win.printer_ip_input.text()
                    self.port = int(self.set_win.printer_port_input.text())
                    self.mode =  '^' + self.set_win.printer_mode_input.text().replace(' ', '')
                    #---------------------------------------------------------------------
                    print_socket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
                    print_socket.settimeout(0.5)
                    print_socket.connect((self.ip, self.port))
                    print_socket.close()
                    #---------------------------------------------------------------------
                    self.printer_flag = "success"
                    logging.debug("connectPrint : 프린트 연결 성공")
                    self.print_status.setStyleSheet("background-color: #55cba7;") #green
                else:
                    self.printer_flag = "unable"
                    logging.debug("connectPrint : 프린트 비활성")
                    self.print_status.setStyleSheet("background-color: #CDCDCD;") #gray
        except:
            self.printer_flag = "failed"
            logging.debug("connectPrint : 프린트 연결 실패")
            self.print_status.setStyleSheet("background-color: #fd97a5;") #red
    
    #---------------------------------------------------------------------------------------------------
    def setCheckFlag(self):
        try:
            self.set_check = []
            for count, checkbox in enumerate(self.checkBoxList):
                if checkbox.isChecked() == True: self.set_check.append(count)
        except: pass
        
    def DBload(self):
        self.setCheckFlag()
        self.tableWidget.clearSelection()
        self.tableWidget.setRowCount(0)
        self.checkBoxList = []
        #----------------------------------------------------------------------------
        D_rows = DaesungQuery.selectDetailList(self, self.REG_NO, '%', '%', self.s_date, PROC_CODE, self.ORDER)
        #----------------------------------------------------------------------------
        if D_rows == 'failed': self.connectDBThread()
        elif D_rows == ():
            logging.debug("DBload : 작업지시서 취소됨")
            MessageWindow(self, "해당 작업지시서가 취소되었습니다.").showModal()
            self.back()
        else:
            try:
                self.tableWidget.setRowCount(len(D_rows))
                for i in range(len(D_rows)):
                    self.tableWidget.setRowHeight(i, 85)
                    ckbox = QCheckBox()
                    ckbox.setStyleSheet(t_checkStyle)
                    self.checkBoxList.append(ckbox)
                    #----------------------------------------------------------------------------
                    if WC_CODE == '02':
                        MAKE_ARRAY = ['', '', '']
                        MAKE_FLAG, PROC = D_rows[i]['P_LK_MAKE_FLAG'].split(','), D_rows[i]['P_PROC_CODE'].split(',')
                        for a in range(len(PROC)):
                            if PROC[a] == S_PROC or PROC[a] == SS_PROC: MAKE_ARRAY[0] = MAKE_FLAG[a]
                            elif PROC[a] == C_PROC or PROC[a] == CS_PROC: MAKE_ARRAY[1] = MAKE_FLAG[a]
                            elif PROC[a] == ZS_PROC or PROC[a] == ZS_PROC: MAKE_ARRAY[2] = MAKE_FLAG[a]
                    #----------------------------------------------------------------------------
                    for count, j in enumerate(self.col):
                        if j == 'CHECK':
                            if WC_CODE == '01' and D_rows[i]['NO_YN'] == 'Y': pass
                            else:
                                cellWidget = QWidget()
                                layoutCB = QHBoxLayout(cellWidget)
                                layoutCB.addWidget(self.checkBoxList[i])
                                layoutCB.setAlignment(QtCore.Qt.AlignCenter)
                                layoutCB.setContentsMargins(0, 0, 0, 0)
                                cellWidget.setLayout(layoutCB)
                                self.tableWidget.setCellWidget(i, count, cellWidget)
                                if i in self.set_check: self.checkBoxList[i].setChecked(True)
                        elif j == 'PRT_FLAG':
                            states = QPushButton()
                            if D_rows[i][j] == '1': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/완료.png);")
                            else: states.setStyleSheet("background: none; border: none;")
                            self.state_group.addButton(states, i)
                            self.tableWidget.setCellWidget(i, count, states)
                        elif j == '' and WC_CODE == '02':
                            print_data = MAKE_ARRAY[count]
                            states = QPushButton()
                            if print_data == '00': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/대기.png);")
                            elif print_data == '99' or print_data == '01': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/진행.png);")
                            elif print_data == '90' or print_data == '02': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/완료.png);")
                            else: states.setStyleSheet(stateBtnStyle + "background: none;")
                            self.make_array[count].addButton(states, i)
                            self.tableWidget.setCellWidget(i, count, states)
                        elif j == '': self.tableWidget.setItem(i, count, QTableWidgetItem(''))
                        else:
                            print_data = D_rows[i][j]
                            if print_data == None: print_data = ''
                            elif j == 'HOPE_DATE': print_data = "{0}/{1}".format(print_data[4:6], print_data[6:8])
                            elif j == 'ABS_LENX' or j == 'ABS_WIDX': print_data = int(print_data)
                            elif j == 'QTY_NO_ALL' and PROC_CODE == '0103': print_data = int(D_rows[i]['ABS_QTY'])
                            elif j == 'QTY': print_data = int(print_data)
                            item_data = QTableWidgetItem(str(print_data))
                            if j == 'REG_SEQ' or j == 'HOPE_DATE': item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                            elif j == 'KYU' or j == 'QTY': item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                            elif j == 'ABS_LENX' or j == 'ABS_WIDX':
                                item_data.setForeground(QBrush(QColor(63, 139, 204))) #blue
                                item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                            elif j == 'QTY_NO_ALL':
                                if PROC_CODE == '0103': item_data.setForeground(QBrush(QColor(63, 139, 204))) #blue
                                item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                            elif j == 'SPCL_NAME' and PROC_CODE == '0103': item_data.setForeground(QBrush(QColor(66, 188, 112))) #green
                            self.tableWidget.setItem(i, count, item_data)
                DaesungFunctions.tableWidth(self, PROC_CODE, WC_CODE, len(D_rows))
                if self.reload_num == 0:
                    try:
                        self.th_rowCount = lotCountThread(self.s_date, self.PROC_CODE, self.JAKUP_APPR_FLAG, self.W_DATA)
                        self.th_rowCount.sig_data.connect(self.newData)
                        self.th_rowCount.start()
                        self.reload_num = 1
                    except: logging.debug("DBload : th_rowCount 실패")
                self.DB_btn.setStyleSheet("background-color: #55cba7;") #green
            except: logging.debug("DBload : table 로드 실패")
    
    #라벨 발행
    def printLabel(self):
        checkArray, l_count = [], 1
        if self.printer_flag == "success":
            for count, checkbox in enumerate(self.checkBoxList):
                if checkbox.isChecked() == True: checkArray.append(count)
            if checkArray != []:
                try:
                    self.mysocket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
                    self.mysocket.settimeout(0.5)
                    self.mysocket.connect((self.ip, self.port))
                    self.print_status.setStyleSheet("background-color: #55cba7;") #green
                    try: self.th_rowCount.terminate()
                    except: pass
                    if WC_CODE == '02':
                        if PROC_CODE[:2] == '01':
                            if self.label_combo.currentIndex() == 0: self.fileName, mode_edit, hole_flag, PROC = ss_label, self.set_win.printer_mode_check, 0, S_PROC
                            elif self.label_combo.currentIndex() == 1: self.fileName, mode_edit, hole_flag, PROC = cs_label, self.set_win.printer2_mode_check, 0, C_PROC
                            elif self.label_combo.currentIndex() == 2: self.fileName, mode_edit, hole_flag, PROC = zs_label, self.set_win.printer3_mode_check, 1, Z_PROC
                        else:
                            if self.label_combo.currentIndex() == 0: self.fileName, mode_edit, hole_flag, PROC = ss_label, self.set_win.printer_mode_check, 0, SS_PROC
                            elif self.label_combo.currentIndex() == 1: self.fileName, mode_edit, hole_flag, PROC = cs_label, self.set_win.printer2_mode_check, 0, CS_PROC
                            elif self.label_combo.currentIndex() == 2: self.fileName, mode_edit, hole_flag, PROC = zs_label, self.set_win.printer3_mode_check, 1, ZS_PROC
                    elif WC_CODE == '19': self.fileName, mode_edit, hole_flag, PROC = if_label, self.set_win.printer_mode_check, 0, PROC_CODE
                    else:
                        if PROC_CODE == '0101' and self.label_combo.currentIndex() == 0: self.fileName, mode_edit, hole_flag, PROC = s_label, self.set_win.printer_mode_check, 0, S_PROC
                        elif PROC_CODE == '0101' and self.label_combo.currentIndex() == 2: self.fileName, mode_edit, hole_flag, PROC = z300_label, self.set_win.printer_mode_check, 1, Z_PROC
                        else: self.fileName, mode_edit, hole_flag, PROC = c_label, self.set_win.printer_mode_check, 0, C_PROC
                    try:
                        t_count = len(checkArray)
                        for i in checkArray:
                            REG_SEQ, QTY_NO, SEQ_QTY = self.tableWidget.item(i, 4).text(), self.tableWidget.item(i, 12).text(), self.tableWidget.item(i, 14).text()
                            f_name = open(self.fileName, 'r', encoding = 'utf-8')
                            textData = f_name.read()
                            f_name.close()
                            if self.mode == '^MMC' and mode_edit.isChecked() == True:
                                if l_count == t_count or l_count == self.count:
                                    textData = textData.replace("^CI28^MMT", "^CI28^MMC")
                                    t_count = t_count - l_count
                                    l_count = 0
                            elif self.mode != '^MMC' and mode_edit.isChecked() == True: textData = textData.replace("^CI28", "^CI28" + self.mode)
                            if self.set_win.printer_po_check.isChecked() == True: textData = textData.replace("^LS0", "^LS0^POI")
                            elif self.set_win.printer_po_check.isChecked() == False: textData = textData.replace("^LS0", "^LS0^PON")
                            #----------------------------------------------------------------------------
                            P_rows = DaesungQuery.selectDetailList(self, self.REG_NO, REG_SEQ, SEQ_QTY, self.s_date, PROC, self.ORDER)
                            #----------------------------------------------------------------------------
                            if P_rows == 'failed': self.connectDBThread()
                            elif P_rows != []:
                                try:
                                    for i in ['REG_NO', 'LOT_NUMB', 'REG_SEQ', 'REG_DATE', 'HOPE_DATE', 'LENX', 'WIDX', 'TIKX', 'LW', 'W', 'L', 'CAL_HOLE_VALUE', 'ITEM_MA_NAME', 'ITEM_NAME', 'SPCL_NAME', 'EDGE_NAME', 'GLAS_NAME', 'CONN_CPROC_NAME', 'QTY_NO_ALL', 'QTY', 'BUYER_NAME', 'TRANS_FLAG_NAME', 'BIGO', 'CPROC_BIGO', 'LABEL_BIGO', 'BAR_CODE', 'FSET_FLAG_NAME', 'CONN_CPROC_NAME_BIGO']:
                                        if i == 'CONN_CPROC_NAME_BIGO':
                                            #----------------------------------------------------------------------------
                                            CONN_CPROC_NAME_BIGO = DaesungQuery.selectConnBigo(self, self.REG_NO, REG_SEQ)
                                            #----------------------------------------------------------------------------
                                            print_data = CONN_CPROC_NAME_BIGO[0]['CONN_CPROC_NAME_BIGO']
                                        elif i == 'BAR_CODE' and self.label_combo.currentIndex() == 0:
                                            row, KYU = P_rows[0], P_rows[0]['KYU'].split('*')
                                            if row['DR3_YN'] == 'Y' or row['DR6_YN'] =='Y' or row['DR16_YN'] =='Y' or row['DR17_YN'] == 'Y':
                                                print_data = str(KYU[1]).zfill(4) + str(int(KYU[0]) - 70).zfill(4) + '1'
                                            elif row['DR1_YN'] == 'Y' or row['DR7_YN'] =='Y' or row['DR9_YN'] =='Y' or row['DR10_YN'] =='Y' or row['DR11_YN'] =='Y' or row['DR12_YN'] =='Y' or row['DR13_YN'] == 'Y':
                                                print_data = str(KYU[1]).zfill(4) + str(KYU[0]).zfill(4) + '1'
                                            elif row['DR14_YN'] =='Y' or row['DR15_YN'] =='Y' or row['DR18_YN'] =='Y' or row['DR19_YN'] =='Y' or row['DR20_YN'] == 'Y' or row['TK_YN'] == 'N':
                                                print_data = str(KYU[1]).zfill(4) + str(KYU[0]).zfill(4) + '1'
                                            elif row['DR4_YN'] == 'Y': print_data = str(KYU[1]).zfill(4) + str(int(KYU[0]) - 35).zfill(4) + 'Y'
                                            else: print_data = str(KYU[1]).zfill(4) + str(KYU[0]).zfill(4) + '0'
                                            print_data = print_data + P_rows[0]['REG_NO'] + P_rows[0]['REG_SEQ'] + str(P_rows[0]['SEQ_QTY'])
                                        elif i == 'LW':
                                            if P_rows[0]['WIDX'] == None: widx = '-'
                                            else: widx = int(P_rows[0]['WIDX'])
                                            if P_rows[0]['LENX'] == None: lenx = '-'
                                            else: lenx = int(P_rows[0]['LENX'])
                                            print_data = str(widx)[:-2].zfill(2) + str(lenx)[:-2].zfill(2)
                                        elif i == 'W':
                                            if P_rows[0]['WIDX'] == None: print_data = ''
                                            else: print_data = int(P_rows[0]['WIDX']) - 10
                                        elif i == 'L':
                                            if P_rows[0]['LENX'] == None: print_data = ''
                                            else: print_data = int(P_rows[0]['LENX']) + 10
                                        else:
                                            print_data = P_rows[0][i]
                                            if print_data == None: print_data = ""
                                            elif i == 'REG_DATE' or i == 'HOPE_DATE': print_data = "{0}.{1}.{2}".format(print_data[2:4], print_data[4:6], print_data[6:8]) 
                                            elif i == 'LENX' or i == 'WIDX' or i == 'TIKX': print_data = int(print_data)
                                            elif i == 'CAL_HOLE_VALUE' and hole_flag == 1: print_data = '(%d)'%int(print_data)
                                            elif i == 'CAL_HOLE_VALUE' and hole_flag == 0: print_data = str(int(print_data))
                                            elif i == 'QTY_NO_ALL': print_data = QTY_NO
                                            elif i == 'QTY': print_data = '{0}/{1}'.format(P_rows[0]['SEQ_QTY'], int(print_data))
                                        textData = textData.replace("{%s}"%i, str(print_data))
                                    self.mysocket.send(textData.encode())
                                    l_count += 1
                                    try:
                                        #----------------------------------------------------------------------------
                                        if P_rows[0]['PRT_FLAG'] != '1': DaesungQuery.LABEL_UPDATE_SQL(self, self.REG_NO, REG_SEQ, SEQ_QTY)
                                        #----------------------------------------------------------------------------
                                        if PROC == '0101': M_rows = DaesungQuery.selectDetailItem(self, self.s_date, P_rows[0]['LOT_NUMB'], self.REG_NO, P_rows[0]['REG_SEQ'])
                                        else: M_rows = DaesungQuery.selectMakeData(self, PROC, P_rows[0]['BAR_CODE'])
                                        #----------------------------------------------------------------------------
                                        if M_rows == (): DaesungQuery.PR_SAVE_MAKE_BAR_DETAIL(self, 'insert', '0', EMPL_CODE, self.REG_NO, P_rows[0]['REG_SEQ'], P_rows[0]['SORT_KEY'], P_rows[0]['BAR_CODE'], self.c_date, 1, 0) #실적등록
                                        time.sleep(0.3)
                                    except: pass
                                except: logging.debug("printLabel : selectDetailList 실패")
                            else: logging.debug("printLabel : 등록된 바코드 없음")
                        #----------------------------------------------------------------------------
                        DaesungQuery.SELECT_PR_PASS_JAKUP_MAKE(self)
                        #----------------------------------------------------------------------------
                        self.mysocket.close()
                    except: logging.debug("printLabel : select 실패")
                    self.DBload()
                    self.th_rowCount.start()
                except:
                    self.print_status.setStyleSheet("background-color: #fd97a5;") #red
                    logging.debug("printLabel : 프린터 연결 실패")
            else: MessageWindow(self, "출력할 라벨을 선택해주세요.").showModal()
        else: MessageWindow(self, "프린터를 연결해주세요.").showModal()
    
    #---------------------------------------------------------------------------------------------------
    def connectTable(self, row):
        self.checkBoxList[row].setChecked(True)
    
    def clickedRow(self, row):
        s_index, s_row = -1, []
        for item in self.tableWidget.selectedIndexes():
            if s_index != item.row():
                    s_index = item.row()
                    s_row.append(s_index)
        for i in range(self.tableWidget.rowCount()):
            if i in s_row: self.checkBoxList[row].setChecked(True)
            else: self.checkBoxList[i].setChecked(False)
        self.checkBoxList[row].setChecked(True)
    
    def selectedAll(self, num):
        if num == 3 and self.check_flag == False:
            for checkbox in self.checkBoxList: checkbox.setChecked(True)
            self.check_flag = True
        elif num == 3 and self.check_flag == True:
            for checkbox in self.checkBoxList: checkbox.setChecked(False)
            self.check_flag = False
    
    #---------------------------------------------------------------------------------------------------
    @pyqtSlot(int, str)
    def newData(self, count, time):
        global DATA_COUNT, JAKUP_TIME
        if time == 'f': self.connectDBThread()
        elif time == 'None':
            MessageWindow(self, "해당 작업지시서가 취소되었습니다.").showModal()
            self.back()
        elif count == DATA_COUNT and int(time) == JAKUP_TIME: self.DB_btn.setStyleSheet("background-color: #55cba7;") #green
        elif count != DATA_COUNT or int(time) != JAKUP_TIME:
            self.DB_btn.setStyleSheet("background-color: #fd97a5;") #red
            DATA_COUNT = count
            JAKUP_TIME = int(time)
    
    @pyqtSlot(int)
    def FormatSlot(self, num):
        if num == 3:
            try: light_ser.write('RY 1 0\r'.encode()) #green light
            except: pass
    
    #---------------------------------------------------------------------------------------------------
    def resetLight(self):
        try: light_ser.write('RY 1 0\r'.encode()) #green light
        except: pass
        self.DB_btn.setStyleSheet("background-color: #55cba7;") #green
    
    def setData(self):
        try:
            SetWindow().showModal()
            self.setPrintIp()
        except: logging.debug("setData : 상세페이지 연결 실패")
    
    def logWindow(self):
        try:
            try: self.th_rowCount.terminate()
            except: pass
            MesLogWindow(TodayData).showModal()
            self.th_rowCount.start()
        except: logging.debug("logWindow : 로그 화면 연결 실패")
    
    #실시간 DB연결
    def connectDBThread(self):
        self.DB_btn.setStyleSheet("background-color: #fd97a5;") #red
        if self.DB_flag == 0:
            self.DB_flag = 1
            logging.debug("DBload : DB로드 실패")
            MessageWindow(self, "DB연결 실패").showModal()
            #----------------------------------------------------------------------------
            self.DB_th = ConnectDBThread()
            self.DB_th.sig_data.connect(self.DbThreadSlot)
            self.DB_th.start()
    
    @pyqtSlot(int)
    def DbThreadSlot(self, con):
        if con == 1:
            self.DB_th.terminate()
            self.DB_flag = 0
            #----------------------------------------------------------------------------
            result = DaesungQuery.connectDB(self, host, port, user, name)
            #----------------------------------------------------------------------------
            if result == 'success':
                self.connectPrint() #바코드 프린터 연결
                self.DBload() #DB로드
    
    def back(self):
        try: self.light_th.terminate()
        except: pass
        try: self.th_rowCount.terminate()
        except: pass
        set_date = self.date_btn.text()
        widget.addWidget(MesLotWindow(set_date))
        widget.setCurrentIndex(widget.currentIndex() + 1)
        self.deleteLater()

################################################################################################################
class ScannerThread(QThread):
    sig_data = pyqtSignal(str)
        
    def run(self):
        STX = '\x02'
        while True:
            try:
                data = scanner_socket.recv(2048)
                data = data.decode()
                data = data.replace(STX,'').replace(" ", '') #공백 제거
                self.sig_data.emit(data)
            except: pass
            time.sleep(1)

#---------------------------------------------------------------------------------------------------
def RS232Addr(flag, data):
    if flag == 'SEQ':
        HEX1 = hex(ord(data[:1])).replace('0x', '')
        HEX2 = hex(ord(data[1:])).replace('0x', '')
        RSDATA = '25 {0} {1}'.format(HEX1, HEX2)
    elif flag.find('L') >= 0:
        if flag == 'L':
            if len(data) > 2: LENGTH = str(hex(13 + len(data))[2:]) + ' 00'
            else: LENGTH = '0' + str(hex(13 + len(data))[2:]) + ' 00'
        elif flag == 'WL': LENGTH = str(hex(17 + len(data))[2:]) + ' 00'
        elif flag == 'RRL': LENGTH = str(hex(15 + len(data))[2:]) + ' 00'
        N_LENGTH = '0' + str(hex(4 + (len(data)-1))[2:]) + ' 00'
        ADDR = data.encode('utf-8').hex()
        RSDATA = [LENGTH, N_LENGTH, ADDR]
    elif flag == 'SIZE':
        if len(str(data)) == 1: D_SIZE1 = '0' + hex(data)[-1:]
        else: D_SIZE1 = hex(data)[-2:]
        if data <= 255: D_SIZE2 = '00'
        elif data <= 4095: D_SIZE2 = '0' + hex(data)[2]
        elif data <= 65535: D_SIZE2 = hex(data)[2:4]
        else: D_SIZE2 = ''
        RSDATA = [D_SIZE1, D_SIZE2]
    return RSDATA

class PlcCompThread(QThread):
    sig_data = pyqtSignal(int)
    read_data = pyqtSignal(list)
    
    def __init__(self, addr, r_addr):
        super().__init__()
        self.addr = addr
        self.r_addr = r_addr
    
    def run(self):
        LENGTH = RS232Addr('L', self.addr)
        s_data = bytes.fromhex(COMPANY + LENGTH[0] + POSTION + '54 00 02 00' + BLOCK + LENGTH[1] + '25 44 57' + LENGTH[2])
        if self.r_addr != '':
            LENGTH1 = RS232Addr('RRL', str(self.r_addr[0]*2))
            s_data1 = bytes.fromhex(COMPANY + LENGTH1[0] + POSTION + '54 00 14 00' + BLOCK + LENGTH1[1] + '25 44 42' + LENGTH1[2] + '0E 00')
            LENGTH2 = RS232Addr('RRL', str(self.r_addr[1]*2))
            s_data2 = bytes.fromhex(COMPANY + LENGTH2[0] + POSTION + '54 00 14 00' + BLOCK + LENGTH2[1] + '25 44 42' + LENGTH2[2] + '0E 00')
        while True:
            try:
                #----------------------------------------------------------------------------
                plc_socket.send(s_data)
                #----------------------------------------------------------------------------
                re = plc_socket.recv(1024).hex()
                comp = int((re[-2:] + re[-4:-2]), 16)
                if self.r_addr != '':
                    try:
                        d_array = []
                        for data in [s_data1, s_data2]:
                            #----------------------------------------------------------------------------
                            plc_socket.send(data)
                            #----------------------------------------------------------------------------
                            re = plc_socket.recv(1024).hex()
                            b = [str(re[60:][4:])[i:i+8] for i in range(0, len(str(re[60:][4:])), 8)]
                            d_array.append(b)
                    except: d_array = [99999]
                    self.read_data.emit(d_array)
            except: comp = 99999
            self.sig_data.emit(comp)
            time.sleep(1)

class PlcWriteThread(QThread):
    sig_data = pyqtSignal(int)
    
    def __init__(self, data, addr):
        super().__init__()
        self.data = data
        self.addr = addr
    
    def run(self):
        try:
            for j in range(2):
                for i in range(len(self.data)):
                    LENGTH = RS232Addr('WL', self.addr[i])
                    D_SIZE = RS232Addr('SIZE', self.data[i])
                    s_data = bytes.fromhex(COMPANY + LENGTH[0] + POSTION + '58 00 02 00' + BLOCK + LENGTH[1] + '25 44 57' + LENGTH[2] + '02 00' + D_SIZE[0] + D_SIZE[1])
                    #-------------------------------------------------------------
                    plc_socket.send(s_data)
                    #-------------------------------------------------------------
                    re = plc_socket.recv(1024)
                    logging.debug("PlcWriteThread : {0} 전송".format(self.data[i]))
                time.sleep(1)
            self.sig_data.emit(j)
        except:
            logging.debug("PlcWriteThread : 전송 실패")
            self.sig_data.emit(99999)

#---------------------------------------------------------------------------------------------------------------
#자동화(접착, 테노너) Window
class MesWindow(QDialog):
    def __init__(self, date):
        super(MesWindow, self).__init__()
        loadUi("ui\DAESUNG_MES_A.ui", self)
        
        DaesungFunctions.setAutoStyle(self, date, PROC_CODE, PROC_NAME) #기본 셋팅
        
        self.hwConnect() #HW연결
        DaesungFunctions.replaceDate(self) #DB로드
        try:
            self.focus_th = FormatThread("focus")
            self.focus_th.sig_data.connect(self.FormatSlot)
            self.focus_th.start()
            #------------------------------------------------
            self.format_th = FormatThread("format")
            self.format_th.sig_data.connect(self.FormatSlot)
            #------------------------------------------------
            self.light_th = FormatThread("light")
            self.light_th.sig_data.connect(self.FormatSlot)
            #------------------------------------------------
            self.clear_th = FormatThread("clear")
            self.clear_th.sig_data.connect(self.FormatSlot)
        except: pass
        
        self.QR_INPUT.returnPressed.connect(self.PressedEnterKey) #QR코드 입력
        self.edgeCode_btn.clicked.connect(self.edgeCode) #엣지코드리스트 화면 연결
        self.state_group.buttonClicked[int].connect(self.itemCancel) #전표 취소
        
        # 상단버튼 --------------------------------------------------
        self.date_btn.clicked.connect(lambda: DaesungFunctions.calendar(self))
        self.set_btn.clicked.connect(self.setData)
        self.log_btn.clicked.connect(self.logWindow) #로그 화면 연결
        self.as_btn.clicked.connect(lambda: DaesungFunctions.openUrl(self))
        self.exit_btn.clicked.connect(self.exitWindow)
    
    def hwConnect(self):
        try:
            self.set_win = SetWindow()
            self.connectScanner() #스캐너 연결
            self.connectPLC() #PLC 연결
            self.connectPrinter() #바코드 프린터 연결
            if PROC_CODE == '0110': self.connectLight() #경광등 연결
        except: logging.debug("hwConnect : 실패")
    
    def connectScanner(self):
        global scanner_socket
        #-------------------------------------------------------------
        try: self.scanner_th.terminate()
        except: pass
        try: self.sc_con_th.terminate()
        except: pass
        try: scanner_socket.close()
        except: pass
        #-------------------------------------------------------------
        if self.set_win.scanner_check.isChecked():
            try:
                self.scanner_ip = self.set_win.scanner_ip_input.text()
                self.scanner_port = int(self.set_win.scanner_port_input.text())
                #-------------------------------------------------------------
                scanner_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                scanner_socket.settimeout(0.5)
                scanner_socket.connect((self.scanner_ip, self.scanner_port))
                logging.debug("connectScanner : 스캐너 연결 성공")
                #-------------------------------------------------------------
                self.scanner_th = ScannerThread()
                self.scanner_th.sig_data.connect(self.StartData)
                self.scanner_th.start()
                #-------------------------------------------------------------
                logging.debug("connectScanner : scanner_th 성공")
                self.scanner_btn.setStyleSheet("background-color: #55cba7;") #green
            except:
                logging.debug("connectScanner : 스캐너 연결 실패")
                self.scanner_btn.setStyleSheet("background-color: #fd97a5;") #red
                try:
                    self.sc_con_th = SocketThread(self.scanner_ip, self.scanner_port)
                    self.sc_con_th.sig_data.connect(self.ScannerConSlot)
                    self.sc_con_th.start()
                except: logging.debug("connectScanner : sc_con_th 실패")
        else:
            logging.debug("connectScanner : 스캐너 비활성")
            self.scanner_btn.setStyleSheet("background-color: #CDCDCD;") #gray
    
    def connectPLC(self):
        global plc_socket
        #-------------------------------------------------------------
        try: self.plc_th.terminate()
        except: pass
        try: self.plc_con_th.terminate()
        except: pass
        try: plc_socket.close()
        except: pass
        #-------------------------------------------------------------
        if self.set_win.plc_check.isChecked():
            try:
                self.plc_ip = self.set_win.plc_ip_input.text()
                self.plc_port = int(self.set_win.plc_port_input.text())
                self.plc_comp = self.set_win.plc_comp_input.text()
                self.plc_addr = [self.set_win.plc_lenx_input.text(), self.set_win.plc_widx_input.text(), self.set_win.plc_tikx_input.text(), self.set_win.plc_edgeC_input.text(),
                                self.set_win.plc_edge1_input.text(), self.set_win.plc_edge2_input.text(), self.set_win.plc_holeP_input.text(), self.set_win.plc_holeF_input.text(), self.set_win.plc_no_input.text()]
                if PROC_CODE == '0110': self.plc_read = [280, 300]
                else: self.plc_read = ''
                #-------------------------------------------------------------
                plc_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                plc_socket.settimeout(0.5)
                plc_socket.connect((self.plc_ip, self.plc_port))
                try:
                    self.plc_th = PlcCompThread(self.plc_comp, self.plc_read)
                    self.plc_th.sig_data.connect(self.PLCCompSlot)
                    if PROC_CODE == '0110': self.plc_th.read_data.connect(self.PLCReadSlot)
                    self.plc_th.start()
                except: logging.debug("connectPLC : plc_th 실패")
                self.PLC_flag = "success"
                logging.debug("connectPLC : PLC 연결 성공")
                self.PLC_btn.setStyleSheet("background-color: #55cba7;") #green
            except:
                self.PLC_flag = "failed"
                logging.debug("connectPLC : PLC 연결 실패")
                self.PLC_btn.setStyleSheet("background-color: #fd97a5;") #red
                try:
                    self.plc_con_th = SocketThread(self.plc_ip, self.plc_port)
                    self.plc_con_th.sig_data.connect(self.PLCConSlot)
                    self.plc_con_th.start()
                except: logging.debug("connectPLC : PLC_con_th 실패")
                self.PLCReadSlot([99999])
        else:
            self.PLC_flag = "unable"
            logging.debug("connectPLC : PLC 비활성")
            self.PLC_btn.setStyleSheet("background-color: #CDCDCD;") #gray
            self.PLCReadSlot([99999])
    
    def connectPrinter(self):
        try: self.p_con_th.close()
        except: pass
        #-------------------------------------------------------------
        if self.set_win.printer_check.isChecked():
            try:
                self.printer_ip = self.set_win.printer_ip_input.text()
                self.printer_port = int(self.set_win.printer_port_input.text())
                self.mode =  '^' + self.set_win.printer_mode_input.text().replace(' ', '')
                #-------------------------------------------------------------
                print_socket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
                print_socket.settimeout(0.5)
                print_socket.connect((self.printer_ip, self.printer_port))
                print_socket.close()
                #-------------------------------------------------------------
                self.printer_flag = "success"
                logging.debug("connectPrinter : 프린트 연결 성공")
                self.Print_btn.setStyleSheet("background-color: #55cba7;") #green
            except:
                self.printer_flag = "failed"
                logging.debug("connectPrinter : 프린트 연결 실패")
                self.Print_btn.setStyleSheet("background-color: #fd97a5;") #red
                try:
                    self.p_con_th = SocketThread(self.printer_ip, self.printer_port)
                    self.p_con_th.sig_data.connect(self.PrinterConSlot)
                    self.p_con_th.start()
                except: logging.debug("connectPrinter : printer_con_th 실패")
        else:
            self.printer_flag = "unable"
            logging.debug("connectPrinter : 프린트 비활성")
            self.Print_btn.setStyleSheet("background-color: #CDCDCD;") #gray
    
    def connectLight(self):
        try: self.light_ser.close()
        except: pass
        #-------------------------------------------------------------
        if self.set_win.light_check.isChecked():
            try:
                light_port = self.set_win.light_port_input.text()
                light_rate = int(self.set_win.light_rate_input.text())
                self.light_ser = serial.Serial(light_port, light_rate, timeout = 0.5)
                logging.debug("connectLight : 경광등 연결 성공")
                self.light_btn.setStyleSheet("background-color: #55cba7;") #green
            except: 
                logging.debug("connectLight : 경광등 연결 실패")
                self.light_btn.setStyleSheet("background-color: #fd97a5;") #red
        else:
            logging.debug("connectLight : 경광등 비활성")
            self.light_btn.setStyleSheet("background-color: #CDCDCD;") #gray
    
    #---------------------------------------------------------------------------------------------------
    def DBload(self):
        self.tableWidget.setRowCount(0)
        self.tableWidget.clearContents()
        #----------------------------------------------------------------------------
        A_rows = DaesungQuery.selectAutoList(self, PROC_CODE, LINE_FLAG, self.s_date)
        #----------------------------------------------------------------------------
        if A_rows == 'failed': self.connectDBThread()
        elif A_rows == ():
            self.tableWidget.setRowCount(1)
            self.tableWidget.setRowHeight(0, 100)
            self.tableWidget.setSpan(0, 0, 1, 10)
            item_data = QTableWidgetItem("해당일자의 데이터가 없습니다.")
            item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
            self.tableWidget.setItem(0, 0, item_data)
            self.DB_btn.setStyleSheet("background-color: #55cba7;") #green
        else:
            try:
                self.tableWidget.setRowCount(len(A_rows))
                for i in range(len(A_rows)):
                    self.tableWidget.setRowHeight(i, 100)
                    for count, j in enumerate(['SEQ', 'LOT_NUMB', 'ITEM_TEXT', 'LENX', 'SPCL_NAME', 'EDGE_NAME', 'CAL_HOLE_VALUE', 'QTY', 'BUTTON', 'BAR_CODE', 'MES_FLAG']):
                        if j == 'BUTTON':
                            buttons = QPushButton("삭제")
                            buttons.setStyleSheet("background:#fff8f2;""color:#f7932f;""border:2px solid #f7932f;""border-radius:5px;""font:bold 25px;")
                            buttons.setFont(QtGui.QFont("맑은 고딕"))
                            self.state_group.addButton(buttons, i)
                            self.tableWidget.setCellWidget(i, count, buttons)
                        else:
                            print_data = A_rows[i][j]
                            if print_data == None: print_data = '-'
                            elif j == 'LOT_NUMB': print_data = print_data + '\n' + A_rows[i]['LK_JAKUP_SEQ']
                            elif j == 'LENX': print_data = str(int(print_data)) + " x " + str(int(A_rows[i]['WIDX']))
                            elif j == 'CAL_HOLE_VALUE': print_data = int(print_data)
                            elif j == 'QTY': print_data = '{0}/{1}'.format(int(print_data), int(A_rows[i]['BAR_CODE'][20:24]))
                            item_data = QTableWidgetItem(str(print_data))
                            item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                            self.tableWidget.setItem(i, count, item_data)
                self.DB_btn.setStyleSheet("background-color: #55cba7;") #green
            except: logging.debug("DBload : table 로드 실패")
        DaesungFunctions.tableWidth(self, 'Auto', '', '')
    
    def PressedEnterKey(self):
        QR_CODE = self.QR_INPUT.text().replace('\x02','').replace(" ", '') #공백 제거
        self.StartData(QR_CODE)
    
    @pyqtSlot(str)
    def StartData(self, QR_CODE):
        try: self.plc_write_th.terminate()
        except: pass
        #----------------------------------------------------------------------------
        logging.debug('StartData : %s'%QR_CODE)
        #----------------------------------------------------------------------------
        if len(QR_CODE) < 35: QR_rows = DaesungQuery.selectCNClabel_REG(self, PROC_CODE, QR_CODE[9:21], QR_CODE[21:25], QR_CODE[25:])
        else: QR_rows = DaesungQuery.selectCNClabel(self, PROC_CODE, QR_CODE)
        #----------------------------------------------------------------------------
        if QR_rows == 'failed': self.connectDBThread()
        elif QR_rows != ():
            try:
                LENX, WIDX, TIKX = int(QR_rows[0]['LENX']), int(QR_rows[0]['WIDX']), int(QR_rows[0]['TIKX'])
                JAKUP_FLAG = QR_rows[0]['NO_YN']
                if JAKUP_FLAG == 'Y': raise('WIDX')
                try:
                    JAKUP_NO, JAKUP_SEQ = QR_rows[0]['REG_NO'], QR_rows[0]['REG_SEQ']
                    JAKUP_SORT_KEY = QR_rows[0]['SORT_KEY']
                    SPCL_CO, SPCL = QR_rows[0]['SPCL_CODE'], QR_rows[0]['SPCL_NAME']
                    H_FLAG, H_VALUE = QR_rows[0]['HOLE_FLAG'], QR_rows[0]['HOLE_VALUE']
                    HOLE_V = QR_rows[0]['CAL_HOLE_VALUE']
                    CONN_CPROC_CODE, CPROC_CODE = QR_rows[0]['CONN_CPROC_CODE'].split(','), ''
                    for C, CODE in enumerate(CONN_CPROC_CODE):
                        if C + 1 == len(CONN_CPROC_CODE): CPROC_CODE = CPROC_CODE + "'%s'"%CODE.replace(' ', '').replace("'", '').replace('"', '')
                        else: CPROC_CODE = CPROC_CODE + "'%s', "%CODE.replace(' ', '').replace("'", '').replace('"', '')
                    #----------------------------------------------------------------------------
                    H_rows = DaesungQuery.selectHoleFlag(self, CPROC_CODE)
                    #----------------------------------------------------------------------------
                    if HOLE_V == None or HOLE_V <= 0: HOLE_F, HOLE_V = 0, 0
                    elif TIKX != 36 or H_rows != None: HOLE_F, HOLE_V = 0, 0
                    elif H_FLAG == '3' and 0 < H_VALUE < 700:
                        HOLE_F, HOLE_V = 0, 0
                        MessageWindow(self, "타공값 700 이상").showModal()
                    else: HOLE_F, HOLE_V = 1, int(HOLE_V)
                    #-------------------------------------------------------------
                    if QR_rows[0]['EDGE_FLAG']  == '1': EDGE, EDGE1, EDGE2 = '일면', 1, 0
                    elif QR_rows[0]['EDGE_FLAG']  == '2': EDGE, EDGE1, EDGE2 = '일면2', 0, 1
                    elif QR_rows[0]['EDGE_FLAG']  == '3': EDGE, EDGE1, EDGE2 = '양면', 1, 1
                    else: EDGE, EDGE1, EDGE2 = '-', 0, 0
                    #-------------------------------------------------------------
                    self.QR_INPUT.setText(QR_CODE)
                    self.LENX_INPUT.setText(str(LENX))
                    self.WIDX_INPUT.setText(str(WIDX))
                    self.SPCL_INPUT.setText(str(SPCL))
                    self.EDGE_INPUT.setText(EDGE)
                    #----------------------------------------------------------------------------
                    QR_CODE = QR_rows[0]['BAR_CODE']
                    B_rows = DaesungQuery.selectJakupData(self, '%', QR_CODE)
                    if B_rows == None or B_rows['MES_FLAG'] == LINE_FLAG or B_rows['PUT_FLAG'] == 'S' or B_rows['PUT_FLAG'] == 'F':
                        result = DaesungQuery.PR_SAVE_MAKE_BAR(self, 'insert', '0', EMPL_CODE, JAKUP_NO, JAKUP_SEQ, JAKUP_SORT_KEY, QR_CODE, self.c_date, 1, 0)
                        if result == 1: self.result_data.setText("생산실적 완료")
                        else: self.result_data.setText("생산실적 실패")
                        #-------------------------------------------------------------
                        if PROC_CODE == '0110' and self.PLC_flag == "success":
                            SEQ = DaesungQuery.selectEdgeMaxSeq(self, LINE_FLAG, self.s_date[:6])
                            if SEQ['SEQ'] == None: SEQ = 1000
                            else: SEQ = int(SEQ['SEQ']) + 1
                            #-------------------------------------------------------------
                            B_rows = DaesungQuery.selectJakupData(self, LINE_FLAG, QR_CODE)
                            if B_rows == None: DaesungQuery.insertEdgeSeq(self, LINE_FLAG, self.s_date[:6], QR_CODE, SEQ, '0')
                            else: DaesungQuery.updateEdgeSeq(self, "PUT_FLAG = '0', SEQ = '{0}'".format(SEQ), self.s_date[:6], LINE_FLAG, QR_CODE, B_rows['SEQ'])
                            #-------------------------------------------------------------
                            self.plc_write_th = PlcWriteThread([LENX, WIDX, TIKX], self.plc_addr) #PLC에 값 전송
                            self.plc_write_th.sig_data.connect(self.PLCWriteSlot)
                            self.plc_write_th.start()
                        elif PROC_CODE == '0115':
                            if self.PLC_flag == "success":
                                if (SPCL_CO in self.edge_code): EDGE_C = int(self.edge_code.index(SPCL_CO)) + 1
                                elif SPCL_CO == None or SPCL_CO == '': EDGE_C = 0
                                else: EDGE_C = 7
                                #-------------------------------------------------------------
                                B_rows = DaesungQuery.selectJakupData(self, LINE_FLAG, QR_CODE)
                                if B_rows == None:
                                    SEQ = DaesungQuery.selectEdgeMaxSeq(self, LINE_FLAG, self.s_date[:6])
                                    if SEQ['SEQ'] == None: SEQ = 1000
                                    else: SEQ = int(SEQ['SEQ']) + 1
                                    DaesungQuery.insertEdgeSeq(self, LINE_FLAG, self.s_date[:6], QR_CODE, SEQ, '0')
                                else: SEQ = int(B_rows['SEQ'])
                                self.plc_write_th = PlcWriteThread([LENX, WIDX, TIKX, EDGE_C, EDGE1, EDGE2, HOLE_V, HOLE_F, SEQ], self.plc_addr) #PLC에 값 전송
                                self.plc_write_th.sig_data.connect(self.PLCWriteSlot)
                                self.plc_write_th.start()
                            if self.printer_flag == "success": self.printCNCLabel(QR_CODE) #라벨 발행
                    else: self.result_data.setText("이미 생산된 제품입니다.")
                except:
                    logging.debug("StartData : QR_SQL 오류")
                    self.result_data.setText("QR값 오류")
            except:
                logging.debug("StartData : 작업 불가 조건")
                self.result_data.setText("작업 불가 조건")
        else:
            logging.debug("StartData : 바코드 없음")
            self.result_data.setText("바코드 없음")
        self.DBload() #DB로드
        self.format_th.start()
    
    #라벨 발행
    def printCNCLabel(self, QR_CODE):
        try:
            self.print_socket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
            self.print_socket.settimeout(0.5)
            self.print_socket.connect((self.printer_ip, self.printer_port))
            logging.debug("printCNCLabel : 프린터 연결 성공")
            self.Print_btn.setStyleSheet("background-color: #55cba7;") #green
            #----------------------------------------------------------------------------
            QR_rows = DaesungQuery.selectCNClabel(self, PROC_CODE, QR_CODE)
            #----------------------------------------------------------------------------
            if QR_rows == 'failed': self.connectDBThread()
            elif QR_rows != []:
                try:
                    f_name = open(c_label, 'r', encoding = 'utf-8')
                    textData = f_name.read()
                    f_name.close()
                    if self.set_win.printer_mode_check.isChecked() == True: textData = textData.replace("^CI28", "^CI28" + self.mode)
                    if self.set_win.printer_po_check.isChecked() == True: textData = textData.replace("^LS0", "^LS0^POI")
                    elif self.set_win.printer_po_check.isChecked() == False: textData = textData.replace("^LS0", "^LS0^PON")
                    for i in ['BAR_CODE', 'LOT_NUMB', 'REG_SEQ', 'REG_DATE', 'HOPE_DATE', 'BUYER_NAME', 'ITEM_MA_NAME', 'ITEM_NAME', 'LENX', 'WIDX', 'TIKX', 'SPCL_NAME', 'EDGE_NAME', 'GLAS_NAME', 'CONN_CPROC_NAME', 'CAL_HOLE_VALUE', 'BIGO', 'CPROC_BIGO', 'LABEL_BIGO', 'TRANS_FLAG_NAME', 'FSET_FLAG_NAME', 'CONN_CPROC_NAME_BIGO']:
                        if i == 'CONN_CPROC_NAME_BIGO':
                            #----------------------------------------------------------------------------
                            CONN_CPROC_NAME_BIGO = DaesungQuery.selectConnBigo(self, QR_rows[0]['REG_NO'], QR_rows[0]['REG_SEQ'])
                            #----------------------------------------------------------------------------
                            print_data = CONN_CPROC_NAME_BIGO[0]['CONN_CPROC_NAME_BIGO']
                        else:
                            print_data = QR_rows[0][i]
                            if print_data == None: print_data = ""
                            elif i == 'REG_DATE' or i == 'HOPE_DATE': print_data = "{0}.{1}.{2}".format(print_data[2:4], print_data[4:6], print_data[6:8])
                            elif i == 'LENX' or i == 'WIDX' or i == 'TIKX': print_data = int(print_data)
                            elif i == 'CAL_HOLE_VALUE': print_data = str(int(print_data))
                            elif i == 'BAR_CODE':
                                if QR_rows[0]['CAL_HOLE_VALUE'] == None: h_value = '0000'
                                else:
                                    h_value = str(int(QR_rows[0]['CAL_HOLE_VALUE']))
                                    if len(h_value) == 3: h_value = '0' + h_value
                                print_data = print_data[:33] + h_value + QR_rows[0]['EDGE_FLAG']
                        textData = textData.replace("{%s}"%i, str(print_data))
                    self.print_socket.send(textData.encode())
                except: logging.debug("printCNCLabel : selectCNClabel 실패")
            self.print_socket.close()
        except: self.connectPrinter() #바코드 프린터 연결
    
    #전표 취소
    def itemCancel(self, row):
        QR_CODE = self.tableWidget.item(row, 9).text()
        #----------------------------------------------------------------------------
        QR_rows = DaesungQuery.selectCNClabel(self, PROC_CODE, QR_CODE)
        #----------------------------------------------------------------------------
        if QR_rows == 'failed':self.connectDBThread()
        else:
            JAKUP_NO = QR_rows[0]['REG_NO']
            JAKUP_SEQ = QR_rows[0]['REG_SEQ']
            JAKUP_SORT_KEY = QR_rows[0]['SORT_KEY']
            #----------------------------------------------------------------------------
            result = DaesungQuery.PR_SAVE_MAKE_BAR(self, 'insert', '0', EMPL_CODE, JAKUP_NO, JAKUP_SEQ, JAKUP_SORT_KEY, QR_CODE, self.c_date, 1, 1) #불량수량 등록
            #----------------------------------------------------------------------------
            if result == 1: r_text = '품목 삭제 완료'
            else: r_text = '품목 삭제 실패'
            logging.debug("itemCancel : %s"%r_text)
            self.result_data.setText(r_text)
            self.DBload() #DB로드
        self.format_th.start()
    
    #----------------------------------------------------------------------------
    @pyqtSlot(list)
    def PLCReadSlot(self, data):
        if data == [99999]:
            for i in range(2):
                for count in range(4): self.plc_tableWidget.setItem(i, count + 1, QTableWidgetItem(''))
        else:
            for i, d_data in enumerate(data):
                for count, h_data in enumerate(d_data):
                    print_data = float(int((h_data[2:4] + h_data[:2]), 16)/10)
                    item_data = QTableWidgetItem(str(print_data))
                    if i == 0: item_data.setForeground(QBrush(QColor(66, 188, 112))) #green
                    else: item_data.setForeground(QBrush(QColor(63, 139, 204))) #blue
                    item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                    self.plc_tableWidget.setItem(i, count + 1, item_data)
        for count, t in enumerate([423, 225, 225, 225]): self.plc_tableWidget.setColumnWidth(count, t)
    
    @pyqtSlot(int)
    def PLCCompSlot(self, data):
        if data == 99999 and data != self.c_data:
            try: self.plc_write_th.terminate()
            except: pass
            self.c_data = data
            self.connectPLC() #PLC 연결
        elif data == 1 and data != self.c_data:
            try: self.plc_write_th.terminate()
            except: pass
            logging.debug("PLCCompSlot : 완료신호 1")
            self.c_data = data
            self.clear_th.start()
        else: self.c_data = data
    
    @pyqtSlot(int)
    def PLCWriteSlot(self, num):
        if num == 99999: self.PLC_btn.setStyleSheet("background-color: #fd97a5;") #red
        else: self.PLC_btn.setStyleSheet("background-color: #55cba7;") #green
    
    #----------------------------------------------------------------------------
    @pyqtSlot(str)
    def ScannerConSlot(self, state):
        if state == "success": self.connectScanner()
        elif state == "failed": self.scanner_btn.setStyleSheet("background-color: #fd97a5;") #red
    
    @pyqtSlot(str)
    def PLCConSlot(self, state):
        if state == "success": self.connectPLC()
        elif state == "failed": self.PLC_btn.setStyleSheet("background-color: #fd97a5;") #red
    
    @pyqtSlot(str)
    def PrinterConSlot(self, state):
        if state == "success": self.connectPrinter()
        elif state == "failed": self.Print_btn.setStyleSheet("background-color: #fd97a5;") #red
    
    @pyqtSlot(int)
    def FormatSlot(self, num):
        if num == 1:
            self.QR_INPUT.setText("")
            self.LENX_INPUT.setText("")
            self.WIDX_INPUT.setText("")
            self.SPCL_INPUT.setText("")
            self.EDGE_INPUT.setText("")
            self.result_data.setText("")
        elif num == 2: self.QR_INPUT.setFocus()
        elif num == 3:
            try: self.light_ser.write('RY 1 0\r'.encode()) #green light
            except: pass
        elif num == 5:
            self.plc_write_th = PlcWriteThread([0, 0, 0, 0, 0, 0, 0, 0, 0], self.plc_addr) #PLC값 초기화
            self.plc_write_th.sig_data.connect(self.PLCWriteSlot)
            self.plc_write_th.start()
    
    def edgeCodeLoad(self):
        self.edge_code = []
        load_wb = load_workbook("엣지코드_리스트.xlsx", data_only = True)
        load_ws = load_wb['Sheet1']
        get_cells = load_ws['C3':'C9']
        for row in get_cells:
            for cell in row: self.edge_code.append(cell.value)
    
    def edgeCode(self):
        try: 
            EdgeCodeWindow().showModal()
            self.edgeCodeLoad()
        except: logging.debug("edgeCode : 연결 실패")
    
    #----------------------------------------------------------------------------    
    #날짜 출력
    def showDate(self, date):
        self.calender.hide()
        a = date.toString().split()
        if len(a[1]) == 1: a[1] = "0" + a[1]
        if len(a[2]) == 1: a[2] = "0" + a[2]
        self.date_btn.setText("%s-%s-%s "%(a[3],a[1],a[2]))
        self.calendar_flag = False
        DaesungFunctions.replaceDate(self) #DB로드
    
    def mousePressEvent(self, event):
        if event.buttons() == Qt.LeftButton and self.calendar_flag == True:
            self.calender.hide()
            self.calendar_flag = False
    
    def setData(self):
        try:
            SetWindow().showModal()
            self.hwConnect() #HW연결
        except: logging.debug("setData : 연결 실패")
    
    def logWindow(self):
        try: MesLogWindow(TodayData).showModal()
        except: logging.debug("logWindow : 로그 화면 연결 실패")
    
    #실시간 DB연결
    def connectDBThread(self):
        self.DB_btn.setStyleSheet("background-color: #fd97a5;") #red
        if self.DB_flag == 0:
            self.DB_flag = 1
            logging.debug("DBload : DB로드 실패")
            MessageWindow(self, "DB연결 실패").showModal()
            #----------------------------------------------------------------------------
            self.DB_th = ConnectDBThread()
            self.DB_th.sig_data.connect(self.DbThreadSlot)
            self.DB_th.start()
    
    @pyqtSlot(int)       
    def DbThreadSlot(self, con):
        if con == 1:
            self.DB_th.terminate()
            self.DB_flag = 0
            #----------------------------------------------------------------------------
            result = DaesungQuery.connectDB(self, host, port, user, name)
            #----------------------------------------------------------------------------
            if result == 'success':
                self.hwConnect() #HW연결
                DaesungFunctions.replaceDate(self) #DB로드
    
    def exitWindow(self):
        MessageWindow(self, 2).showModal()
        #----------------------------------
        if RECONNECT_FLAG == 1:
            try: self.scanner_th.terminate()
            except: pass
            try: self.sc_con_th.terminate()
            except: pass
            try: scanner_socket.close()
            except: pass
            #------------------------------------------
            try: self.plc_th.terminate()
            except: pass
            try: self.plc_con_th.terminate()
            except: pass
            try: plc_socket.close()
            except: pass
            #------------------------------------------
            try: self.p_con_th.close()
            except: pass
            try: self.light_ser.close()
            except: pass
            #------------------------------------------
            try: self.focus_th.terminate()
            except: pass
            try: self.format_th.terminate()
            except: pass
            try: self.light_th.terminate()
            except: pass

################################################################################################################
#자동화(접착, 테노너) Window > 엣지코드리스트 Window
class EdgeCodeWindow(QDialog):
    def __init__(self):
        super(EdgeCodeWindow, self).__init__()
        loadUi("ui\EdgeCodeList.ui", self)
        
        self.setFixedSize(722, 641)
        self.edit_name = [self.edge_1_input, self.edge_2_input, self.edge_3_input, self.edge_4_input, self.edge_5_input, self.edge_6_input, self.edge_7_input]
        self.edgeCodeLoad()
        
        DaesungFunctions.clickable(self, self.edge_1_input, self.edge_1_input)
        DaesungFunctions.clickable(self, self.edge_2_input, self.edge_2_input)
        DaesungFunctions.clickable(self, self.edge_3_input, self.edge_3_input)
        DaesungFunctions.clickable(self, self.edge_4_input, self.edge_4_input)
        DaesungFunctions.clickable(self, self.edge_5_input, self.edge_5_input)
        DaesungFunctions.clickable(self, self.edge_6_input, self.edge_6_input)
        DaesungFunctions.clickable(self, self.edge_7_input, self.edge_7_input)
        
        self.one.clicked.connect(lambda state, button = self.one : DaesungFunctions.NumClicked(self, state, button))
        self.two.clicked.connect(lambda state, button = self.two : DaesungFunctions.NumClicked(self, state, button))
        self.three.clicked.connect(lambda state, button = self.three : DaesungFunctions.NumClicked(self, state, button))
        self.four.clicked.connect(lambda state, button = self.four : DaesungFunctions.NumClicked(self, state, button))
        self.five.clicked.connect(lambda state, button = self.five : DaesungFunctions.NumClicked(self, state, button))
        self.six.clicked.connect(lambda state, button = self.six : DaesungFunctions.NumClicked(self, state, button))
        self.seven.clicked.connect(lambda state, button = self.seven : DaesungFunctions.NumClicked(self, state, button))
        self.eight.clicked.connect(lambda state, button = self.eight : DaesungFunctions.NumClicked(self, state, button))
        self.nine.clicked.connect(lambda state, button = self.nine : DaesungFunctions.NumClicked(self, state, button))
        self.zero.clicked.connect(lambda state, button = self.zero : DaesungFunctions.NumClicked(self, state, button))
        self.dot.clicked.connect(lambda state, button = self.dot : DaesungFunctions.NumClicked(self, state, button))
        self.backspace.clicked.connect(lambda: DaesungFunctions.NumDeleted(self))
        
        self.save_btn.clicked.connect(self.saveData)
        self.cancel_btn.clicked.connect(lambda: DaesungFunctions.closeWindow(self))
    
    def edgeCodeLoad(self):
        try:
            self.load_wb = load_workbook("엣지코드_리스트.xlsx", data_only = True)
            self.load_ws = self.load_wb['Sheet1']
            get_cells = self.load_ws['C3':'C9']
            for count, row in enumerate(get_cells):
                for cell in row:
                    if cell.value == None: data = ''
                    else: data = cell.value
                    self.edit_name[count].setText(data)
        except: MessageWindow(self, "엣지코드_리스트.xlsx를 찾을 수 없습니다.").showModal()
    
    def saveData(self):
        try:
            for count, i in enumerate(self.edit_name): self.load_ws.cell(row = count + 3, column = 3, value = i.text())
            self.load_wb.save("엣지코드_리스트.xlsx")
            logging.debug("saveData : 엣지코드_리스트 저장 성공")
            DaesungFunctions.closeWindow(self)
        except: MessageWindow(self, "데이터를 저장할 수 없습니다.").showModal()
    
    def showModal(self):
        return super().exec_()

################################################################################################################
#실시간 DB연결 및 rowCount 확인
class reloadThread(QThread):
    sig_data = pyqtSignal(int)

    def run(self):
        while True:
            self.sig_data.emit(1)
            time.sleep(1)

class PlcEdgeThread(QThread):
    sig_data = pyqtSignal(list)
    
    def __init__(self, addr1, addr2):
        super().__init__()
        self.addr1 = addr1
        self.addr2 = addr2
    
    def run(self):
        LENGTH = RS232Addr('L', self.addr1)
        s_data = bytes.fromhex(COMPANY + LENGTH[0] + POSTION + '54 00 02 00' + BLOCK + LENGTH[1] + '25 44 57' + LENGTH[2])
        LENGTH2 = RS232Addr('L', self.addr2)
        s_data2 = bytes.fromhex(COMPANY + LENGTH2[0] + POSTION + '54 00 02 00' + BLOCK + LENGTH2[1] + '25 44 57' + LENGTH2[2])
        while True:
            try:
                #----------------------------------------------------------------------------
                plc_socket.send(s_data)
                #----------------------------------------------------------------------------
                re = plc_socket.recv(1024).hex()
                edge1 = int((re[-2:] + re[-4:-2]), 16)
                #----------------------------------------------------------------------------
                plc_socket.send(s_data2)
                #----------------------------------------------------------------------------
                re = plc_socket.recv(1024).hex()
                edge2 = int((re[-2:] + re[-4:-2]), 16)
                self.sig_data.emit([edge1, edge2])
            except: self.sig_data.emit([99999, 99999])
            time.sleep(0.2)

class SensorThread(QThread):
    sig_data = pyqtSignal(str)
    
    def __init__(self, ser):
        super().__init__()
        self.ser = ser
    
    def run(self):
        while True:
            try:
                self.ser.write("IN\r".encode())
                data = self.ser.readline()
                data = data.decode()[:len(data)-1]
                self.sig_data.emit(data)
            except: self.sig_data.emit("failed")
            time.sleep(0.5)

#---------------------------------------------------------------------------------------------------------------
#자동화(엣지) Window
class MesEdgeWindow(QDialog):
    def __init__(self, date):
        super(MesEdgeWindow, self).__init__()
        loadUi("ui\DAESUNG_MES_E.ui", self)
        
        DaesungFunctions.setAutoStyle(self, date, PROC_CODE, PROC_NAME) #기본 셋팅
        
        self.hwConnect() #HW연결
        DaesungFunctions.replaceDate(self) #DB로드
        try:
            self.light_th = FormatThread("light")
            self.light_th.sig_data.connect(self.FormatSlot)
            #-------------------------------------------------------------
            self.format_th = FormatThread("format")
            self.format_th.sig_data.connect(self.FormatSlot)
            #-------------------------------------------------------------
            self.dorna_th = FormatThread("dorna")
            self.dorna_th.sig_data.connect(self.FormatSlot)
        except: pass
        
        self.sensor_btn.clicked.connect(lambda: self.clickedSensor('0100'))
        
        self.test_btn.clicked.connect(self.testMode)
        self.state_group.buttonClicked[int].connect(self.item1Cancel) #엣지1 취소
        self.state_group2.buttonClicked[int].connect(self.item2Cancel) #엣지2 취소

        # 상단버튼 --------------------------------------------------
        self.date_btn.clicked.connect(lambda: DaesungFunctions.calendar(self))
        self.set_btn.clicked.connect(self.setData)
        self.log_btn.clicked.connect(self.logWindow)
        self.as_btn.clicked.connect(lambda: DaesungFunctions.openUrl(self))
        self.exit_btn.clicked.connect(self.exitWindow)
    
    def hwConnect(self):
        try:
            self.set_win = SetWindow()
            self.connectPLC() #PLC 연결
            self.connectSensor() #센서 연결
            self.connectPrinter() #바코드 프린터 연결
        except: logging.debug("hwConnect : 실패")
    
    def connectPLC(self):
        global plc_socket
        #-------------------------------------------------------------
        try: self.edge_th.terminate()
        except: pass
        try: self.plc_con_th.terminate()
        except: pass
        try: plc_socket.close()
        except: pass
        #-------------------------------------------------------------
        self.e1_data, self.e2_data = 0, 0
        if self.set_win.plc_check.isChecked():
            try:
                self.plc_ip = self.set_win.plc_ip_input.text()
                self.plc_port = int(self.set_win.plc_port_input.text())
                self.plc_edge1, self.plc_edge2 = self.set_win.plc_edgeR1_input.text(), self.set_win.plc_edgeR2_input.text()
                #-------------------------------------------------------------
                plc_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                plc_socket.settimeout(0.5)
                plc_socket.connect((self.plc_ip, self.plc_port))
                try:
                    self.edge_th = PlcEdgeThread(self.plc_edge1, self.plc_edge2)
                    self.edge_th.sig_data.connect(self.PLCEdgeSlot)
                    self.edge_th.start()
                except: logging.debug("connectPLC : edge_th 실패")
                self.PLC_flag = "success"
                logging.debug("connectPLC : PLC 연결 성공")
                self.PLC_btn.setStyleSheet("background-color: #55cba7;") #green
            except:
                self.PLC_flag = "failed"
                logging.debug("connectPLC : PLC 연결 실패")
                self.PLC_btn.setStyleSheet("background-color: #fd97a5;") #red
                try:
                    self.plc_con_th = SocketThread(self.plc_ip, self.plc_port)
                    self.plc_con_th.sig_data.connect(self.PLCConSlot)
                    self.plc_con_th.start()
                except: logging.debug("connectPLC : PLC_con_th 실패")
        else:
            self.PLC_flag = "unable"
            logging.debug("connectPLC : PLC 비활성")
            self.PLC_btn.setStyleSheet("background-color: #CDCDCD;") #gray
    
    def connectSensor(self):
        try: self.sensor_th.terminate()
        except: pass
        try: self.sensor_con_th.terminate()
        except: pass
        try: self.ser.close()
        except: pass
        #-------------------------------------------------------------
        self.es1_data, self.s_data, self.es2_data = 0, 0, 0
        if self.set_win.sensor_check.isChecked():
            try:
                sensor_port = self.set_win.sensor_port_input.text()
                sensor_rate = int(self.set_win.sensor_rate_input.text())
                #-------------------------------------------------------------
                self.ser = serial.Serial(sensor_port, sensor_rate, timeout = 0.5)
                self.sensor_th = SensorThread(self.ser)
                self.sensor_th.sig_data.connect(self.Edge1SensorSlot)
                self.sensor_th.sig_data.connect(self.SensorSlot)
                self.sensor_th.sig_data.connect(self.Edge2SensorSlot)
                self.sensor_th.start()
                #-------------------------------------------------------------
                self.sensor_flag = "success"
                logging.debug("connectSensor : 센서 연결 성공")
                self.sensor_btn.setStyleSheet("background-color: #55cba7;") #green
                self.light_btn.setStyleSheet("background-color: #55cba7;") #green
            except:
                self.sensor_flag = "failed"
                logging.debug("connectSensor : 센서 연결 실패")
                self.sensor_btn.setStyleSheet("background-color: #fd97a5;") #red
                self.light_btn.setStyleSheet("background-color: #fd97a5;") #red
                try:
                    self.sensor_con_th = SerialThread(sensor_port, sensor_rate)
                    self.sensor_con_th.sig_data.connect(self.SensorConSlot)
                    self.sensor_con_th.start()
                except: logging.debug("connectSensor : sensor_con_th 실패")
        else:
            self.sensor_flag = "unable"
            logging.debug("connectSensor : 센서 비활성")
            self.sensor_btn.setStyleSheet("background-color: #CDCDCD;") #gray
            self.light_btn.setStyleSheet("background-color: #CDCDCD;") #gray
    
    def connectPrinter(self):
        try: self.prt_con_th.close()
        except: pass
        #-------------------------------------------------------------
        if self.set_win.printer_check.isChecked():
            try:
                self.printer_ip = self.set_win.printer_ip_input.text()
                self.printer_port = int(self.set_win.printer_port_input.text())
                self.mode =  '^' + self.set_win.printer_mode_input.text().replace(' ', '')
                #-------------------------------------------------------------
                print_socket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
                print_socket.settimeout(0.5)
                print_socket.connect((self.printer_ip, self.printer_port))
                print_socket.close()
                #-------------------------------------------------------------
                self.printer_flag = "success"
                logging.debug("connectPrinter : 프린트 연결 성공")
                self.Print_btn.setStyleSheet("background-color: #55cba7;") #green
            except:
                self.printer_flag = "failed"
                logging.debug("connectPrinter : 프린트 연결 실패")
                self.Print_btn.setStyleSheet("background-color: #fd97a5;") #red
                try:
                    self.prt_con_th = SocketThread(self.printer_ip, self.printer_port)
                    self.prt_con_th.sig_data.connect(self.PrinterConSlot)
                    self.prt_con_th.start()
                except: logging.debug("connectPrinter : prt_con_th 실패")
        else:
            self.printer_flag = "unable"
            logging.debug("connectPrinter : 프린트 비활성")
            self.Print_btn.setStyleSheet("background-color: #CDCDCD;") #gray
    
    #---------------------------------------------------------------------------------------------------
    def DBload(self):
        btn_group = [self.state_group, self.state_group2]
        for t_count, table_name in enumerate([self.tableWidget, self.tableWidget2]):
            table_name.removeRow(0)
            if t_count == 0: FLAG, LIMIT = "'1', '2'", 'LIMIT 2'
            elif t_count == 1: FLAG, LIMIT = "'3', '4'", 'LIMIT 2'
            #----------------------------------------------------------------------------
            A_rows = DaesungQuery.selectAutoEdgeList(self, LINE_FLAG, self.s_date, '%', FLAG, LIMIT)
            #----------------------------------------------------------------------------
            if A_rows == 'failed': self.connectDBThread()
            else:
                if A_rows == () and t_count == 0:
                    table_name.setRowCount(1)
                    table_name.setRowHeight(0, 105)
                    table_name.setSpan(0, 0, 1, 9)
                    item_data = QTableWidgetItem("해당일자의 데이터가 없습니다.")
                    item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                    table_name.setItem(0, 0, item_data)
                    self.noData = 1
                elif A_rows != ():
                    table_name.setRowCount(len(A_rows))
                    for i in range(len(A_rows)):
                        table_name.setRowHeight(i, 105)
                        for count, j in enumerate(['GUBUN', 'LOT_NUMB', 'SEQ', 'ITEM_TEXT', 'SPCL_NAME', 'EDGE_NAME', 'PUT_FLAG', 'BAR_CODE', 'BUTTON']):
                            if j == 'BUTTON':
                                buttons = QPushButton("삭제")
                                buttons.setStyleSheet("background:#fff8f2;""color:#f7932f;""border:2px solid #f7932f;""border-radius:5px;""font:bold 25px;")
                                buttons.setFont(QtGui.QFont("맑은 고딕"))
                                btn_group[t_count].addButton(buttons, i)
                                table_name.setCellWidget(i, count, buttons)
                            else:
                                if j == 'GUBUN' and t_count == 0: print_data = '엣지1'
                                elif j == 'GUBUN' and t_count == 1: print_data = '엣지2'
                                else:
                                    print_data = A_rows[i][j]
                                    if print_data == None: print_data = '-'
                                    elif j == 'LOT_NUMB': print_data = print_data + '\n' + A_rows[i]['LK_JAKUP_SEQ']
                                    elif j == 'ITEM_TEXT':
                                        if A_rows[i]['CAL_HOLE_VALUE'] == None: print_data = print_data
                                        else: print_data = print_data + '/상' + str(int(A_rows[i]['CAL_HOLE_VALUE']))
                                    elif j == 'SPCL_NAME':
                                        if A_rows[i]['SPCL_CODE'] == None: print_data = print_data
                                        else: print_data = print_data + '\n(' + A_rows[i]['SPCL_CODE'] + ')'
                                item_data = QTableWidgetItem(str(print_data))
                                if j == 'GUBUN': item_data.setBackground(QtGui.QColor(241, 241, 241)) #gray
                                elif j == 'PUT_FLAG' and print_data != '대기': item_data.setForeground(QBrush(QColor(247, 49, 49))) #red
                                if j != 'ITEM_TEXT': item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                                table_name.setItem(i, count, item_data)
                self.DB_btn.setStyleSheet("background-color: #55cba7;") #green
                if self.reload_num == 0:
                    try:
                        self.th_rowCount = reloadThread()
                        self.th_rowCount.sig_data.connect(self.rowCountSlot)
                        self.th_rowCount.start()
                        self.reload_num = 1
                    except: logging.debug("DBload : th_rowCount 연결 실패")
            DaesungFunctions.tableWidth(self, 'EDGE', '', '')
    
    def testMode(self):
        c_text = self.test_btn.text()
        if c_text == '테스트 사용':
            self.test_btn.setText('테스트 멈춤')
            self.test_btn.setStyleSheet('background:#fd983c;')
        else:
            self.test_btn.setText('테스트 사용')
            self.test_btn.setStyleSheet('background:#5D76B8;')
    
    #라벨 발행
    def printData(self):
        try:
            if self.test_btn.text() == '테스트 멈춤': QR_CODE = '00120221018005900011000119200620110253'
            else:
                SPCL = self.tableWidget.item(self.tableWidget.rowCount() - 1, 4).text()
                EDGE = self.tableWidget.item(self.tableWidget.rowCount() - 1, 5).text()
                QR_CODE = self.tableWidget.item(self.tableWidget.rowCount() - 1, 7).text()
            try:
                if SPCL.find('(177)') >= 0: self.ser.write('RY 3 0\r'.encode()) #엣지 없음
                elif EDGE == '-' or EDGE == '일면': self.ser.write('RY 3 0\r'.encode()) #엣지 없음
                elif EDGE == '일면2' or EDGE == '양면': self.ser.write('RY 3 1\r'.encode()) #엣지 있음
                self.ser.write('RY 2 1\r'.encode())
                logging.debug("printData : dorna 전송 성공")
                try:
                    self.print_socket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
                    self.print_socket.settimeout(0.5)
                    self.print_socket.connect((self.printer_ip, self.printer_port))
                    logging.debug("AutoLabeler : 프린터 연결 성공")
                    self.Print_btn.setStyleSheet("background-color: #55cba7;") #green
                    try:
                        #----------------------------------------------------------------------------
                        QR_rows = DaesungQuery.selectCNClabel(self, PROC_CODE, QR_CODE)
                        #----------------------------------------------------------------------------
                        if QR_rows == 'failed': self.connectDBThread()
                        elif QR_rows != []:
                            f_name = open(z_label, 'r', encoding = 'utf-8')
                            textData = f_name.read()
                            f_name.close()
                            if self.set_win.printer_mode_check.isChecked() == True: textData = textData.replace("^CI28", "^CI28" + self.mode)
                            if self.set_win.printer_po_check.isChecked() == True: textData = textData.replace("^LS0", "^LS0^POI")
                            elif self.set_win.printer_po_check.isChecked() == False: textData = textData.replace("^LS0", "^LS0^PON")
                            for i in ['BUYER_NAME', 'REG_NO', 'REG_SEQ', 'LOT_NUMB', 'TRANS_FLAG_NAME', 'SPCL_NAME', 'ITEM_NAME', 'LENX', 'WIDX', 'TIKX', 'LW', 'W', 'L', 'KYU', 'CAL_HOLE_VALUE', 'QTY', 'LABEL_BIGO', 'FSET_FLAG_NAME', 'CONN_CPROC_NAME_BIGO']:
                                if i == 'CONN_CPROC_NAME_BIGO':
                                    #----------------------------------------------------------------------------
                                    CONN_CPROC_NAME_BIGO = DaesungQuery.selectConnBigo(self, QR_rows[0]['REG_NO'], QR_rows[0]['REG_SEQ'])
                                    #----------------------------------------------------------------------------
                                    print_data = CONN_CPROC_NAME_BIGO[0]['CONN_CPROC_NAME_BIGO']
                                elif i == 'LW':
                                    if QR_rows[0]['WIDX'] == None: widx = '-'
                                    else: widx = int(QR_rows[0]['WIDX'])
                                    if QR_rows[0]['LENX'] == None: lenx = '-'
                                    else: lenx = int(QR_rows[0]['LENX'])
                                    print_data = str(widx)[:-2].zfill(2) + str(lenx)[:-2].zfill(2)
                                elif i == 'W':
                                    if QR_rows[0]['WIDX'] == None: print_data = ''
                                    else: print_data = int(QR_rows[0]['WIDX']) - 10
                                elif i == 'L':
                                    if QR_rows[0]['LENX'] == None: print_data = ''
                                    else: print_data = int(QR_rows[0]['LENX']) + 10
                                else:
                                    print_data = QR_rows[0][i]
                                    if print_data == None: print_data = ""
                                    elif i == 'LENX' or i == 'WIDX' or i == 'TIKX': print_data = int(print_data)
                                    elif i == 'CAL_HOLE_VALUE': print_data = '(%d)'%int(QR_rows[0]['CAL_HOLE_VALUE'])
                                    elif i == 'QTY': print_data = '{0}/{1}'.format(QR_rows[0]['SEQ_QTY'], int(print_data))
                                textData = textData.replace("{%s}"%i, str(print_data))
                            self.print_socket.send(textData.encode())
                    except: logging.debug("printData : selectCNClabel 실패")
                    self.print_socket.close()
                except:
                    logging.debug("printData : 프린터 연결 실패")
                    self.connectPrinter() #바코드 프린터 연결
                self.dorna_th.start()
            except: logging.debug("printData : dorna 전송 실패")
        except: pass
    
    #엣지1 전표 취소
    def item1Cancel(self, index):
        SEQ = self.tableWidget.item(index, 2).text()
        QR_CODE = self.tableWidget.item(index, 7).text()
        #----------------------------------------------------------------------------
        QR_rows = DaesungQuery.selectCNClabel(self, PROC_CODE, QR_CODE)
        #----------------------------------------------------------------------------
        if QR_rows == 'failed':self.connectDBThread()
        else:
            JAKUP_NO = QR_rows[0]['REG_NO']
            JAKUP_SEQ = QR_rows[0]['REG_SEQ']
            JAKUP_SORT_KEY = QR_rows[0]['SORT_KEY']
            #----------------------------------------------------------------------------
            result = DaesungQuery.PR_SAVE_MAKE_BAR(self, 'insert', '0', EMPL_CODE, JAKUP_NO, JAKUP_SEQ, JAKUP_SORT_KEY, QR_CODE, self.c_date, 1, 1) #불량수량 등록
            #----------------------------------------------------------------------------
            if result == 1:
                #----------------------------------------------------------------------------
                DaesungQuery.updateEdgeSeq(self, "PUT_FLAG = 'F'", self.s_date[:6], LINE_FLAG, QR_CODE, SEQ)
                #----------------------------------------------------------------------------
                r_text = '품목 삭제 완료'
            else: r_text = '품목 삭제 실패'
            logging.debug("item1Cancel : %s"%r_text)
            self.result_data.setText(r_text)
            self.DBload() #DB로드
        self.format_th.start()
    
    #엣지2 전표 취소
    def item2Cancel(self, index):
        SEQ = self.tableWidget2.item(index, 2).text()
        QR_CODE = self.tableWidget2.item(index, 7).text()
        #----------------------------------------------------------------------------
        QR_rows = DaesungQuery.selectCNClabel(self, PROC_CODE, QR_CODE)
        #----------------------------------------------------------------------------
        if QR_rows == 'failed':self.connectDBThread()
        else:
            JAKUP_NO = QR_rows[0]['REG_NO']
            JAKUP_SEQ = QR_rows[0]['REG_SEQ']
            JAKUP_SORT_KEY = QR_rows[0]['SORT_KEY']
            #----------------------------------------------------------------------------
            result = DaesungQuery.PR_SAVE_MAKE_BAR(self, 'insert', '0', EMPL_CODE, JAKUP_NO, JAKUP_SEQ, JAKUP_SORT_KEY, QR_CODE, self.c_date, 1, 1) #불량수량 등록
            #----------------------------------------------------------------------------
            if result == 1:
                #----------------------------------------------------------------------------
                DaesungQuery.updateEdgeSeq(self, "PUT_FLAG = 'S'", self.s_date[:6], LINE_FLAG, QR_CODE, SEQ)
                #----------------------------------------------------------------------------
                r_text = '품목 삭제 완료'
            else: r_text = '품목 삭제 실패'
            logging.debug("item2Cancel : %s"%r_text)
            self.result_data.setText(r_text)
            self.DBload() #DB로드
        self.format_th.start()
    
    #---------------------------------------------------------------------------------------------------
    def clickedSensor(self, data):
        self.Edge1SensorSlot(data)
        self.SensorSlot(data)
        self.Edge2SensorSlot(data)
    
    @pyqtSlot(list)
    def PLCEdgeSlot(self, data):
        try:
            #엣지1 대기
            if data[0] != 99999 and data[0] != self.e1_data:
                self.e1_data = data[0]
                logging.debug("PLCEdgeSlot 1 : {0}".format(data[0]))
                #----------------------------------------------------------------------------
                SEQ = DaesungQuery.selectEdgeSeq(self, LINE_FLAG, self.e1_data, self.s_date[:6], "(PUT_FLAG = '0' OR PUT_FLAG = 'F')")
                #----------------------------------------------------------------------------
                if SEQ != None:
                    QR_rows = DaesungQuery.selectCNClabel(self, PROC_CODE, SEQ['BAR_CODE'])
                    #----------------------------------------------------------------------------
                    result = DaesungQuery.PR_SAVE_MAKE_BAR(self, 'insert', '0', EMPL_CODE, QR_rows[0]['REG_NO'], QR_rows[0]['REG_SEQ'], QR_rows[0]['SORT_KEY'], QR_rows[0]['BAR_CODE'], self.c_date, 1, 0)
                    #----------------------------------------------------------------------------
                    if result == 1:
                        #----------------------------------------------------------------------------
                        DaesungQuery.updateEdgeSeq(self, "PUT_FLAG = '1'", self.s_date[:6], SEQ['MES_FLAG'], SEQ['BAR_CODE'], SEQ['SEQ'])
                        #----------------------------------------------------------------------------
                        self.DBload()
                else: logging.debug("PLCEdgeSlot 1 : SEQ 없음")
            #엣지2 대기
            if data[1] != 99999 and data[1] != self.e2_data:
                self.e2_data = data[1]
                logging.debug("PLCEdgeSlot 2 : {0}".format(data[1]))
                #----------------------------------------------------------------------------
                SEQ = DaesungQuery.selectEdgeSeq(self, LINE_FLAG, self.e2_data, self.s_date[:6], "(PUT_FLAG = '1' OR PUT_FLAG = '2' OR PUT_FLAG = 'S')")
                #----------------------------------------------------------------------------
                if SEQ != None:
                    #----------------------------------------------------------------------------
                    DaesungQuery.updateEdgeSeq(self, "PUT_FLAG = '3'", self.s_date[:6], SEQ['MES_FLAG'], SEQ['BAR_CODE'], SEQ['SEQ'])
                    #----------------------------------------------------------------------------
                    self.DBload()     
                else: logging.debug("PLCEdgeSlot 2 : SEQ 없음")   
        except Exception as e: logging.debug("PLCEdgeSlot : failed(%s)"%e)
    
    @pyqtSlot(str)
    def Edge1SensorSlot(self, data):
        if data != "failed" and data[3:4] != self.es1_data:
            logging.debug("Edge1SensorSlot : {0}".format(data[3:4]))
            self.es1_data = data[3:4]
            if data[3:4] == '1': #엣지1 진행
                try:
                    rowCount = self.tableWidget.rowCount() - 1
                    if self.tableWidget.item(rowCount, 0).text() != '해당일자의 데이터가 없습니다.':
                        for i in range(rowCount, -1, -1):
                            if self.tableWidget.item(i, 6).text() == '대기':
                                #----------------------------------------------------------------------------
                                SEQ = DaesungQuery.selectEdgeSeq(self, LINE_FLAG, self.tableWidget.item(i, 2).text(), self.s_date[:6], "PUT_FLAG = '1'")
                                if SEQ != None: DaesungQuery.updateEdgeSeq(self, "PUT_FLAG = '2'", self.s_date[:6], SEQ['MES_FLAG'], SEQ['BAR_CODE'], SEQ['SEQ'])
                                else: logging.debug("Edge1SensorSlot : SEQ 없음")
                                #----------------------------------------------------------------------------
                                break
                            else: logging.debug("Edge1SensorSlot : 대기건 없음")
                        self.DBload()
                    else: logging.debug("Edge1SensorSlot : 데이터 없음")
                except Exception as e: logging.debug("Edge1SensorSlot : failed(%s)"%e)
    
    @pyqtSlot(str)
    def Edge2SensorSlot(self, data):
        if data != 'failed' and data[1:2] != self.es2_data:
            logging.debug("Edge2SensorSlot : {0}".format(data[1:2]))
            self.es2_data = data[1:2]
            if data[1:2] == '1':
                try:
                    rowCount = self.tableWidget2.rowCount() - 1
                    if rowCount >= 0:
                        for i in range(rowCount, -1, -1):
                            if self.tableWidget2.item(i, 6).text() == '대기':
                                #----------------------------------------------------------------------------
                                SEQ = DaesungQuery.selectEdgeSeq(self, LINE_FLAG, self.tableWidget2.item(i, 2).text(), self.s_date[:6], "PUT_FLAG = '3'")
                                if SEQ != None: DaesungQuery.updateEdgeSeq(self, "PUT_FLAG = '4'", self.s_date[:6], SEQ['MES_FLAG'], SEQ['BAR_CODE'], SEQ['SEQ'])
                                else: logging.debug("Edge2SensorSlot : SEQ 없음")
                                #----------------------------------------------------------------------------
                                if self.sensor_flag == 'success':
                                    EDGE = self.tableWidget2.item(i, 5).text()
                                    if EDGE == '-':
                                        try:
                                            self.ser.write('RY 1 1\r'.encode()) #red light
                                            self.ser.write('RY 1 1\r'.encode()) #red light
                                            logging.debug("Edge2SensorSlot : RY 1 1 성공")
                                            self.light_btn.setStyleSheet("background-color: #55cba7;") #green
                                        except:
                                            logging.debug("Edge2SensorSlot : RY 1 1 실패")
                                            self.light_btn.setStyleSheet("background-color: #fd97a5;") #red
                                break
                            else: logging.debug("Edge2SensorSlot : 대기건 없음")
                        self.DBload()
                    else: logging.debug("Edge2SensorSlot : 데이터 없음")
                except Exception as e: logging.debug("Edge2SensorSlot : failed(%s)"%e)
    
    @pyqtSlot(str)
    def SensorSlot(self, data):
        if data != 'failed' and data[2:3] != self.s_data:
            logging.debug("SensorSlot : {0}".format(data[2:3]))
            self.s_data = data[2:3]
            if data[2:3] == '1' and self.printer_flag == "success": self.printData()
        elif data == 'failed': self.connectSensor()
    
    #---------------------------------------------------------------------------------------------------
    @pyqtSlot(int)
    def rowCountSlot(self, data):
        if data == 1: self.DBload()
    
    @pyqtSlot(str)
    def PLCConSlot(self, state):
        if state == "success": self.connectPLC()
        elif state == "failed": self.PLC_btn.setStyleSheet("background-color: #fd97a5;") #red
    
    @pyqtSlot(str)
    def SensorConSlot(self, state):
        if state == "success": self.connectSensor()
        elif state == "failed": self.sensor_btn.setStyleSheet("background-color: #fd97a5;") #red
    
    @pyqtSlot(str)
    def PrinterConSlot(self, state):
        if state == "success": self.connectPrinter()
        elif state == "failed": self.Print_btn.setStyleSheet("background-color: #fd97a5;") #red
    
    @pyqtSlot(int)
    def FormatSlot(self, num):
        try:
            if num == 1: self.result_data.setText('')
            if num == 3 and self.sensor_flag == 'success':
                try:
                    self.ser.write('RY 1 0\r'.encode()) #green light
                    logging.debug("FormatSlot : RY 1 0 성공")
                except: logging.debug("FormatSlot : RY 1 0 실패")
            elif num == 4 and self.sensor_flag == 'success':
                try:
                    self.ser.write('RY 2 0\r'.encode())
                    logging.debug("FormatSlot : RY 2 0 성공")
                    self.ser.write('RY 3 0\r'.encode())
                    logging.debug("FormatSlot : RY 3 0 성공")
                except: logging.debug("FormatSlot : RY 2 0, RY 3 0 실패")
        except: pass
    
    #---------------------------------------------------------------------------------------------------
    def showDate(self, date):
        self.calender.hide()
        a = date.toString().split()
        if len(a[1]) == 1: a[1] = "0" + a[1]
        if len(a[2]) == 1: a[2] = "0" + a[2]
        self.date_btn.setText("%s-%s-%s "%(a[3],a[1],a[2])) # 월, 일자가 1자 일때
        self.calendar_flag = False
        DaesungFunctions.replaceDate(self) #DB로드
    
    def mousePressEvent(self, event):
        if event.buttons() == Qt.LeftButton and self.calendar_flag == True:
            self.calender.hide()
            self.calendar_flag = False
    
    def setData(self):
        try:
            SetWindow().showModal()
            self.hwConnect() #HW연결
        except: logging.debug("setData : 연결 실패")
    
    def logWindow(self):
        try: MesLogWindow(TodayData).showModal()
        except: logging.debug("logWindow : 로그 화면 연결 실패")
    
    #실시간 DB연결
    def connectDBThread(self):
        self.DB_btn.setStyleSheet("background-color: #fd97a5;") #red
        if self.DB_flag == 0:
            self.DB_flag = 1
            logging.debug("DBload : DB로드 실패")
            MessageWindow(self, "DB연결 실패").showModal()
            #----------------------------------------------------------------------------
            self.DB_th = ConnectDBThread()
            self.DB_th.sig_data.connect(self.DbThreadSlot)
            self.DB_th.start()
    
    @pyqtSlot(int)       
    def DbThreadSlot(self, con):
        if con == 1:
            self.DB_th.terminate()
            self.DB_flag = 0
            #----------------------------------------------------------------------------
            result = DaesungQuery.connectDB(self, host, port, user, name)
            #----------------------------------------------------------------------------
            if result == 'success':
                self.hwConnect() #HW연결
                DaesungFunctions.replaceDate(self) #DB로드
    
    def exitWindow(self):
        MessageWindow(self, 2).showModal()
        #----------------------------------
        if RECONNECT_FLAG == 1:
            try: self.th_rowCount.terminate()
            except: pass
            try: self.edge_th.terminate()
            except: pass
            try: self.plc_con_th.terminate()
            except: pass
            try: plc_socket.close()
            except: pass
            #------------------------------------------
            try: self.sensor_th.terminate()
            except: pass
            try: self.sensor_con_th.terminate()
            except: pass
            try: self.ser.close()
            except: pass
            #------------------------------------------
            try: self.prt_con_th.close()
            except: pass
            #------------------------------------------
            try: self.light_th.close()
            except: pass
            try: self.format_th.close()
            except: pass
            try: self.dorna_th.close()
            except: pass

################################################################################################################
class PlcPackThread(QThread):
    sig_data = pyqtSignal(int)
    
    def __init__(self, addr):
        super().__init__()
        self.addr = addr
    
    def run(self):
        LENGTH = RS232Addr('L', self.addr)
        s_data = bytes.fromhex(COMPANY + LENGTH[0] + POSTION + '54 00 02 00' + BLOCK + LENGTH[1] + '25 44 57' + LENGTH[2])
        while True:
            try:
                #----------------------------------------------------------------------------
                plc_socket.send(s_data)
                #----------------------------------------------------------------------------
                re = plc_socket.recv(1024).hex()
                pack = int((re[-2:] + re[-4:-2]), 16)
                self.sig_data.emit(pack)
            except: self.sig_data.emit(99999)
            time.sleep(3)

#---------------------------------------------------------------------------------------------------------------
#포장검수 Window
class MesPackWindow(QDialog):
    def __init__(self, date, flag):
        super(MesPackWindow, self).__init__()
        loadUi("ui\DAESUNG_MES_P.ui", self)
        
        if flag == 1: PROC = DaesungQuery.selectProc(self, EMPL_CODE, WC_CODE, "AND PROC.PROC_CODE IN ('0101', '0102', '0103', '0120')")
        else: PROC = ''
        
        DaesungFunctions.setComboStyle(self, date, WC_CODE, PROC_CODE, PROC_NAME, PROC) #기본 셋팅
        
        #----------------------------------------------------------------------------
        C_rows = DaesungQuery.selectCproc(self)
        #----------------------------------------------------------------------------
        self.C_rows = C_rows[0]['CPROC_CODE'].split(',')
        
        self.hwConnect() #HW연결
        DaesungFunctions.replaceDate(self) #DB로드
        try:
            self.light_th = FormatThread("light")
            self.light_th.sig_data.connect(self.FormatSlot)
        except: pass
        
        self.aproc_combo.currentIndexChanged.connect(self.procChanged) #공정 변경
        self.make_btn.clicked.connect(self.clickedMakeHw) #생산현황 화면 연결
        
        # 상단버튼 --------------------------------------------------
        self.date_btn.clicked.connect(lambda: DaesungFunctions.calendar(self))
        self.set_btn.clicked.connect(self.setData)
        self.log_btn.clicked.connect(self.logWindow) #로그 화면 연결
        self.as_btn.clicked.connect(lambda: DaesungFunctions.openUrl(self))
        self.exit_btn.clicked.connect(self.exitWindow)
    
    def hwConnect(self):
        try:
            self.set_win = SetWindow()
            self.connectPLC() #PLC 연결
            self.connectLight() #경광등 연결
        except: logging.debug("hwConnect : 실패")
    
    def connectPLC(self):
        global plc_socket
        #-------------------------------------------------------------
        try: self.pack_th.terminate()
        except: pass
        try: self.plc_con_th.terminate()
        except: pass
        try: plc_socket.close()
        except: pass
        #-------------------------------------------------------------
        if self.set_win.plc_check.isChecked():
            try:
                self.plc_ip = self.set_win.plc_ip_input.text()
                self.plc_port = int(self.set_win.plc_port_input.text())
                self.plc_pack = self.set_win.plc_pack_input.text()
                #-------------------------------------------------------------
                plc_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                plc_socket.settimeout(0.5)
                plc_socket.connect((self.plc_ip, self.plc_port))
                try:
                    self.pack_th = PlcPackThread(self.plc_pack)
                    self.pack_th.sig_data.connect(self.PLCPackSlot)
                    self.pack_th.start()
                except: logging.debug("connectPLC : pack_th 실패")
                self.PLC_flag = "success"
                logging.debug("connectPLC : PLC 연결 성공")
                self.PLC_btn.setStyleSheet("background-color: #55cba7;") #green
            except:
                self.PLC_flag = "failed"
                logging.debug("connectPLC : PLC 연결 실패")
                self.PLC_btn.setStyleSheet("background-color: #fd97a5;") #red
                try:
                    self.plc_con_th = SocketThread(self.plc_ip, self.plc_port)
                    self.plc_con_th.sig_data.connect(self.PLCConSlot)
                    self.plc_con_th.start()
                except: logging.debug("connectPLC : plc_con_th 실패")
        else:
            self.PLC_flag = "unable"
            logging.debug("connectPLC : PLC 비활성")
            self.PLC_btn.setStyleSheet("background-color: #CDCDCD;") #gray
    
    def connectLight(self):
        try: self.light_con_th.terminate()
        except: pass
        try: self.light_ser.close()
        except: pass
        #-------------------------------------------------------------
        if self.set_win.light_check.isChecked():
            try:
                light_port = self.set_win.light_port_input.text()
                light_rate = int(self.set_win.light_rate_input.text())
                #-------------------------------------------------------------
                self.light_ser = serial.Serial(light_port, light_rate, timeout = 0.5)
                self.light_flag = "success"
                logging.debug("connectLight : 경광등 연결 성공")
                self.light_btn.setStyleSheet("background-color: #55cba7;") #green
            except: 
                self.light_flag = "failed"
                logging.debug("connectLight : 경광등 연결 실패")
                self.light_btn.setStyleSheet("background-color: #fd97a5;") #red
                try:
                    self.light_con_th = SerialThread(light_port, light_rate)
                    self.light_con_th.sig_data.connect(self.LightConSlot)
                    self.light_con_th.start()
                except: logging.debug("connectLight : light_con_th 실패")
        else:
            self.light_flag = "unable"
            logging.debug("connectLight : 경광등 비활성")
            self.light_btn.setStyleSheet("background-color: #CDCDCD;") #gray
    
    #---------------------------------------------------------------------------------------------------
    def DBload(self):
        self.tableWidget.setRowCount(0)
        #----------------------------------------------------------------------------
        L_rows = DaesungQuery.selectPackList(self, LINE_FLAG, self.s_date, 'LIMIT 15')
        #----------------------------------------------------------------------------
        if L_rows == 'failed': self.connectDBThread()
        elif L_rows == ():
            self.tableWidget.setRowCount(1)
            self.tableWidget.setRowHeight(0, 100)
            self.tableWidget.setSpan(0, 0, 1, 10)
            item_data = QTableWidgetItem('해당일자의 데이터가 없습니다.')
            item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
            self.tableWidget.setItem(0, 0, item_data)
        else:
            self.tableWidget.setRowCount(len(L_rows))
            try:
                for i in range(len(L_rows)):
                    self.tableWidget.setRowHeight(i, 100)
                    proc_code = L_rows[i]['PROC_CODE'].split(',')
                    for count, j in enumerate(['SEQ', 'LOT_NUMB', 'ITEM_TEXT', 'CONN_CPROC_CODE', '0110', '0115', '0117', '0120']):
                        if 4 <= count <= 7:
                            states = QPushButton()
                            if j in proc_code: states.setStyleSheet(stateBtnStyle + "background-image: url(./img/complete.png);")
                            else: states.setStyleSheet(stateBtnStyle + "background-image: url(./img/none.png);")
                            self.tableWidget.setCellWidget(i, count, states)
                        else:
                            print_data = L_rows[i][j]
                            if print_data == None or print_data == '000': print_data = '-'
                            elif j == 'LOT_NUMB': print_data = print_data + '\n' + L_rows[i]['LK_JAKUP_SEQ']
                            elif j == 'ITEM_TEXT':
                                if L_rows[i]['SPCL_NAME'] == None: spcl = '/-'
                                elif L_rows[i]['EDGE_NAME'] == '-': spcl = '/' + L_rows[i]['SPCL_NAME']
                                else: spcl = '/{0}({1})'.format(L_rows[i]['SPCL_NAME'], L_rows[i]['EDGE_NAME'])
                                #-------------------------------------------------------------
                                if L_rows[i]['CAL_HOLE_VALUE'] == None or int(L_rows[i]['CAL_HOLE_VALUE']) == 0: hole = '/-'
                                else: hole = "/상(%d)"%int(L_rows[i]['CAL_HOLE_VALUE'])
                                #-------------------------------------------------------------
                                print_data = print_data + spcl + hole
                            elif j == 'CONN_CPROC_CODE':
                                c_data = str(print_data.replace(' ', '').replace("'", '')).split(',')
                                print_data = '-'
                                for code in c_data:
                                    if code in self.C_rows: print_data = 'O'
                            item_data = QTableWidgetItem(str(print_data))
                            if j != 'ITEM_TEXT': item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                            self.tableWidget.setItem(i, count, item_data)
            except: logging.debug("DBload : table 로드 실패")
            if self.reload_num == 0:
                try:
                    self.th_rowCount = reloadThread()
                    self.th_rowCount.sig_data.connect(self.rowCountSlot)
                    self.th_rowCount.start()
                    self.reload_num = 1
                except: logging.debug("DBload : th_rowCount 실패")
        DaesungFunctions.tableWidth(self, 'PACK', '', len(L_rows))
    
    @pyqtSlot(int)
    def PLCPackSlot(self, data):
        try:
            if data != 99999 and data != self.p_data:
                logging.debug("PLCPackSlot : {0}".format(data))
                self.p_data = data
                #----------------------------------------------------------------------------
                SEQ = DaesungQuery.selectEdgeSeq(self, LINE_FLAG, self.p_data, self.s_date[:6], "(PUT_FLAG = '3' OR PUT_FLAG = '4')")
                #----------------------------------------------------------------------------
                if SEQ != None:
                    QR_rows = DaesungQuery.selectCNClabel(self, PROC_CODE, SEQ['BAR_CODE'])
                    #----------------------------------------------------------------------------
                    result = DaesungQuery.PR_SAVE_MAKE_BAR(self, 'insert', '0', EMPL_CODE, QR_rows[0]['REG_NO'], QR_rows[0]['REG_SEQ'], QR_rows[0]['SORT_KEY'], QR_rows[0]['BAR_CODE'], self.c_date, 1, 0) #실적등록
                    #----------------------------------------------------------------------------
                    if result == 1:
                        #----------------------------------------------------------------------------
                        DaesungQuery.updateEdgeSeq(self, "PUT_FLAG = '5'", self.s_date[:6], SEQ['MES_FLAG'], SEQ['BAR_CODE'], SEQ['SEQ'])
                        #----------------------------------------------------------------------------
                        self.DBload()
                        if self.light_flag == "success" and self.tableWidget.item(0, 3).text() == 'O':
                            try:
                                self.light_ser.write('RY 1 1\r'.encode()) #red light
                                logging.debug("PLCPackSlot : RY 1 1 성공")
                                self.light_th.start()
                            except: logging.debug("PLCPackSlot : RY 1 1 실패")
            elif data == 99999: self.connectPLC()
        except: pass
    
    #---------------------------------------------------------------------------------------------------
    @pyqtSlot(str)
    def PLCConSlot(self, state):
        if state == "success": self.connectPLC()
        elif state == "failed": self.PLC_btn.setStyleSheet("background-color: #fd97a5;") #red
    
    @pyqtSlot(str)
    def LightConSlot(self, state):
        if state == "success": self.connectLight()
        elif state == "failed": self.light_btn.setStyleSheet("background-color: #fd97a5;") #red
    
    @pyqtSlot(int)
    def rowCountSlot(self, data):
        if data == 1: self.DBload()
    
    @pyqtSlot(int)
    def FormatSlot(self, num):
        if num == 3:
            try:
                self.light_ser.write('RY 1 0\r'.encode()) #green light
                logging.debug("FormatSlot : RY 1 0 성공")
            except: logging.debug("FormatSlot : RY 1 0 실패")
    
    #---------------------------------------------------------------------------------------------------
    #공정 변경
    def procChanged(self):
        self.threadTerminate()
        global PROC_CODE, PROC_NAME
        PROC_CODE = self.aproc_combo.currentText()[:4]
        PROC_NAME = self.aproc_combo.currentText()[4:]
        date = self.date_btn.text()
        widget.addWidget(MesLotWindow(date))
        widget.setCurrentIndex(widget.currentIndex() + 1)
        self.deleteLater()
    
    def showDate(self, date):
        self.calender.hide()
        a = date.toString().split()
        if len(a[1]) == 1: a[1] = "0" + a[1]
        if len(a[2]) == 1: a[2] = "0" + a[2]
        self.date_btn.setText("%s-%s-%s "%(a[3],a[1],a[2])) # 월, 일자가 1자 일때
        self.calendar_flag = False
        DaesungFunctions.replaceDate(self) #DB로드
    
    def mousePressEvent(self, event):
        if event.buttons() == Qt.LeftButton and self.calendar_flag == True:
            self.calender.hide()
            self.calendar_flag = False
    
    #생산현황 화면 연결
    def clickedMakeHw(self):
        try:
            try: self.th_rowCount.terminate()
            except: pass
            s_date = self.date_btn.text()
            MesMakeHwWindow(s_date).showModal()
            self.th_rowCount.start()
        except: logging.debug("clickRow : 상세페이지 연결 실패")
    
    def setData(self):
        try:
            SetWindow().showModal()
            self.hwConnect() #HW연결
        except: logging.debug("setData : 상세페이지 연결 실패")
    
    def logWindow(self):
        try: MesLogWindow(TodayData).showModal()
        except: logging.debug("logWindow : 로그 화면 연결 실패")
    
    #실시간 DB연결
    def connectDBThread(self):
        if self.DB_flag == 0:
            self.DB_flag = 1
            logging.debug("DBload : DB로드 실패")
            MessageWindow(self, "DB연결 실패").showModal()
            #----------------------------------------------------------------------------
            self.DB_th = ConnectDBThread()
            self.DB_th.sig_data.connect(self.DbThreadSlot)
            self.DB_th.start()
    
    @pyqtSlot(int)       
    def DbThreadSlot(self, con):
        if con == 1:
            self.DB_th.terminate()
            self.DB_flag = 0
            #----------------------------------------------------------------------------
            result = DaesungQuery.connectDB(self, host, port, user, name)
            #----------------------------------------------------------------------------
            if result == 'success':
                self.hwConnect() #HW연결
                DaesungFunctions.replaceDate(self) #DB로드
    
    def threadTerminate(self):
        try: self.th_rowCount.terminate()
        except: pass
        #------------------------------------------
        try: self.pack_th.terminate()
        except: pass
        try: self.plc_con_th.terminate()
        except: pass
        try: plc_socket.close()
        except: pass
        #------------------------------------------
        try: self.light_con_th.terminate()
        except: pass
        try: self.light_ser.close()
        except: pass
        #------------------------------------------
        try: self.light_th.terminate()
        except: pass
    
    def exitWindow(self):
        MessageWindow(self, 2).showModal()
        #----------------------------------
        if RECONNECT_FLAG == 1: self.threadTerminate()

################################################################################################################
#포장검수 Window > 생산현황 Window
class MesMakeHwWindow(QDialog):
    def __init__(self, date):
        super(MesMakeHwWindow, self).__init__()
        loadUi("ui\DAESUNG_MES_HW.ui", self)
        
        DaesungFunctions.setAutoStyle(self, date, 'MAKE', '') #기본 셋팅
        
        DaesungFunctions.replaceDate(self) #DB로드
        
        self.all_btn.clicked.connect(lambda: self.clickedSelectBtns('PACK(A)'))
        self.progress_btn.clicked.connect(lambda: self.clickedSelectBtns('PACK(P)'))
        self.complete_btn.clicked.connect(lambda: self.clickedSelectBtns('PACK(C)'))
        
        self.state_group.buttonClicked[int].connect(self.itemComplete)
        
        # 화살표버튼 --------------------------------------------------
        self.top.clicked.connect(lambda: DaesungFunctions.topData(self))
        self.prev.clicked.connect(lambda: DaesungFunctions.prevData(self))
        self.next.clicked.connect(lambda: DaesungFunctions.nextData(self))
        self.bottom.clicked.connect(lambda: DaesungFunctions.bottomData(self))
        
        # 상단버튼 --------------------------------------------------
        self.date_btn.clicked.connect(lambda: DaesungFunctions.calendar(self))
        self.exit_btn.clicked.connect(lambda: DaesungFunctions.closeWindow(self))
    
    def clickedSelectBtns(self, flag):
        self.s_flag = flag
        if flag == 'PACK(A)': all_style, pro_style, com_style = selectBtnStyle, '', '생산된 데이터가 없습니다.'
        elif flag == 'PACK(P)': all_style, pro_style, com_style, self.t_text = '', selectBtnStyle, '', '진행중인 데이터가 없습니다.'
        else: all_style, pro_style, com_style, self.t_text = '', '', selectBtnStyle, '완료된 데이터가 없습니다.'
        self.all_btn.setStyleSheet(all_style)
        self.progress_btn.setStyleSheet(pro_style)
        self.complete_btn.setStyleSheet(com_style)
        self.DBload()
    
    def DBload(self):
        self.tableWidget.setRowCount(0)
        #----------------------------------------------------------------------------
        L_rows = DaesungQuery.selectPackList(self, LINE_FLAG, self.s_date, '')
        #----------------------------------------------------------------------------
        if L_rows == 'failed': self.connectDBThread()
        elif L_rows == ():
            self.tableWidget.setRowCount(1)
            self.tableWidget.setRowHeight(0, 100)
            self.tableWidget.setSpan(0, 0, 1, 9)
            item_data = QTableWidgetItem(self.t_text)
            item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
            self.tableWidget.setItem(0, 0, item_data)
            self.qty_label.setText('0')
        else:
            try:
                for i in reversed(range(len(L_rows))):
                    if self.s_flag == 'PACK(P)' and (len(L_rows[i]['PROC_CODE'].split(',')) == 0 or len(L_rows[i]['PROC_CODE'].split(',')) == 4): L_rows.pop(i)
                    elif self.s_flag == 'PACK(C)' and len(L_rows[i]['PROC_CODE'].split(',')) < 4: L_rows.pop(i)
                if L_rows == []:
                    self.tableWidget.setRowCount(1)
                    self.tableWidget.setRowHeight(0, 100)
                    self.tableWidget.setSpan(0, 0, 1, 9)
                    item_data = QTableWidgetItem(self.t_text)
                    item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                    self.tableWidget.setItem(0, 0, item_data)
                else:
                    self.tableWidget.setRowCount(len(L_rows))
                    for i in range(len(L_rows)):
                        self.tableWidget.setRowHeight(i, 100)
                        proc_code = L_rows[i]['PROC_CODE'].split(',')
                        for count, j in enumerate(['SEQ', 'LOT_NUMB', 'ITEM_TEXT', '0110', '0115', '0117', '0120', 'BUTTON', 'BAR_CODE']):
                            if 3 <= count <= 6:
                                states = QPushButton()
                                if j in proc_code: states.setStyleSheet(stateBtnStyle + "background-image: url(./img/complete.png);")
                                else: states.setStyleSheet(stateBtnStyle + "background-image: url(./img/none.png);")
                                self.tableWidget.setCellWidget(i, count, states)
                            elif j == 'BUTTON' and self.s_flag == 'PACK(P)':
                                buttons = QPushButton("완료")
                                buttons.setStyleSheet("background:#fff8f2; color:#f7932f; border:2px solid #f7932f; border-radius:5px; font:bold 25px;")
                                buttons.setFont(QtGui.QFont("맑은 고딕"))
                                self.state_group.addButton(buttons, i)
                                self.tableWidget.setCellWidget(i, count, buttons)
                            elif j == 'BUTTON' and self.s_flag != 'PACK(P)': pass
                            else:
                                print_data = L_rows[i][j]
                                if print_data == None or print_data == '000': print_data = '-'
                                elif j == 'LOT_NUMB': print_data = print_data + '\n' + L_rows[i]['LK_JAKUP_SEQ']
                                elif j == 'ITEM_TEXT':
                                    if L_rows[i]['SPCL_NAME'] == None: spcl = '/-'
                                    elif L_rows[i]['EDGE_NAME'] == '-': spcl = '/' + L_rows[i]['SPCL_NAME']
                                    else: spcl = '/{0}({1})'.format(L_rows[i]['SPCL_NAME'], L_rows[i]['EDGE_NAME'])
                                    #-------------------------------------------------------------
                                    if L_rows[i]['CAL_HOLE_VALUE'] == None or int(L_rows[i]['CAL_HOLE_VALUE']) == 0: hole = '/-'
                                    else: hole = "/상(%d)"%int(L_rows[i]['CAL_HOLE_VALUE'])
                                    #-------------------------------------------------------------
                                    print_data = print_data + spcl + hole
                                item_data = QTableWidgetItem(str(print_data))
                                if j != 'ITEM_TEXT': item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                                self.tableWidget.setItem(i, count, item_data)
                self.qty_label.setText(str(len(L_rows)))
            except: logging.debug("DBload : table 로드 실패")
            DaesungFunctions.tableWidth(self, self.s_flag, '', len(L_rows))
    
    def itemComplete(self, row):
        QR_CODE = self.tableWidget.item(row, 8).text()
        #----------------------------------------------------------------------------
        QR_rows = DaesungQuery.selectCNClabel(self, '0110', QR_CODE)
        #----------------------------------------------------------------------------
        if QR_rows == 'failed':self.connectDBThread()
        else:
            JAKUP_NO = QR_rows[0]['REG_NO']
            JAKUP_SEQ = QR_rows[0]['REG_SEQ']
            JAKUP_SORT_KEY = QR_rows[0]['SORT_KEY']
            #----------------------------------------------------------------------------
            result = DaesungQuery.PR_SAVE_MAKE_BAR(self, 'insert', '1', EMPL_CODE, JAKUP_NO, JAKUP_SEQ, JAKUP_SORT_KEY, QR_CODE, self.s_date, 1, 0) #실적등록
            #----------------------------------------------------------------------------
            if result == 1: self.DBload()
    
    #----------------------------------------------------------------------------
    def showDate(self, date):
        self.calender.hide()
        a = date.toString().split()
        if len(a[1]) == 1: a[1] = "0" + a[1]
        if len(a[2]) == 1: a[2] = "0" + a[2]
        self.date_btn.setText("%s-%s-%s "%(a[3],a[1],a[2])) # 월, 일자가 1자 일때
        self.calendar_flag = False
        DaesungFunctions.replaceDate(self) #DB로드
    
    def mousePressEvent(self, event):
        if event.buttons() == Qt.LeftButton and self.calendar_flag == True:
            self.calender.hide()
            self.calendar_flag = False
    
    def setData(self):
        try: SetWindow().showModal()
        except: logging.debug("setData : 상세페이지 연결 실패")
    
    #실시간 DB연결
    def connectDBThread(self):
        self.DB_btn.setStyleSheet("background-color: #fd97a5;") #red
        if self.DB_flag == 0:
            self.DB_flag = 1
            logging.debug("DBload : DB로드 실패")
            MessageWindow(self, "DB연결 실패").showModal()
            #----------------------------------------------------------------------------
            self.DB_th = ConnectDBThread()
            self.DB_th.sig_data.connect(self.DbThreadSlot)
            self.DB_th.start()
    
    @pyqtSlot(int)       
    def DbThreadSlot(self, con):
        if con == 1:
            self.DB_th.terminate()
            self.DB_flag = 0
            #----------------------------------------------------------------------------
            result = DaesungQuery.connectDB(self, host, port, user, name)
            #----------------------------------------------------------------------------
            if result == 'success': self.DBload() #DB로드
    
    def showModal(self):
        return super().exec_()

################################################################################################################
#몰딩부 LOT Window
class MesMoldingLotWindow(QDialog):
    def __init__(self, date):
        super(MesMoldingLotWindow, self).__init__()
        loadUi("ui\DAESUNG_MES_MD_L.ui", self)
        
        DaesungFunctions.setComboStyle(self, date, WC_CODE, PROC_CODE, PROC_NAME, '') #기본 셋팅
        self.flag_combo.setCurrentText(PRT_NAME)
        self.sort_num, self.key_flag = [], 0
        
        self.hwConnect() #HW연결
        DaesungFunctions.replaceDate(self) #DB로드
        
        self.print_btn.clicked.connect(self.printLabel) #라벨 발행
        self.aproc_combo.currentIndexChanged.connect(self.procChanged) #공정 변경
        self.flag_combo.currentIndexChanged.connect(self.flagChanged)
        
        self.jackup_btn.clicked.connect(lambda: self.jackupPrint(1)) #기본 인쇄
        self.jackup_set_btn.clicked.connect(lambda: self.jackupPrint(2)) #세트 인쇄
        
        self.tableWidget.clicked.connect(self.click_row)
        self.tableWidget.horizontalHeader().sectionClicked.connect(self.selectedAll)
        
        keyboard.on_press_key("ctrl", lambda _:self.clickedCtrl(1))
        keyboard.on_release_key("ctrl", lambda _:self.clickedCtrl(0))
        
        # 화살표버튼 --------------------------------------------------
        self.top.clicked.connect(lambda: DaesungFunctions.topData(self))
        self.prev.clicked.connect(lambda: DaesungFunctions.prevData(self))
        self.next.clicked.connect(lambda: DaesungFunctions.nextData(self))
        self.bottom.clicked.connect(lambda: DaesungFunctions.bottomData(self))
        
        # 상단버튼 --------------------------------------------------
        self.DB_btn.clicked.connect(self.resetLight)
        self.date_btn.clicked.connect(lambda: DaesungFunctions.calendar(self))
        self.set_btn.clicked.connect(self.setData)
        self.log_btn.clicked.connect(self.logWindow)
        self.as_btn.clicked.connect(lambda: DaesungFunctions.openUrl(self))
        self.exit_btn.clicked.connect(self.exitWindow)
    
    def hwConnect(self):
        try:
            self.set_win = SetWindow()
            self.connectPrinter() #바코드 프린터 연결
            self.connectLight() #경광등 연결
        except: logging.debug("hwConnect : 실패")
    
    def connectPrinter(self):
        try: self.p_con_th.terminate()
        except: pass
        #-------------------------------------------------------------
        self.printer_ip = self.set_win.printer_ip_input.text()
        self.printer_port = self.set_win.printer_port_input.text()
        #-------------------------------------------------------------
        if self.set_win.printer_check.isChecked():
            try:
                self.mode =  '^' + self.set_win.printer_mode_input.text().replace(' ', '')
                #-------------------------------------------------------------
                print_socket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
                print_socket.settimeout(0.5)
                print_socket.connect((self.printer_ip, int(self.printer_port)))
                print_socket.close()
                #-------------------------------------------------------------
                self.printer_flag = "success"
                logging.debug("connectPrinter : 프린트 연결 성공")
                self.print_status.setStyleSheet("background-color: #55cba7;") #green
            except:
                self.printer_flag = "failed"
                logging.debug("connectPrinter : 프린트({0}:{1}) 연결 실패".format(self.printer_ip, self.printer_port))
                self.print_status.setStyleSheet("background-color: #fd97a5;") #red
                try:
                    self.p_con_th = SocketThread(self.printer_ip, int(self.printer_port))
                    self.p_con_th.sig_data.connect(self.PrinterConSlot)
                    self.p_con_th.start()
                except: logging.debug("connectPrinter : printer_con_th 실패")
        else:
            self.printer_flag = "unable"
            logging.debug("connectPrinter : 프린트({0}:{1}) 비활성".format(self.printer_ip, self.printer_port))
            self.print_status.setStyleSheet("background-color: #CDCDCD;") #gray
        
    def connectLight(self):
        global light_ser
        #--------------------------------------------------
        try: light_ser.close()
        except: pass
        #--------------------------------------------------
        light_port = self.set_win.light_port_input.text()
        if self.set_win.light_check.isChecked():
            light_rate = int(self.set_win.light_rate_input.text())
            try:
                light_ser = serial.Serial(light_port, light_rate, timeout = 0.5)
                self.light_flag = "success"
                logging.debug("connectLight : 경광등(%s) 연결 성공"%light_port)
            except:
                self.light_flag = "failed"
                logging.debug("connectLight : 경광등(%s) 연결 실패"%light_port)
        else:
            self.light_flag = "unable"
            logging.debug("connectLight : 경광등(%s) 비활성"%light_port)
    
    #---------------------------------------------------------------------------------------------------
    def setCheckFlag(self):
        try:
            self.set_check = []
            for count, checkbox in enumerate(self.checkBoxList):
                if checkbox.isChecked() == True: self.set_check.append(count)
        except: pass
    
    #미출력 LOT 알림
    def printAlarm(self, index):
        try:
            global PRT_INDEX
            if index != PRT_INDEX:
                PRT_INDEX = index
                if self.light_flag == 'success':
                    try:
                        light_ser.write('RY 1 1\r'.encode()) #red light
                        logging.debug("printAlarm : RY 1 1 성공")
                    except: logging.debug("printAlarm : RY 1 1 실패")
                    self.DB_btn.setStyleSheet("background-color: #fd97a5;") #red
                    MessageWindow(self, '미출력 로트 : [ %s ]'%self.tableWidget.item(index, 2).text()).showModal()
                    self.DB_btn.setStyleSheet("background-color: #55cba7;") #green
                    try:
                        light_ser.write('RY 1 0\r'.encode()) #green light
                        logging.debug("printAlarm : RY 1 0 성공")
                    except: logging.debug("printAlarm : RY 1 0 실패")
                else:
                    self.DB_btn.setStyleSheet("background-color: #fd97a5;") #red
                    MessageWindow(self, '미출력 로트 : [ %s ]'%self.tableWidget.item(index, 2).text()).showModal()
                    self.DB_btn.setStyleSheet("background-color: #55cba7;") #green
        except: pass
    
    def DBload(self):
        self.setCheckFlag()
        self.tableWidget.setSortingEnabled(False)
        self.tableWidget.setRowCount(0)
        try:
            global DATA_COUNT, JAKUP_TIME, PRT_INDEX
            DATA_COUNT, JAKUP_TIME = 0, 0
            self.checkBoxList, qty, prt_index = [], 0, -1
            #----------------------------------------------------------------------------
            S_rows = DaesungQuery.selectMoldingLotList(self, self.s_date, PRT_FLAG)
            #----------------------------------------------------------------------------
            if S_rows == 'failed': self.connectDBThread()
            elif S_rows == ():
                self.tableWidget.setRowCount(1)
                self.tableWidget.setRowHeight(0, 85)
                self.tableWidget.setSpan(0, 0, 1, 10)
                item_data = QTableWidgetItem("해당일자의 데이터가 없습니다.")
                item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                self.tableWidget.setItem(0, 0, item_data)
            else:
                DATA_COUNT = len(S_rows)
                self.tableWidget.setRowCount(len(S_rows))
                for i in range(len(S_rows)):
                    self.tableWidget.setRowHeight(i, 85)
                    ckbox = QCheckBox()
                    ckbox.setStyleSheet(t_checkStyle)
                    self.checkBoxList.append(ckbox)
                    if JAKUP_TIME < int(S_rows[i]['JAKUP_APPR_TIME']): JAKUP_TIME = int(S_rows[i]['JAKUP_APPR_TIME'])
                    #----------------------------------------------------------------------------
                    for count, j in enumerate(['CHECK', 'JAKUP_FLAG', 'LOT_NUMB', 'ITEM_TEXT', 'BUYER_TEXT', 'QTY', 'LK_PRT_FLAG', 'MES_PRT_FLAG', 'REG_NO']):
                        if j == 'CHECK':
                            cellWidget = QWidget()
                            layoutCB = QHBoxLayout(cellWidget)
                            layoutCB.addWidget(self.checkBoxList[i])
                            layoutCB.setAlignment(QtCore.Qt.AlignCenter)
                            layoutCB.setContentsMargins(0, 0, 0, 0)
                            cellWidget.setLayout(layoutCB)
                            self.tableWidget.setCellWidget(i, count, cellWidget)
                            if i in self.set_check: self.checkBoxList[i].setChecked(True)
                        elif j == 'LK_PRT_FLAG' or j == 'MES_PRT_FLAG':
                            states = QPushButton()
                            #------------------------------------------------------------
                            print_data = S_rows[i][j]
                            if print_data == None: states.setStyleSheet(stateBtnStyle + "background-image: url(./img/대기.png);")
                            elif print_data == '0' or print_data == '00': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/대기.png);")
                            elif print_data == '2' or print_data == '01': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/진행.png);")
                            elif print_data == '1' or print_data == '02': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/완료.png);")
                            #------------------------------------------------------------
                            if j == 'LK_PRT_FLAG': self.state_group.addButton(states, i)
                            else:
                                self.state2_group.addButton(states, i)
                                if print_data == None or print_data != '1':
                                    print_data = ''
                                    if prt_index == -1: prt_index = i
                                self.tableWidget.setItem(i, 9, QTableWidgetItem(str(print_data)))
                            self.tableWidget.setCellWidget(i, count, states)
                        elif j == '': self.tableWidget.setItem(i, count, QTableWidgetItem(''))
                        else:
                            print_data = S_rows[i][j]
                            if j == 'QTY':
                                qty = qty + int(print_data)
                                item_data = QTableWidgetItem()
                                item_data.setData(Qt.DisplayRole, int(print_data))
                                item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                                self.tableWidget.setItem(i, count, item_data)
                            else:
                                if print_data == None: print_data = ''
                                elif j == 'JAKUP_FLAG':
                                    if print_data == '1': print_data = "시판"
                                    elif print_data == '2': print_data = "LX"
                                item_data = QTableWidgetItem(str(print_data))
                                if j == 'ITEM_TEXT' or j == 'BUYER_TEXT': item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                                else: item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                                self.tableWidget.setItem(i, count, item_data)
            self.DB_btn.setStyleSheet("background-color: #55cba7;") #green
            if self.reload_num == 0:
                try:
                    self.th_rowCount = lotLoadThread(20)
                    self.th_rowCount.sig_data.connect(self.newData)
                    self.th_rowCount.start()
                    self.reload_num = 1
                except: logging.debug("DBload : th_rowCount 실패")
            self.qty_label.setText(str(qty))
            #-----------------------------------------------
            if prt_index > -1: self.printAlarm(prt_index) #미출력 LOT 알림
            else: PRT_INDEX = -1
            #-----------------------------------------------
            DaesungFunctions.tableWidth(self, 'LOT', WC_CODE, len(S_rows))
        except: logging.debug("DBload : DB로드 실패")
    
    #라벨 발행
    def printLabel(self):
        checkArray = []
        if self.printer_flag == "success":
            for count, checkbox in enumerate(self.checkBoxList):
                if checkbox.isChecked() == True: checkArray.append(count)
            if checkArray != []:
                try:
                    self.mysocket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
                    self.mysocket.settimeout(0.5)
                    self.mysocket.connect((self.printer_ip, int(self.printer_port)))
                    self.print_status.setStyleSheet("background-color: #55cba7;") #green
                    try: self.th_rowCount.terminate()
                    except: pass
                    self.fileName = m_label
                    try:
                        for i in checkArray:
                            REG_NO = self.tableWidget.item(i, 8).text()
                            f_name = open(self.fileName, 'r', encoding = 'utf-8')
                            m_textData = f_name.read()
                            f_name.close()
                            if self.set_win.printer_mode_check.isChecked() == True: m_textData = m_textData.replace("^CI28", "^CI28" + self.mode)
                            if self.set_win.printer_po_check.isChecked() == True: m_textData = m_textData.replace("^LS0", "^LS0^POI")
                            elif self.set_win.printer_po_check.isChecked() == False: m_textData = m_textData.replace("^LS0", "^LS0^PON")
                            #----------------------------------------------------------------------------
                            P_rows = DaesungQuery.selectMoldingLotLabel(self, self.s_date, REG_NO)
                            #----------------------------------------------------------------------------
                            if P_rows == 'failed': self.connectDBThread()
                            elif P_rows != []:
                                for row in P_rows:
                                    textData, BUYER_CODE = m_textData, row['BUYER_CODE']
                                    try:
                                        for t in ['REG_NO', 'LOT_NUMB', 'REG_DATE', 'HOPE_DATE', 'BUYER_NAME', 'BUYER_CODE', 'BIGO', 'LABEL_BIGO', 'ITEM_NAMES']:
                                            if t == 'ITEM_NAMES':
                                                print_data = ''
                                                if row['SPCL_NAMES'] == None: spcl_names = []
                                                else: spcl_names = row['SPCL_NAMES'].split(',')
                                                item_names = row['ITEM_NAMES'].split(',')
                                                kyus = row['KYUS'].split(',')
                                                qtys = row['QTYS'].split(',')
                                                for d in range(len(item_names)):
                                                    detail_data = '^FO35,{0}^CFJ^AKN40,40^TBN,750,710^FD'.format(550 + 80 * d)
                                                    #컬러명 --------------------------------------------------------------------
                                                    if d < len(spcl_names): detail_data = detail_data + spcl_names[d].replace(' ', '') + ' '
                                                    else: detail_data = detail_data + '-' + ' '
                                                    #품명 ---------------------------------------------------------------------
                                                    detail_data = detail_data + item_names[d].replace(' ', '') + ' '
                                                    #규격 ---------------------------------------------------------------------
                                                    if d < len(kyus): detail_data = detail_data + kyus[d].replace(' ', '') + ' '
                                                    else: detail_data = detail_data + '-' + ' '
                                                    #수량 ---------------------------------------------------------------------
                                                    if d < len(qtys): detail_data = detail_data + str(int(float(qtys[d])))
                                                    else: detail_data = detail_data + '-'
                                                    #------------------------------------------------------------------------
                                                    print_data = print_data + detail_data + '^FS\n'
                                            else:
                                                print_data = row[t]
                                                if print_data == None: print_data = ""
                                                elif t == 'REG_DATE' or t == 'HOPE_DATE': print_data = "{0}.{1}.{2}".format(print_data[2:4], print_data[4:6], print_data[6:8])
                                            textData = textData.replace("{%s}"%t, str(print_data))
                                        self.mysocket.send(textData.encode())
                                        try:
                                            #----------------------------------------------------------------------------
                                            D_rows = DaesungQuery.selectMoldingDetailList(self, REG_NO, '%', '%', BUYER_CODE, self.s_date)
                                            #----------------------------------------------------------------------------
                                            for d_row in D_rows:
                                                #----------------------------------------------------------------------------
                                                if d_row['PRT_FLAG'] != '1': DaesungQuery.LABEL_UPDATE_SQL(self, REG_NO, d_row['REG_SEQ'], d_row['SEQ_QTY']) #라벨빌헹 FLAG UPDATE
                                                #----------------------------------------------------------------------------
                                            # QTY = row['QTYS'].split(',')
                                            # for c, C_REG_SEQ in enumerate(row['REG_SEQ'].split(',')):
                                            #     #----------------------------------------------------------------------------
                                            #     M_rows = DaesungQuery.selectMakeRegData(self, '0404', REG_NO, C_REG_SEQ)
                                            #     #----------------------------------------------------------------------------
                                            #     if M_rows == (): c_date = self.c_date
                                            #     else: c_date = M_rows[0]['REG_DATE']
                                            #     #----------------------------------------------------------------------------
                                            #     DaesungQuery.PR_SAVE_MAKE(self, 'insert', '0', EMPL_CODE, REG_NO, C_REG_SEQ.replace(' ', ''), row['SORT_KEY'], c_date, int(float(QTY[c])), 0) #실적등록
                                            #     #----------------------------------------------------------------------------
                                        except: pass
                                    except: logging.debug("printLabel : selectDetailList 실패")
                                DaesungQuery.DBCommit(self) #DB commit
                            else: logging.debug("printLabel : 등록된 바코드 없음")
                        #----------------------------------------------------------------------------
                        # DaesungQuery.SELECT_PR_PASS_JAKUP_MAKE(self)
                        #----------------------------------------------------------------------------
                        self.mysocket.close()
                    except: logging.debug("printLabel : select 실패")
                    self.DBload()
                    self.th_rowCount.start()
                except:
                    self.print_status.setStyleSheet("background-color: #fd97a5;") #red
                    logging.debug("printLabel : 프린터 연결 실패")
            else: MessageWindow(self, "출력할 라벨을 선택해주세요.").showModal()
        else: MessageWindow(self, "프린터를 연결해주세요.").showModal()
    
    #---------------------------------------------------------------------------------------------------
    def click_row(self, index):
        try:
            JAKUP_FLAG = self.tableWidget.item(index.row(), 1).text()
            LOT_NO = self.tableWidget.item(index.row(), 2).text()
            REG_NO = self.tableWidget.item(index.row(), 8).text()
            s_date = self.date_btn.text()
            print_window = MesMoldingDetailWindow(JAKUP_FLAG, LOT_NO, REG_NO, s_date) #DETAIL 화면 연결
            try: self.p_con_th.terminate()
            except: pass
            try: self.th_rowCount.terminate()
            except: pass
            widget.addWidget(print_window)
            widget.setCurrentIndex(widget.currentIndex() + 1)
            self.deleteLater()
        except: logging.debug("clickRow : 상세페이지 연결 실패")
        
    def selectedAll(self, num):
        global SORT_ARRAY
        if len(SORT_ARRAY) == 5: SORT_ARRAY.pop(0)
        SORT_ARRAY.append(num)
        #-----------------------------------------------------------------
        if num == 0 and self.check_flag == False:
            for checkbox in self.checkBoxList: checkbox.setChecked(True)
            self.check_flag = True
        elif num == 0 and self.check_flag == True:
            for checkbox in self.checkBoxList: checkbox.setChecked(False)
            self.check_flag = False
    
    def clickedCtrl(self, flag):
        self.key_flag = flag
    
    #---------------------------------------------------------------------------------------------------
    @pyqtSlot(int)
    def newData(self, num):
        if num == 1: self.DBload() #DB로드
    
    @pyqtSlot(str)
    def PrinterConSlot(self, state):
        if state == "success": self.connectPrinter()
        elif state == "failed": self.print_status.setStyleSheet("background-color: #fd97a5;") #red
    
    #---------------------------------------------------------------------------------------------------
    def jackupPrint(self, flag):
        try:
            checkArray, P_REG_NO = [], ''
            for count, checkbox in enumerate(self.checkBoxList):
                if checkbox.isChecked() == True:
                    checkArray.append(count)
            if checkArray != []:
                for i in checkArray:
                    REG_NO = self.tableWidget.item(i, 8).text()
                    P_REG_NO = P_REG_NO + '&reg_no[]=%s'%REG_NO
                    self.checkBoxList[i].setChecked(False)
            if flag == 1: DaesungFunctions.jackupPrint(self, WC_CODE, PROC_CODE, P_REG_NO, EMPL_CODE, '')
            else: DaesungFunctions.jackupPrint(self, WC_CODE, PROC_CODE, P_REG_NO, EMPL_CODE, 'set')
        except: logging.debug("jackupPrint : failed")
    
    #공정 변경
    def procChanged(self):
        global PROC_CODE, PROC_NAME
        self.tableWidget.clearSelection()
        if PROC_CODE[:2] != self.aproc_combo.currentText()[:2]: self.checkBoxList = []
        PROC_CODE = self.aproc_combo.currentText()[:4]
        PROC_NAME = self.aproc_combo.currentText()[4:]
        self.DB_flag, self.reload_num = 0, 0
        DaesungFunctions.replaceDate(self) #DB로드
        self.hwConnect() #HW연결
    
    def flagChanged(self):
        global PRT_FLAG, PRT_NAME
        self.tableWidget.clearSelection()
        self.checkBoxList = []
        PRT_NAME = self.flag_combo.currentText()
        if PRT_NAME == '전체': PRT_FLAG = '%'
        elif PRT_NAME == '대기': PRT_FLAG = '0'
        elif PRT_NAME == '부분': PRT_FLAG = '2'
        elif PRT_NAME == '완료': PRT_FLAG = '1'
        DaesungFunctions.replaceDate(self) #DB로드
    
    def resetLight(self):
        if self.light_flag == 'success':
            try:
                light_ser.write('RY 1 0\r'.encode()) #green light
                logging.debug("resetLight : RY 1 0 성공")
            except: logging.debug("resetLight : RY 1 0 실패")
        self.DB_btn.setStyleSheet("background-color: #55cba7;") #green
    
    def showDate(self, date):
        self.calender.hide()
        a = date.toString().split()
        if len(a[1]) == 1: a[1] = "0" + a[1]
        if len(a[2]) == 1: a[2] = "0" + a[2]
        self.date_btn.setText("%s-%s-%s "%(a[3],a[1],a[2]))
        self.calendar_flag = False
        global PRT_INDEX
        PRT_INDEX = -1
        DaesungFunctions.replaceDate(self) #DB로드
    
    def mousePressEvent(self, event):
        if event.buttons() == Qt.LeftButton and self.calendar_flag == True:
            self.calender.hide()
            self.calendar_flag = False
    
    def setData(self):
        try:
            SetWindow().showModal()
            self.hwConnect() #HW연결
        except: logging.debug("setData : 상세페이지 연결 실패")
    
    def logWindow(self):
        try:
            try: self.th_rowCount.terminate()
            except: pass
            MesLogWindow(TodayData).showModal()
            self.th_rowCount.start()
        except: logging.debug("logWindow : 로그 화면 연결 실패")
    
    #실시간 DB연결
    def connectDBThread(self):
        if self.DB_flag == 0:
            self.DB_flag = 1
            MessageWindow(self, "DB연결 실패").showModal()
            self.DB_th = ConnectDBThread()
            self.DB_th.sig_data.connect(self.DbThreadSlot)
            self.DB_th.start()
    
    @pyqtSlot(int)       
    def DbThreadSlot(self, con):
        if con == 1:
            self.DB_th.terminate()
            self.DB_flag = 0
            #----------------------------------------------------------------------------
            result = DaesungQuery.connectDB(self, host, port, user, name)
            #----------------------------------------------------------------------------
            if result == 'success': DaesungFunctions.replaceDate(self) #DB로드
    
    def threadTerminate(self):
        try: self.th_rowCount.terminate()
        except: pass
        #--------------------------------------------------
        try: self.p_con_th.terminate()
        except: pass
        try: light_ser.close()
        except: pass
    
    def exitWindow(self):
        MessageWindow(self, 2).showModal()
        #----------------------------------
        if RECONNECT_FLAG == 1: self.threadTerminate()

################################################################################################################
#몰딩부 DETAIL Window
class MesMoldingDetailWindow(QDialog):
    def __init__(self, gubun, lot, reg_no, date):
        super(MesMoldingDetailWindow, self).__init__()
        loadUi("ui\DAESUNG_MES_MD_D.ui", self)
        
        DaesungFunctions.setDetailStyle(self, date, WC_CODE, PROC_CODE, gubun, lot, reg_no) #기본 셋팅
        
        self.hwConnect() #HW연결
        self.DBload() #DB로드
        
        self.print_btn.clicked.connect(self.printLabel) #라벨 발행
        self.jackup_btn.clicked.connect(lambda: self.jackupPrint(1)) #기본 인쇄
        
        self.select_all.clicked.connect(lambda: self.selectedAll(0))
        self.tableWidget.horizontalHeader().sectionClicked.connect(self.selectedAll)
        
        self.tableWidget.currentCellChanged.connect(self.connectTable)
        self.tableWidget.cellPressed.connect(self.clickedRow)
        self.tableWidget.cellClicked.connect(self.clickedRow)
        
        # 화살표버튼 --------------------------------------------------
        self.top.clicked.connect(lambda: DaesungFunctions.topData(self))
        self.prev.clicked.connect(lambda: DaesungFunctions.prevData(self))
        self.next.clicked.connect(lambda: DaesungFunctions.nextData(self))
        self.bottom.clicked.connect(lambda: DaesungFunctions.bottomData(self))
        
        # 상단버튼 --------------------------------------------------
        self.DB_btn.clicked.connect(self.resetLight)
        self.set_btn.clicked.connect(self.setData)
        self.log_btn.clicked.connect(self.logWindow)
        self.as_btn.clicked.connect(lambda: DaesungFunctions.openUrl(self))
        self.back_btn.clicked.connect(self.back)
    
    def hwConnect(self):
        try:
            self.set_win = SetWindow()
            self.connectPrinter() #바코드 프린터 연결
        except: logging.debug("hwConnect : failed")
    
    def connectPrinter(self):
        try: self.p_con_th.terminate()
        except: pass
        #-------------------------------------------------------------
        self.printer_ip = self.set_win.printer_ip_input.text()
        self.printer_port = self.set_win.printer_port_input.text()
        #-------------------------------------------------------------
        if self.set_win.printer_check.isChecked():
            try:
                self.mode =  '^' + self.set_win.printer_mode_input.text().replace(' ', '')
                #-------------------------------------------------------------
                print_socket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
                print_socket.settimeout(0.5)
                print_socket.connect((self.printer_ip, int(self.printer_port)))
                print_socket.close()
                #-------------------------------------------------------------
                self.printer_flag = "success"
                logging.debug("connectPrinter : 프린트 연결 성공")
                self.print_status.setStyleSheet("background-color: #55cba7;") #green
            except:
                self.printer_flag = "failed"
                logging.debug("connectPrinter : 프린트({0}:{1}) 연결 실패".format(self.printer_ip, self.printer_port))
                self.print_status.setStyleSheet("background-color: #fd97a5;") #red
                try:
                    self.p_con_th = SocketThread(self.printer_ip, int(self.printer_port))
                    self.p_con_th.sig_data.connect(self.PrinterConSlot)
                    self.p_con_th.start()
                except: logging.debug("connectPrinter : printer_con_th 실패")
        else:
            self.printer_flag = "unable"
            logging.debug("connectPrinter : 프린트({0}:{1}) 비활성".format(self.printer_ip, self.printer_port))
            self.print_status.setStyleSheet("background-color: #CDCDCD;") #gray
    
    #---------------------------------------------------------------------------------------------------
    def setCheckFlag(self):
        try:
            self.set_check = []
            for count, checkbox in enumerate(self.checkBoxList):
                if checkbox.isChecked() == True: self.set_check.append(count)
        except: pass
    
    def DBload(self):
        self.setCheckFlag()
        self.tableWidget.clearSelection()
        self.tableWidget.setRowCount(0)
        self.checkBoxList = []
        self.BUYER_DATA = {}
        try:
            #----------------------------------------------------------------------------
            D_rows = DaesungQuery.selectMoldingDetailList(self, self.REG_NO, '%', '%', '%', self.s_date)
            #----------------------------------------------------------------------------
            if D_rows == 'failed': self.connectDBThread()
            elif D_rows == ():
                logging.debug("DBload : 작업지시서 취소됨")
                MessageWindow(self, "해당 작업지시서가 취소되었습니다.").showModal()
                self.back()
            else:
                self.tableWidget.setRowCount(len(D_rows))
                for i in range(len(D_rows)):
                    self.tableWidget.setRowHeight(i, 85)
                    ckbox = QCheckBox()
                    ckbox.setStyleSheet(t_checkStyle)
                    self.checkBoxList.append(ckbox)
                    for count, j in enumerate(['CHECK', 'REG_SEQ', 'HOPE_DATE', 'ITEM_TEXT', 'BUYER_NAME', 'KYU', 'QTY_NO_ALL', 'PRT_FLAG', 'MES_PRT_FLAG', 'SEQ_QTY']):
                        if j == 'CHECK':
                            cellWidget = QWidget()
                            layoutCB = QHBoxLayout(cellWidget)
                            layoutCB.addWidget(self.checkBoxList[i])
                            layoutCB.setAlignment(QtCore.Qt.AlignCenter)
                            layoutCB.setContentsMargins(0, 0, 0, 0)
                            cellWidget.setLayout(layoutCB)
                            self.tableWidget.setCellWidget(i, count, cellWidget)
                            if i in self.set_check: self.checkBoxList[i].setChecked(True)
                        elif j == 'PRT_FLAG':
                            states = QPushButton()
                            if D_rows[i][j] == '1': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/완료.png);")
                            else: states.setStyleSheet("background: none; border: none;")
                            self.state_group.addButton(states, i)
                            self.tableWidget.setCellWidget(i, count, states)
                        elif j == 'MES_PRT_FLAG':
                            states = QPushButton()
                            if D_rows[i]['MES_PRT_FLAG'] == '1': states.setStyleSheet(stateBtnStyle + " background-image: url(./img/완료.png);")
                            elif D_rows[i]['MES_PRT_FLAG'] == '2': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/진행.png);")
                            else: states.setStyleSheet("background: none; border: none;")
                            self.select_group.addButton(states, i)
                            self.tableWidget.setCellWidget(i, count, states)
                        else:
                            print_data = D_rows[i][j]
                            if print_data == None: print_data = ''
                            elif j == 'HOPE_DATE': print_data = "{0}/{1}".format(print_data[4:6], print_data[6:8])
                            elif j == 'BUYER_NAME':
                                if (print_data in self.BUYER_DATA) == False: self.BUYER_DATA[print_data] = D_rows[i]['BUYER_CODE']
                            elif j == 'QTY': print_data = int(print_data)
                            item_data = QTableWidgetItem(str(print_data))
                            if j == 'REG_SEQ' or j == 'HOPE_DATE': item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                            elif j == 'KYU' or j == 'QTY': item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                            elif j == 'QTY_NO_ALL': item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                            self.tableWidget.setItem(i, count, item_data)
                DaesungFunctions.tableWidth(self, PROC_CODE, WC_CODE, len(D_rows))
                if self.reload_num == 0:
                    try:
                        self.th_rowCount = lotLoadThread(20)
                        self.th_rowCount.sig_data.connect(self.newData)
                        self.th_rowCount.start()
                        self.reload_num = 1
                    except: logging.debug("DBload : th_rowCount 실패")
                self.DB_btn.setStyleSheet("background-color: #55cba7;") #green
                #----------------------------------------------------------------------------
                for data in self.BUYER_DATA: self.buyer_combo.addItem(data)
        except: logging.debug("DBload : table 로드 실패")
    
    #라벨 발행
    def printLabel(self):
        checkArray = []
        if self.printer_flag == "success":
            for count, checkbox in enumerate(self.checkBoxList):
                if checkbox.isChecked() == True: checkArray.append(count)
            if checkArray != []:
                try:
                    self.mysocket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
                    self.mysocket.settimeout(0.5)
                    self.mysocket.connect((self.printer_ip, int(self.printer_port)))
                    self.print_status.setStyleSheet("background-color: #55cba7;") #green
                    try: self.th_rowCount.terminate()
                    except: pass
                    self.fileName = m_label
                    try:
                        for i in checkArray:
                            REG_SEQ, QTY_NO, SEQ_QTY = self.tableWidget.item(i, 1).text(), self.tableWidget.item(i, 6).text(), self.tableWidget.item(i, 9).text()
                            f_name = open(self.fileName, 'r', encoding = 'utf-8')
                            textData = f_name.read()
                            f_name.close()
                            if self.set_win.printer_mode_check.isChecked() == True: textData = textData.replace("^CI28", "^CI28" + self.mode)
                            if self.set_win.printer_po_check.isChecked() == True: textData = textData.replace("^LS0", "^LS0^POI")
                            elif self.set_win.printer_po_check.isChecked() == False: textData = textData.replace("^LS0", "^LS0^PON")
                            #----------------------------------------------------------------------------
                            P_rows = DaesungQuery.selectMoldingDetailList(self, self.REG_NO, REG_SEQ, SEQ_QTY, '%', self.s_date)
                            #----------------------------------------------------------------------------
                            if P_rows == 'failed': self.connectDBThread()
                            elif P_rows != []:
                                try:
                                    for j in ['REG_NO', 'LOT_NUMB', 'REG_DATE', 'HOPE_DATE', 'BUYER_NAME', 'BUYER_CODE', 'BIGO', 'LABEL_BIGO', 'ITEM_NAMES']:
                                        if j == 'ITEM_NAMES':
                                            #컬러명 --------------------------------------------------------------------
                                            spcl_names = P_rows[0]['SPCL_NAME']
                                            if spcl_names == None or spcl_names == '': spcl_names = '-'
                                            #품명 ---------------------------------------------------------------------
                                            item_names = P_rows[0]['ITEM_NAME']
                                            if item_names == None or item_names == '': item_names = '-'
                                            #규격 ---------------------------------------------------------------------
                                            kyus = P_rows[0]['KYU']
                                            if kyus == None or kyus == '': kyus = '-'
                                            #------------------------------------------------------------------------
                                            print_data = '^FO35,550^CFJ^AKN40,40^TBN,750,710^FD{0} {1} {2} {3}^FS'.format(spcl_names, item_names, kyus, QTY_NO)
                                        else:
                                            print_data = P_rows[0][j]
                                            if print_data == None: print_data = ""
                                            elif j == 'REG_DATE' or j == 'HOPE_DATE': print_data = "{0}.{1}.{2}".format(print_data[2:4], print_data[4:6], print_data[6:8]) 
                                        textData = textData.replace("{%s}"%j, str(print_data))
                                    self.mysocket.send(textData.encode())
                                    try:
                                        #----------------------------------------------------------------------------
                                        if P_rows[0]['PRT_FLAG'] != '1': DaesungQuery.LABEL_UPDATE_SQL(self, self.REG_NO, REG_SEQ, SEQ_QTY) #라벨발행 FLAG UPDATE
                                        #----------------------------------------------------------------------------
                                        # M_rows = DaesungQuery.selectMakeData(self, PROC_CODE, P_rows[0]['BAR_CODE'])
                                        #----------------------------------------------------------------------------
                                        # if M_rows == (): DaesungQuery.PR_SAVE_MAKE_BAR_DETAIL(self, 'insert', '0', EMPL_CODE, self.REG_NO, P_rows[0]['REG_SEQ'], P_rows[0]['SORT_KEY'], P_rows[0]['BAR_CODE'], self.c_date, 1, 0) #실적등록
                                        time.sleep(0.3)
                                    except: pass
                                except: logging.debug("printLabel : selectDetailList 실패")
                            else: logging.debug("printLabel : 등록된 바코드 없음")
                        #----------------------------------------------------------------------------
                        DaesungQuery.DBCommit(self) #DB commit
                        # DaesungQuery.SELECT_PR_PASS_JAKUP_MAKE(self)
                        #----------------------------------------------------------------------------
                        self.mysocket.close()
                    except: logging.debug("printLabel : select 실패")
                    self.DBload() #DB로드
                    self.th_rowCount.start()
                except:
                    self.print_status.setStyleSheet("background-color: #fd97a5;") #red
                    logging.debug("printLabel : 프린터 연결 실패")
            else: MessageWindow(self, "출력할 라벨을 선택해주세요.").showModal()
        else: MessageWindow(self, "프린터를 연결해주세요.").showModal()
    
    #---------------------------------------------------------------------------------------------------
    def connectTable(self, row):
        self.checkBoxList[row].setChecked(True)
        
    def clickedRow(self, row):
        s_index, s_row = -1, []
        for item in self.tableWidget.selectedIndexes():
            if s_index != item.row():
                    s_index = item.row()
                    s_row.append(s_index)
        for i in range(self.tableWidget.rowCount()):
            if i in s_row: self.checkBoxList[row].setChecked(True)
            else: self.checkBoxList[i].setChecked(False)
        self.checkBoxList[row].setChecked(True)
    
    def selectedAll(self, num):
        if num == 0 and self.check_flag == False:
            for checkbox in self.checkBoxList: checkbox.setChecked(True)
            self.check_flag = True
        elif num == 0 and self.check_flag == True:
            for checkbox in self.checkBoxList: checkbox.setChecked(False)
            self.check_flag = False
    
    #---------------------------------------------------------------------------------------------------
    @pyqtSlot(int)
    def newData(self, num):
        if num == 1: self.DBload() #DB로드
    
    @pyqtSlot(str)
    def PrinterConSlot(self, state):
        if state == "success": self.connectPrinter() #바코드 프린터 연결
        elif state == "failed": self.print_status.setStyleSheet("background-color: #fd97a5;") #red
    
    @pyqtSlot(int)
    def FormatSlot(self, num):
        if num == 3:
            try: light_ser.write('RY 1 0\r'.encode()) #green light
            except: pass
    
    #---------------------------------------------------------------------------------------------------
    #작업지시서 인쇄
    def jackupPrint(self, flag):
        try:
            buyer_name = self.buyer_combo.currentText()
            buyer_code = self.BUYER_DATA[buyer_name]
            if flag == 1: url = 'http://namuip.com/wood/door_ds/ordr_jakup_master_m.php?e_date={0}&wc_code={1}&proc_code={2}&reg_no[]={3}&buyer_code[]={4}'.format(self.s_date, WC_CODE, '', self.REG_NO, buyer_code)
            else: url = 'http://namuip.com/wood/door_ds/ordr_jakup_master_m.php?e_date={0}&wc_code={1}&proc_code={2}&reg_no[]={3}&buyer_code[]={4}'.format(self.s_date, WC_CODE, '', self.REG_NO, buyer_code)
            webbrowser.open(url)
        except: logging.debug("jackupPrint : failed")
    
    def resetLight(self):
        try: light_ser.write('RY 1 0\r'.encode()) #green light
        except: pass
        self.DB_btn.setStyleSheet("background-color: #55cba7;") #green
    
    #----------------------------------------------------------------------------    
    def setData(self):
        try:
            SetWindow().showModal()
            self.hwConnect() #HW연결
        except: logging.debug("setData : 상세페이지 연결 실패")
    
    def logWindow(self):
        try:
            try: self.th_rowCount.terminate()
            except: pass
            MesLogWindow(TodayData).showModal()
            self.th_rowCount.start()
        except: logging.debug("logWindow : 로그 화면 연결 실패")
    
    #실시간 DB연결
    def connectDBThread(self):
        self.DB_btn.setStyleSheet("background-color: #fd97a5;") #red
        if self.DB_flag == 0:
            self.DB_flag = 1
            logging.debug("DBload : DB로드 실패")
            MessageWindow(self, "DB연결 실패").showModal()
            #----------------------------------------------------------------------------
            self.DB_th = ConnectDBThread()
            self.DB_th.sig_data.connect(self.DbThreadSlot)
            self.DB_th.start()
     
    @pyqtSlot(int)
    def DbThreadSlot(self, con):
        if con == 1:
            self.DB_th.terminate()
            self.DB_flag = 0
            #----------------------------------------------------------------------------
            result = DaesungQuery.connectDB(self, host, port, user, name)
            #----------------------------------------------------------------------------
            if result == 'success':
                self.connectPrinter() #바코드 프린터 연결
                self.DBload() #DB로드
    
    def back(self):
        try: self.p_con_th.terminate()
        except: pass
        try: self.light_th.terminate()
        except: pass
        try: self.th_rowCount.terminate()
        except: pass
        widget.addWidget(MesMoldingLotWindow(self.date_btn.text()))
        widget.setCurrentIndex(widget.currentIndex() + 1)
        self.deleteLater()

################################################################################################################
class BogangThread(QThread):
    sig_data = pyqtSignal(int)
    
    def __init__(self, flag):
        super().__init__()
        self.flag = flag
    
    def run(self):
        if self.flag == "load":
            while True:
                time.sleep(0.5)
                self.sig_data.emit(1)
        elif self.flag == "delete":
            while True:
                time.sleep(5)
                self.sig_data.emit(2)

#---------------------------------------------------------------------------------------------------------------
#보강재 부착 Window
class MesBogangWindow(QDialog):
    def __init__(self, date):
        super(MesBogangWindow, self).__init__()
        loadUi("ui\DAESUNG_MES_BG.ui", self)

        print("BOGANG")
        
        DaesungFunctions.setFrameStyle(self, date, PROC_CODE, PROC_NAME) #기본 셋팅
        
        self.DBload(0) #DB로드
        
        # 화살표버튼 --------------------------------------------------
        self.top.clicked.connect(lambda: DaesungFunctions.topData(self))
        self.prev.clicked.connect(lambda: DaesungFunctions.prevData(self))
        self.next.clicked.connect(lambda: DaesungFunctions.nextData(self))
        self.bottom.clicked.connect(lambda: DaesungFunctions.bottomData(self))
        
        # 상단버튼 --------------------------------------------------
        self.date_btn.clicked.connect(lambda: DaesungFunctions.calendar(self))
        self.set_btn.clicked.connect(self.setData)
        self.log_btn.clicked.connect(self.logWindow)
        self.as_btn.clicked.connect(lambda: DaesungFunctions.openUrl(self))
        self.exit_btn.clicked.connect(self.exitWindow)

    def DBload(self, flag):
        try:
            if flag == 1: self.s_time = time.strftime('%H%M%S')
            self.num_count = 0
            self.tableWidget.setRowCount(0)
            self.tableWidget.setRowCount(10)
            for i in range(10):
                t_index = 9 - i
                self.tableWidget.setRowHeight(t_index, 85)
                row = self.db_array[i]
                if row != '':
                    LOTID = row['LOT_ID']
                    if LOTID == None or LOTID == '':
                        self.tableWidget.setSpan(t_index, 0, 1, 8)
                        item_data = QTableWidgetItem('자투리')
                        if t_index == 0: item_data.setBackground(QtGui.QColor(255, 241, 209)) # light yellow
                        item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                        self.tableWidget.setItem(t_index, 0, item_data)
                    else:
                        BAR_CODE = row['BAR_CODE']
                        TBLR = LOTID[11:12]
                        #----------------------------------------------------------------------------
                        BOGANG = DaesungQuery.selectBogangData(self, BAR_CODE)
                        #----------------------------------------------------------------------------
                        for count, j in enumerate(['WIDX', 'LENX', 'CAL_HOLE_VALUE', 'FRAME_WIDX', 'FLAG', 'LOT_ID', 'SPCL_NAME', 'BUYER_NAME']):
                            if j == 'CAL_HOLE_VALUE':
                                if TBLR == 'T' or TBLR == 'B': print_data = '-'
                                else:
                                    if BOGANG[0][j] == None or int(BOGANG[0][j]) == 0: print_data = '-'
                                    else: print_data = int(BOGANG[0][j])
                                    CONN = BOGANG[0]['CONN_CPROC_CODE']
                                    # 철판작업 X
                                    if CONN.find('350') >= 0 or CONN.find('369') >= 0: print_data = '-'
                                    # 미는당김
                                    if CONN.find('242') >= 0 or CONN.find('245') >= 0: GC1  = 'L'    #래치홀(좌), 경첩(우)
                                    elif CONN.find('243') >= 0 or CONN.find('244') >= 0: GC1  = 'R'  #래치홀(우), 경첩(좌)
                                    elif CONN.find('191') >= 0 or CONN.find('194') >= 0: GC1  = 'R'  #래치홀(우), 경첩(-)
                                    elif CONN.find('192') >= 0 or CONN.find('193') >= 0: GC1  = 'L'  #래치홀(좌), 경첩(-)
                                    else: GC1  = 'L'
                                    # 홀가공X
                                    CONN_CPROC_CODE, CPROC_CODE = CONN.split(','), ''
                                    for C, CODE in enumerate(CONN_CPROC_CODE):
                                        if C + 1 == len(CONN_CPROC_CODE): CPROC_CODE = CPROC_CODE + "'%s'"%CODE.replace(' ', '').replace("'", '').replace('"', '')
                                        else: CPROC_CODE = CPROC_CODE + "'%s', "%CODE.replace(' ', '').replace("'", '').replace('"', '')
                                    #----------------------------------------------------------------------------
                                    H_rows = DaesungQuery.selectHoleFlag(self, CPROC_CODE)
                                    #----------------------------------------------------------------------------
                                    if H_rows != None: print_data = '-'
                                    #----------------------------------------------------------------------------
                                    if GC1 == 'L' and TBLR == 'R':  print_data = '-' #L에 찍힘
                                    elif GC1 == 'R' and TBLR == 'L': print_data = '-' #R에 찍힘
                            elif j == 'FLAG':
                                if TBLR == 'T' or TBLR == 'B': print_data = '요꼬(W)'
                                elif TBLR == 'L' or TBLR == 'R': print_data = '다대(H)'
                            elif j == 'SPCL_NAME' or j == 'BUYER_NAME':
                                print_data = BOGANG[0][j]
                                if print_data == None: print_data = ''
                            else:
                                print_data = row[j]
                                if print_data == None: print_data = ''
                                elif j == 'LENX' or j == 'WIDX': print_data = int(print_data)
                                elif j == 'FRAME_WIDX' and (TBLR == 'T' or TBLR == 'B'): print_data = ''
                            item_data =  QTableWidgetItem(str(print_data))
                            if TBLR == 'L' or TBLR == 'R': item_data.setForeground(QBrush(QColor(63, 139, 204))) #blue
                            if t_index == 0 or (t_index == 1 and self.db_array[9] != '' and self.db_array[9]['LOT_ID'] == ''): item_data.setBackground(QtGui.QColor(255, 241, 209)) # light yellow
                            if count == 3: item_data.setBackground(QtGui.QColor(218, 245, 226)) # light green
                            if j == 'SPCL_NAME' or j == 'BUYER_NAME': item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                            else: item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                            self.tableWidget.setItem(t_index, count, item_data)
                else: self.tableWidget.setItem(t_index, 0, QTableWidgetItem(''))
            if self.reload_num == 0:
                try:
                    self.th_rowCount = BogangThread('load')
                    self.th_rowCount.sig_data.connect(self.newData)
                    self.th_rowCount.start()
                    self.reload_num = 1
                except: logging.debug("DBload : th_rowCount 실패")
        except: logging.debug("DBload : failed")
    
    @pyqtSlot(int)
    def newData(self, num):
        if num == 1: 
            flag, self.num_count = 0, self.num_count + 1
            #----------------------------------------------------------------------------
            L_rows = DaesungQuery.selectBogangList(self, LINE_FLAG, self.s_date, self.s_time) #보강재 리스트 조회
            #----------------------------------------------------------------------------
            if L_rows != ():
                for i in range(len(L_rows) -1 , -1, -1):
                    if (L_rows[i]['SEQ'] in self.db_seq) == False:
                        flag = 1
                        if len(self.db_seq) == 10:
                            self.db_seq.pop(0)
                            self.db_array.pop(0)
                        self.db_seq.append(L_rows[i]['SEQ'])
                        self.db_array.append(L_rows[i])
            if self.num_count == 30 and flag == 0:
                if len(self.db_seq) == 10:
                    self.db_seq.pop(0)
                    self.db_array.pop(0)
                self.db_seq.append('')
                self.db_array.append('')
            if flag == 1 or self.num_count == 30: self.DBload(flag) #DB로드
    
    #----------------------------------------------------------------------------
    def showDate(self, date):
        self.calender.hide()
        a = date.toString().split()
        if len(a[1]) == 1: a[1] = "0" + a[1]
        if len(a[2]) == 1: a[2] = "0" + a[2]
        self.date_btn.setText("%s-%s-%s "%(a[3],a[1],a[2])) # 월, 일자가 1자 일때
        self.calendar_flag = False
        DaesungFunctions.replaceDate(self) #DB로드
    
    def mousePressEvent(self, event):
        if event.buttons() == Qt.LeftButton and self.calendar_flag == True:
            self.calender.hide()
            self.calendar_flag = False
    
    def setData(self):
        try: SetWindow().showModal()
        except: logging.debug("setData : 상세페이지 연결 실패")
    
    def logWindow(self):
        try: MesLogWindow(TodayData).showModal()
        except: logging.debug("logWindow : 로그 화면 연결 실패")
    
    #실시간 DB연결
    def connectDBThread(self):
        if self.DB_flag == 0:
            self.DB_flag = 1
            MessageWindow(self, "DB연결 실패").showModal()
            self.DB_th = ConnectDBThread()
            self.DB_th.sig_data.connect(self.DbThreadSlot)
            self.DB_th.start()
    
    @pyqtSlot(int)       
    def DbThreadSlot(self, con):
        if con == 1:
            self.DB_th.terminate()
            self.DB_flag = 0
            #----------------------------------------------------------------------------
            result = DaesungQuery.connectDB(self, host, port, user, name)
            #----------------------------------------------------------------------------
            if result == 'success': DaesungFunctions.replaceDate(self) #DB로드
    
    def exitWindow(self):
        MessageWindow(self, 2).showModal()
        #----------------------------------
        if RECONNECT_FLAG == 1:
            try: self.th_rowCount.terminate()
            except: pass

################################################################################################################
#문틀부 LOT Window
class MesFrameWindow(QDialog):
    def __init__(self, date):
        super(MesFrameWindow, self).__init__()
        loadUi("ui\DAESUNG_MES_F.ui", self)
        
        DaesungFunctions.setFrameStyle(self, date, PROC_CODE, PROC_NAME) #기본 셋팅
        if WC_CODE != '09': self.flag_radio_frame.hide()
        
        self.hwConnect() #HW연결
        self.flagChanged('L')
        
        self.lot_radio.clicked.connect(lambda: self.flagChanged('L')) #로트 라디오 버튼 선택
        self.detail_radio.clicked.connect(lambda: self.flagChanged('D')) #자동자재 라디오 버튼 선택
        
        self.widx_combo.currentIndexChanged.connect(self.widxChanged) #규격 변경
        self.search_input.returnPressed.connect(self.searchData)
        self.search_btn.clicked.connect(self.searchData)
        self.spcl_btn.clicked.connect(self.splcList) #자재리스트
        
        self.tableWidget.clicked.connect(self.click_row)
        self.tableWidget.horizontalHeader().sectionClicked.connect(self.selectedAll)
        
        self.sudong_ok_btn.clicked.connect(lambda: self.changedPrtFlag('1', '0')) #수동확정
        self.sudong_cancel_btn.clicked.connect(lambda: self.changedPrtFlag('0', '0')) #수동취소
        self.auto_ok_btn.clicked.connect(lambda: self.changedPrtFlag('0', '1')) #자동확정
        self.auto_cancel_btn.clicked.connect(lambda: self.changedPrtFlag('0', '0')) #자동취소
        self.auto2_ok_btn.clicked.connect(lambda: self.changedPrtFlag('0', '2')) #자동확정
        self.auto2_cancel_btn.clicked.connect(lambda: self.changedPrtFlag('0', '0')) #자동취소
        
        self.sudong_print_btn.clicked.connect(lambda: DaesungFunctions.jackupPrint(self, WC_CODE, PROC_CODE, '', EMPL_CODE, 0)) #수동 인쇄
        self.auto_print_btn.clicked.connect(lambda: DaesungFunctions.jackupPrint(self, WC_CODE, PROC_CODE, '', EMPL_CODE, 1)) #자동1 인쇄
        self.auto2_print_btn.clicked.connect(lambda: DaesungFunctions.jackupPrint(self, WC_CODE, PROC_CODE, '', EMPL_CODE, 2)) #자동2 인쇄
        
        self.print_btn.clicked.connect(self.printLabel) #라벨 발행
        
        # 화살표버튼 --------------------------------------------------
        self.top.clicked.connect(lambda: DaesungFunctions.topData(self))
        self.prev.clicked.connect(lambda: DaesungFunctions.prevData(self))
        self.next.clicked.connect(lambda: DaesungFunctions.nextData(self))
        self.bottom.clicked.connect(lambda: DaesungFunctions.bottomData(self))
        
        # 상단버튼 --------------------------------------------------
        self.date_btn.clicked.connect(lambda: DaesungFunctions.calendar(self))
        self.set_btn.clicked.connect(self.setData)
        self.log_btn.clicked.connect(self.logWindow)
        self.as_btn.clicked.connect(lambda: DaesungFunctions.openUrl(self))
        self.exit_btn.clicked.connect(self.exitWindow)
    
    def hwConnect(self):
        try:
            self.set_win = SetWindow()
            self.connectPrinter() #바코드 프린터 연결
            self.connectLight() #경광등 연결
        except: logging.debug("hwConnect : 실패")
    
    def connectPrinter(self):
        try: self.p_con_th.terminate()
        except: pass
        #-------------------------------------------------------------
        self.printer_ip = self.set_win.printer_ip_input.text()
        self.printer_port = self.set_win.printer_port_input.text()
        #-------------------------------------------------------------
        if self.set_win.printer_check.isChecked():
            try:
                self.mode =  '^' + self.set_win.printer_mode_input.text().replace(' ', '')
                #-------------------------------------------------------------
                print_socket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
                print_socket.settimeout(0.5)
                print_socket.connect((self.printer_ip, int(self.printer_port)))
                print_socket.close()
                #-------------------------------------------------------------
                self.printer_flag = "success"
                logging.debug("connectPrinter : 프린트({0}:{1}) 연결 성공".format(self.printer_ip, self.printer_port))
                self.print_status.setStyleSheet("background-color: #55cba7;") #green
            except:
                self.printer_flag = "failed"
                logging.debug("connectPrinter : 프린트({0}:{1}) 연결 실패".format(self.printer_ip, self.printer_port))
                self.print_status.setStyleSheet("background-color: #fd97a5;") #red
                try:
                    self.p_con_th = SocketThread(self.printer_ip, int(self.printer_port))
                    self.p_con_th.sig_data.connect(self.PrinterConSlot)
                    self.p_con_th.start()
                except: logging.debug("connectPrinter : printer_con_th 실패")
        else:
            self.printer_flag = "unable"
            logging.debug("connectPrinter : 프린트({0}:{1}) 비활성".format(self.printer_ip, self.printer_port))
            self.print_status.setStyleSheet("background-color: #CDCDCD;") #gray
    
    def connectLight(self):
        global light_ser
        #--------------------------------------------------
        try: light_ser.close()
        except: pass
        #--------------------------------------------------
        light_port = self.set_win.light_port_input.text()
        if self.set_win.light_check.isChecked():
            light_rate = int(self.set_win.light_rate_input.text())
            try:
                light_ser = serial.Serial(light_port, light_rate, timeout = 0.5)
                self.light_flag = "success"
                logging.debug("connectLight : 경광등(%s) 연결 성공"%light_port)
            except:
                self.light_flag = "failed"
                logging.debug("connectLight : 경광등(%s) 연결 실패"%light_port)
        else:
            self.light_flag = "unable"
            logging.debug("connectLight : 경광등(%s) 비활성"%light_port)
    
    #---------------------------------------------------------------------------------------------------
    #미출력 LOT 알림
    def printAlarm(self, index):
        try:
            global PRT_INDEX
            if index != PRT_INDEX:
                PRT_INDEX = index
                if self.light_flag == 'success':
                    try:
                        light_ser.write('RY 1 1\r'.encode()) #red light
                        logging.debug("printAlarm : RY 1 1 성공")
                    except: logging.debug("printAlarm : RY 1 1 실패")
                    MessageWindow(self, '미출력 로트 : [ %s ]'%self.tableWidget.item(index, 2).text()).showModal()
                    try:
                        light_ser.write('RY 1 0\r'.encode()) #green light
                        logging.debug("printAlarm : RY 1 0 성공")
                    except: logging.debug("printAlarm : RY 1 0 실패")
                else: MessageWindow(self, '미출력 로트 : [ %s ]'%self.tableWidget.item(index, 2).text()).showModal()
        except: pass
    
    def setCheckFlag(self, flag):
        try:
            self.set_check = []
            if flag == 1:
                for count, checkbox in enumerate(self.checkBoxList):
                    if checkbox.isChecked() == True: self.set_check.append(count)
        except: pass
    
    def DBload(self):
        self.tableWidget.clearSelection()
        self.tableWidget.setRowCount(0)
        #----------------------------------------------------------------------------
        self.checkBoxList, qty, prt_index = [], 0, -1
        make_qty, prt_num, auto_prt_num, auto2_prt_num = 0, 0, 0, 0
        #----------------------------------------------------------------------------
        if self.s_flag == 'L': M_rows = DaesungQuery.selectFrameLotList(self, self.s_date)
        else: M_rows = DaesungQuery.selectFrameList(self, self.s_date, PROC_CODE, self.WIDX, self.W_DATA)
        #----------------------------------------------------------------------------
        if M_rows == 'failed': self.connectDBThread()
        else:
            if M_rows == ():
                self.tableWidget.setRowCount(1)
                self.tableWidget.setRowHeight(0, 85)
                self.tableWidget.setSpan(0, 0, 1, 15)
                item_data = QTableWidgetItem("해당일자의 데이터가 없습니다.")
                item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                self.tableWidget.setItem(0, 0, item_data)
                self.qty_label.setText('수량 합계 : 0')
                self.c_data = 0
            else:
                try:
                    global PRT_INDEX
                    self.tableWidget.setRowCount(len(M_rows))
                    for i in range(len(M_rows)):
                        self.tableWidget.setRowHeight(i, 85)
                        ckbox = QCheckBox()
                        ckbox.setStyleSheet(t_checkStyle)
                        self.checkBoxList.append(ckbox)
                        if self.s_flag == 'L':
                            if M_rows[i]['MES_PRT_FLAG_QTY'] != None: prt_num = prt_num + int(M_rows[i]['MES_PRT_FLAG_QTY'])
                            if M_rows[i]['MES_PRT_AUTO_FLAG_QTY'] != None: auto_prt_num = auto_prt_num + int(M_rows[i]['MES_PRT_AUTO_FLAG_QTY'])
                            if M_rows[i]['MES_PRT_AUTO_FLAG_QTY2'] != None: auto2_prt_num = auto2_prt_num + int(M_rows[i]['MES_PRT_AUTO_FLAG_QTY2'])
                            make_qty = make_qty + (int(M_rows[i]['QTY']) - int(M_rows[i]['LK_MAKE_QTY']))
                        for count, j in enumerate(self.col):
                            if j == 'CHECK':
                                cellWidget = QWidget()
                                layoutCB = QHBoxLayout(cellWidget)
                                layoutCB.addWidget(self.checkBoxList[i])
                                layoutCB.setAlignment(QtCore.Qt.AlignCenter)
                                layoutCB.setContentsMargins(0, 0, 0, 0)
                                cellWidget.setLayout(layoutCB)
                                self.tableWidget.setCellWidget(i, count, cellWidget)
                                if i in self.set_check: self.checkBoxList[i].setChecked(True)
                            elif j == 'LK_MAKE_FLAG':
                                states = QPushButton()
                                if M_rows[i]['LK_MAKE_FLAG'] == '00': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/대기.png);")
                                elif M_rows[i]['LK_MAKE_FLAG'] == '01': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/진행.png);")
                                elif M_rows[i]['LK_MAKE_FLAG'] == '90': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/완료.png);")
                                self.state_group.addButton(states, i)
                                self.tableWidget.setCellWidget(i, count, states)
                            elif j == 'BAR_FLAG':
                                states = QPushButton()
                                if M_rows[i]['BAR_FLAG'] == '0': states.setStyleSheet("background: none; border: none;")
                                else: states.setStyleSheet(stateBtnStyle + "background-image: url(./img/완료.png);")
                                self.state_group.addButton(states, i)
                                self.tableWidget.setCellWidget(i, count, states)
                            elif j == 'MES_PRT_FLAG' or j == 'MES_PRT_FLAG_ALL':
                                states = QPushButton()
                                if M_rows[i][j] == '1': states.setStyleSheet(stateBtnStyle + "background-color: #fff5db; border-radius:0px; background-image: url(./img/완료.png);")
                                elif M_rows[i][j] == '2': states.setStyleSheet(stateBtnStyle + "background-color: #fff5db; border-radius:0px; background-image: url(./img/진행.png);")
                                else: states.setStyleSheet("background-color: #fff5db; border: none;")
                                self.select_group.addButton(states, i)
                                self.tableWidget.setCellWidget(i, count, states)
                            elif j == 'MES_PRT_AUTO_FLAG':
                                states = QPushButton()
                                if M_rows[i][j] == '1' or M_rows[i][j] == '2':
                                    qty1, qty2 = M_rows[i]['MES_PRT_AUTO_FLAG_QTY'], M_rows[i]['MES_PRT_AUTO_FLAG_QTY2']
                                    if count == 12 and qty1 != None and int(qty1) > 0: states.setStyleSheet(stateBtnStyle + "background-color: #fff5db; border-radius:0px; background-image: url(./img/완료.png);")
                                    elif count == 13 and qty2 != None and int(qty2) > 0: states.setStyleSheet(stateBtnStyle + "background-color: #fff5db; border-radius:0px; background-image: url(./img/완료.png);")
                                    else: states.setStyleSheet("background-color: #fff5db; border: none;")
                                else: states.setStyleSheet("background-color: #fff5db; border: none;")
                                self.select_group.addButton(states, i)
                                self.tableWidget.setCellWidget(i, count, states)
                            elif j == 'LK_PRT_FLAG':
                                states = QPushButton()
                                print_data = M_rows[i][j]
                                if print_data == '0' or print_data == '00': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/대기.png);")
                                elif print_data == '2' or print_data == '01': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/진행.png);")
                                elif print_data == '1' or print_data == '02': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/완료.png);")
                                self.prt_group.addButton(states, i)
                                self.tableWidget.setCellWidget(i, count, states)
                            elif j == '': self.tableWidget.setItem(i, count, QTableWidgetItem(''))
                            elif j == 'MES_PRT_INDEX':
                                if M_rows[i]['MES_PRT_FLAG'] == '0' and M_rows[i]['MES_PRT_AUTO_FLAG'] == '0': 
                                    print_data = ''
                                    prt_index = i
                                    #if prt_index == -1: prt_index = i
                                else: print_data = M_rows[i]['MES_PRT_FLAG_ALL']
                                item_data = QTableWidgetItem(str(print_data))
                                self.tableWidget.setItem(i, count, item_data)
                            else:
                                print_data = M_rows[i][j]
                                if print_data == None: print_data = ''
                                elif j == 'JAKUP_FLAG' and print_data == '1': print_data = "시판"
                                elif j == 'JAKUP_FLAG' and print_data == '2': print_data = "LX"
                                elif j == 'LOT_NUMB' and self.s_flag == 'D': print_data = print_data + '\n' + M_rows[i]['REG_SEQ']
                                elif j == 'CAL_HOLE_VALUE': print_data = int(print_data)
                                elif j == 'QTY':
                                    print_data = int(print_data)
                                    qty = qty + print_data
                                elif j == 'QTY_NO_ALL': print_data =  str(M_rows[i]['SEQ_QTY'])+ '/' + str(int(print_data))
                                item_data = QTableWidgetItem(str(print_data))
                                if j == 'HOPE_TEXT' or j == 'ITEM_TEXT' or j == 'ITEM_NAME' or j == 'SPCL_NAME' or j == 'KYU': item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                                elif j == 'CAL_HOLE_VALUE' or j == 'QTY_NO_ALL' or j == 'QTY': item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                                else: item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                                self.tableWidget.setItem(i, count, item_data)
                except: logging.debug("DBload : table 로드 실패")
            if self.s_flag == 'L':
                if self.reload_num == 0:
                    try:
                        self.th_rowCount = lotLoadThread(20)
                        self.th_rowCount.sig_data.connect(self.newData)
                        self.th_rowCount.start()
                        self.reload_num = 1
                    except: logging.debug("DBload : th_rowCount 실패")
                self.qty_label.setText('총 합계 {0}/{1}, 수동합계 {2}, 자동1합계 {3}, 자동2합계 {4}'.format(make_qty, qty, prt_num, auto_prt_num, auto2_prt_num))
                if prt_index > -1: self.printAlarm(prt_index) #미출력 LOT 알림
                else: PRT_INDEX = -1
            else: self.qty_label.setText('수량 합계 : %d'%len(M_rows))
        DaesungFunctions.tableWidth(self, 'FRAME', self.s_flag, len(M_rows))
    
    #로트/자동자재 라디오 버튼 변경
    def flagChanged(self, flag):
        try: self.th_rowCount.terminate()
        except: pass
        self.s_flag = flag
        self.tableWidget.hideColumn(1)
        self.tableWidget.hideColumn(10)
        if self.s_flag == 'L':
            self.col = ['CHECK', 'JAKUP_FLAG', 'LOT_NUMB', 'REG_NO', 'HOPE_TEXT', '', '', 'MES_PRT_INDEX', 'QTY', 'LK_MAKE_FLAG', 'MES_PRT_FLAG_ALL', 'MES_PRT_FLAG', 'MES_PRT_AUTO_FLAG', 'MES_PRT_AUTO_FLAG', 'LK_PRT_FLAG']
            self.stackedWidget.setCurrentWidget(self.lot_page)
            self.tableWidget.showColumn(0)
            self.tableWidget.showColumn(11)
            self.tableWidget.showColumn(12)
            self.tableWidget.showColumn(13)
            for i in [3, 5, 6, 7]: self.tableWidget.hideColumn(i)
            self.setCheckFlag(0)
            DaesungFunctions.replaceDate(self) #DB로드
        else:
            self.col = ['', '', 'LOT_NUMB', 'BUYER_NAME', 'ITEM_NAME', 'KYU', 'SET_FLAG', 'SPCL_NAME', 'QTY_NO_ALL', 'BAR_FLAG', '', '', '', '', '']
            self.stackedWidget.setCurrentWidget(self.jajae_page)
            self.tableWidget.hideColumn(0)
            self.tableWidget.hideColumn(11)
            self.tableWidget.hideColumn(12)
            self.tableWidget.hideColumn(13)
            for i in [3, 5, 6, 7]: self.tableWidget.showColumn(i)
            if WC_CODE == '09': self.ItemWidxload()
    
    #규격 조회
    def ItemWidxload(self):
        try:
            self.widx_combo.clear()
            s_date = self.date_btn.text().replace(' ', '').replace('-', '')
            #----------------------------------------------------------------------------
            Q_rows = DaesungQuery.selectItemList(self, '%', s_date)
            #----------------------------------------------------------------------------
            if Q_rows == ():
                self.widx_combo.addItem('폼목없음')
            elif Q_rows != 'failed':
                self.widx_combo.addItem('전체')
                for i in range(len(Q_rows)): self.widx_combo.addItem(str(int(Q_rows[i]['FWIDX'])) + 'mm')
        except: pass
    
    #규격 변경
    def widxChanged(self):
        try:
            self.WIDX = ''
            FWIDX = self.widx_combo.currentText().replace('mm', '')
            if FWIDX != '' and FWIDX != '폼목없음':
                if FWIDX == '전체': self.WIDX = ''
                else: self.WIDX = 'AND MESF.WIDX = %s'%FWIDX
            self.setCheckFlag(0)
            DaesungFunctions.replaceDate(self) #DB로드
        except: pass
    
    def searchData(self):
        s_data = '%' + self.search_input.text() + '%'
        if self.search_combo.currentIndex() == 0: self.W_DATA = "AND BUYER.BUYER_NAME LIKE '{0}'".format(s_data)
        elif self.search_combo.currentIndex() == 1: self.W_DATA = "AND MJAKUP.LOT_NUMB LIKE '{0}'".format(s_data)
        elif self.search_combo.currentIndex() == 2: self.W_DATA = "AND DJAKUP.SET_FLAG LIKE '{0}'".format(s_data)
        elif self.search_combo.currentIndex() == 3: self.W_DATA = "AND SPCL.SPCL_NAME LIKE '{0}'".format(s_data)
        self.setCheckFlag(0)
        DaesungFunctions.replaceDate(self) #DB로드
    
    def changedPrtFlag(self, prt, auto_prt):
        try:
            if self.s_flag == 'L':
                PRT_DATE = time.strftime('%Y%m%d')
                checkArray = []
                for count, checkbox in enumerate(self.checkBoxList):
                    if checkbox.isChecked() == True: checkArray.append(count)
                if checkArray != []:
                    for i in checkArray:
                        REG_NO = self.tableWidget.item(i, 3).text()
                        D_rows = DaesungQuery.selectFrameDetailList(self, REG_NO, self.s_date, PROC_CODE)
                        for row in D_rows:
                            PRT_TIME = time.strftime('%H%M%S')
                            #----------------------------------------------------------------------------
                            DaesungQuery.updateFrameDetailFlag(self, prt, auto_prt, row['REG_NO'], row['REG_SEQ'], PRT_DATE, PRT_TIME, EMPL_CODE)
                            #----------------------------------------------------------------------------
                        result = DaesungQuery.DBCommit(self)
                        if result == 1:
                            logging.debug("changedPrtFlag : updatePrtFlag 성공")
                            #----------------------------------------------------------------------------
                            prt_flag = DaesungQuery.selectFramePrtFlag(self, REG_NO)
                            #----------------------------------------------------------------------------
                            result = DaesungQuery.updateFrameFlag(self, prt_flag['MES_PRT_FLAG'], PRT_DATE, PRT_TIME, EMPL_CODE, REG_NO)
                            #----------------------------------------------------------------------------
                            if result == 1: logging.debug("changedPrtFlag : updateFrameFlag 성공")
                            else: logging.debug("changedPrtFlag : updateFrameFlag 실패")
                        else: logging.debug("changedPrtFlag : updatePrtFlag 실패")
                    self.setCheckFlag(0)
                    self.DBload() #DB로드
        except: logging.debug("changedPrtFlag : failed")
    
    #라벨 발행
    def printLabel(self):
        if self.s_flag == 'L':
            checkArray = []
            if self.printer_flag == "success":
                for count, checkbox in enumerate(self.checkBoxList):
                    if checkbox.isChecked() == True: checkArray.append(count)
                if checkArray != []:
                    try:
                        self.mysocket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
                        self.mysocket.settimeout(0.5)
                        self.mysocket.connect((self.printer_ip, int(self.printer_port)))
                        self.print_status.setStyleSheet("background-color: #55cba7;") #green
                        try: self.th_rowCount.terminate()
                        except: pass
                        self.fileName = f_label
                        try:
                            for i in checkArray:
                                REG_NO = self.tableWidget.item(i, 3).text()
                                f_name = open(self.fileName, 'r', encoding = 'utf-8')
                                m_textData = f_name.read()
                                f_name.close()
                                if self.set_win.printer_mode_check.isChecked() == True: m_textData = m_textData.replace("^CI28", "^CI28" + self.mode)
                                if self.set_win.printer_po_check.isChecked() == True: m_textData = m_textData.replace("^LS0", "^LS0^POI")
                                elif self.set_win.printer_po_check.isChecked() == False: m_textData = m_textData.replace("^LS0", "^LS0^PON")
                                #----------------------------------------------------------------------------
                                P_rows = DaesungQuery.selectFrameDetaiBarcodelList(self, REG_NO, '%', self.s_date, PROC_CODE)
                                #----------------------------------------------------------------------------
                                if P_rows == 'failed': self.connectDBThread()
                                elif P_rows != []:
                                    for row in P_rows:
                                        textData = m_textData
                                        try:
                                            for t in ['REG_NO', 'LOT_NUMB', 'DATE', 'ITEM_NAME', 'KYU', 'SPCL_NAME', 'BUYER_NAME', 'BUYER_CODE', 'LABEL_BIGO']:
                                                if t == 'DATE':
                                                    print_data = row['REG_DATE']
                                                    print_data = "{0}.{1}.{2}".format(print_data[2:4], print_data[4:6], print_data[6:8])
                                                else:
                                                    print_data = row[t]
                                                    if print_data == None: print_data = ""
                                                textData = textData.replace("{%s}"%t, str(print_data))
                                            self.mysocket.send(textData.encode())
                                            try:
                                                #----------------------------------------------------------------------------
                                                if row['PRT_FLAG'] != 1: DaesungQuery.LABEL_UPDATE_SQL(self, REG_NO, row['REG_SEQ'], row['SEQ_QTY']) #라벨발행 FLAG UPDATE
                                                #----------------------------------------------------------------------------
                                                M_rows = DaesungQuery.selectMakeData(self, row['PROC_CODE'], row['BAR_CODE'])
                                                if M_rows == (): DaesungQuery.PR_SAVE_MAKE_BAR_DETAIL(self, 'insert', '0', EMPL_CODE, self.REG_NO, P_rows[0]['REG_SEQ'], P_rows[0]['SORT_KEY'], P_rows[0]['BAR_CODE'], self.c_date, 1, 0) #실적등록
                                            except: pass
                                        except: logging.debug("printLabel : selectDetailList 실패")
                                    DaesungQuery.DBCommit(self) #DB commit
                                    time.sleep(0.3)
                                else: logging.debug("printLabel : 등록된 바코드 없음")
                            #----------------------------------------------------------------------------
                            DaesungQuery.SELECT_PR_PASS_JAKUP_MAKE(self)
                            #----------------------------------------------------------------------------
                        except: logging.debug("printLabel : select 실패")
                        self.mysocket.close()
                        self.DBload() #DB로드
                        self.th_rowCount.start()
                    except:
                        self.print_status.setStyleSheet("background-color: #fd97a5;") #red
                        logging.debug("printLabel : 프린터 연결 실패")
                else: MessageWindow(self, "출력할 라벨을 선택해주세요.").showModal()
            else: MessageWindow(self, "프린터를 연결해주세요.").showModal()
    
    #----------------------------------------------------------------------------
    @pyqtSlot(int)
    def newData(self, num):
        if num == 1:
            self.setCheckFlag(1)
            self.DBload() #DB로드
    
    def click_row(self, index):
        try:
            if self.s_flag == 'L':
                JAKUP_FLAG = self.tableWidget.item(index.row(), 1).text()
                LOT_NO = self.tableWidget.item(index.row(), 2).text()
                REG_NO = self.tableWidget.item(index.row(), 3).text()
                s_date = self.date_btn.text()
                print_window = MesFrameDetailWindow(JAKUP_FLAG, LOT_NO, REG_NO, s_date) #DETAIL 화면 연결
                try: self.p_con_th.terminate()
                except: pass
                try: self.th_rowCount.terminate()
                except: pass
                widget.addWidget(print_window)
                widget.setCurrentIndex(widget.currentIndex() + 1)
                self.deleteLater()
        except: logging.debug("clickRow : 상세페이지 연결 실패")
    
    def selectedAll(self, num):
        if num == 0 and self.check_flag == False:
            for checkbox in self.checkBoxList: checkbox.setChecked(True)
            self.check_flag = True
        elif num == 0 and self.check_flag == True:
            for checkbox in self.checkBoxList: checkbox.setChecked(False)
            self.check_flag = False
    
    #----------------------------------------------------------------------------
    def showDate(self, date):
        self.calender.hide()
        a = date.toString().split()
        if len(a[1]) == 1: a[1] = "0" + a[1]
        if len(a[2]) == 1: a[2] = "0" + a[2]
        self.date_btn.setText("%s-%s-%s "%(a[3],a[1],a[2])) # 월, 일자가 1자 일때
        self.calendar_flag = False
        global PRT_INDEX
        PRT_INDEX = -1
        self.setCheckFlag(0)
        DaesungFunctions.replaceDate(self) #DB로드
    
    def mousePressEvent(self, event):
        if event.buttons() == Qt.LeftButton and self.calendar_flag == True:
            self.calender.hide()
            self.calendar_flag = False
    
    def splcList(self):
        try: SpclListWindow(self.s_date, self.WIDX, self.W_DATA).showModal()
        except: logging.debug("splcList : 색상정보 페이지 연결실패")
    
    def setData(self):
        try:
            SetWindow().showModal()
            self.hwConnect() #HW연결
        except: logging.debug("setData : 상세페이지 연결 실패")
    
    def logWindow(self):
        try:
            try: self.th_rowCount.terminate()
            except: pass
            MesLogWindow(TodayData).showModal()
            self.th_rowCount.start()
        except: logging.debug("logWindow : 로그 화면 연결 실패")
    
    def connectDBThread(self):
        if self.DB_flag == 0:
            self.DB_flag = 1
            MessageWindow(self, "DB연결 실패").showModal()
            self.DB_th = ConnectDBThread()
            self.DB_th.sig_data.connect(self.DbThreadSlot)
            self.DB_th.start()
    
    @pyqtSlot(int)       
    def DbThreadSlot(self, con):
        if con == 1:
            self.DB_th.terminate()
            self.DB_flag = 0
            #----------------------------------------------------------------------------
            result = DaesungQuery.connectDB(self, host, port, user, name)
            #----------------------------------------------------------------------------
            if result == 'success': DaesungFunctions.replaceDate(self) #DB로드
    
    def exitWindow(self):
        MessageWindow(self, 2).showModal()
        #----------------------------------
        if RECONNECT_FLAG == 1:
            try: self.p_con_th.terminate()
            except: pass
            try: self.th_rowCount.terminate()
            except: pass

################################################################################################################
#문틀부 DETAIL Window
class MesFrameDetailWindow(QDialog):
    def __init__(self, gubun, lot, reg_no, date):
        super(MesFrameDetailWindow, self).__init__()
        loadUi("ui\DAESUNG_MES_FD.ui", self)
        
        DaesungFunctions.setDetailStyle(self, date, WC_CODE, PROC_CODE, gubun, lot, reg_no) #기본 셋팅
        
        self.ORDER = 'DJAKUP.WIDX, DJAKUP.LENX,'
        self.JAKUP_APPR_FLAG = '2'
        self.PROC_CODE, self.W_DATA = "MPJAKUP.PROC_CODE = '{0}'".format(PROC_CODE), ''
        
        self.hwConnect() #HW연결
        self.DBload() #DB로드
        
        self.tableWidget.horizontalHeader().sectionClicked.connect(self.selectedAll)
        self.select_all.clicked.connect(lambda: self.selectedAll(0))
        
        self.print_btn.clicked.connect(self.printLabel) #라벨 발행
        
        self.sudong_ok_btn.clicked.connect(lambda: self.changedPrtFlag('1', '0')) #수동확정
        self.sudong_cancel_btn.clicked.connect(lambda: self.changedPrtFlag('0', '0')) #수동취소
        self.auto_ok_btn.clicked.connect(lambda: self.changedPrtFlag('0', '1')) #자동확정
        self.auto_cancel_btn.clicked.connect(lambda: self.changedPrtFlag('0', '0')) #자동취소
        self.auto2_ok_btn.clicked.connect(lambda: self.changedPrtFlag('0', '2')) #자동확정
        self.auto2_cancel_btn.clicked.connect(lambda: self.changedPrtFlag('0', '0')) #자동취소
        
        self.sudong_print_btn.clicked.connect(lambda: DaesungFunctions.jackupPrint(self, WC_CODE, PROC_CODE, lot, EMPL_CODE, 0)) #수동 인쇄
        self.auto_print_btn.clicked.connect(lambda: DaesungFunctions.jackupPrint(self, WC_CODE, PROC_CODE, lot, EMPL_CODE, 1)) #자동1 인쇄
        self.auto2_print_btn.clicked.connect(lambda: DaesungFunctions.jackupPrint(self, WC_CODE, PROC_CODE, lot, EMPL_CODE, 2)) #자동2 인쇄
        
        self.tableWidget.currentCellChanged.connect(self.connectTable)
        self.tableWidget.cellPressed.connect(self.clickedRow)
        self.tableWidget.cellClicked.connect(self.clickedRow)
        
        # 화살표버튼 --------------------------------------------------
        self.top.clicked.connect(lambda: DaesungFunctions.topData(self))
        self.prev.clicked.connect(lambda: DaesungFunctions.prevData(self))
        self.next.clicked.connect(lambda: DaesungFunctions.nextData(self))
        self.bottom.clicked.connect(lambda: DaesungFunctions.bottomData(self))
        
        # 상단버튼 --------------------------------------------------
        self.set_btn.clicked.connect(self.setData)
        self.log_btn.clicked.connect(self.logWindow)
        self.as_btn.clicked.connect(lambda: DaesungFunctions.openUrl(self))
        self.back_btn.clicked.connect(self.back)
    
    def hwConnect(self):
        try:
            self.set_win = SetWindow()
            self.connectPrinter() #바코드 프린터 연결
        except: logging.debug("hwConnect : 실패")
    
    def connectPrinter(self):
        try: self.p_con_th.terminate()
        except: pass
        #-------------------------------------------------------------
        self.printer_ip = self.set_win.printer_ip_input.text()
        self.printer_port = self.set_win.printer_port_input.text()
        #-------------------------------------------------------------
        if self.set_win.printer_check.isChecked():
            try:
                self.mode =  '^' + self.set_win.printer_mode_input.text().replace(' ', '')
                #-------------------------------------------------------------
                print_socket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
                print_socket.settimeout(0.5)
                print_socket.connect((self.printer_ip, int(self.printer_port)))
                print_socket.close()
                #-------------------------------------------------------------
                self.printer_flag = "success"
                logging.debug("connectPrinter : 프린트({0}:{1}) 연결 성공".format(self.printer_ip, self.printer_port))
                self.print_status.setStyleSheet("background-color: #55cba7;") #green
            except:
                self.printer_flag = "failed"
                logging.debug("connectPrinter : 프린트({0}:{1}) 연결 실패".format(self.printer_ip, self.printer_port))
                self.print_status.setStyleSheet("background-color: #fd97a5;") #red
                try:
                    self.p_con_th = SocketThread(self.printer_ip, int(self.printer_port))
                    self.p_con_th.sig_data.connect(self.PrinterConSlot)
                    self.p_con_th.start()
                except: logging.debug("connectPrinter : printer_con_th 실패")
        else:
            self.printer_flag = "unable"
            logging.debug("connectPrinter : 프린트({0}:{1}) 비활성".format(self.printer_ip, self.printer_port))
            self.print_status.setStyleSheet("background-color: #CDCDCD;") #gray
    
    #---------------------------------------------------------------------------------------------------
    def setCheckFlag(self):
        try:
            self.set_check = []
            for count, checkbox in enumerate(self.checkBoxList):
                if checkbox.isChecked() == True: self.set_check.append(count)
        except: pass
    
    def DBload(self):
        self.setCheckFlag()
        self.tableWidget.clearSelection()
        self.tableWidget.setRowCount(0)
        self.checkBoxList = []
        #----------------------------------------------------------------------------
        D_rows = DaesungQuery.selectFrameDetailList(self, self.REG_NO, self.s_date, PROC_CODE)
        #----------------------------------------------------------------------------
        if D_rows == 'failed': self.connectDBThread()
        elif D_rows == ():
            logging.debug("DBload : 작업지시서 취소됨")
            MessageWindow(self, "해당 작업지시서가 취소되었습니다.").showModal()
            self.back()
        else:
            try:
                make_qty, qty, prt_num, auto_prt_num, auto2_prt_num = 0, 0, 0, 0, 0
                self.tableWidget.setRowCount(len(D_rows))
                for i in range(len(D_rows)):
                    self.tableWidget.setRowHeight(i, 85)
                    ckbox = QCheckBox()
                    ckbox.setStyleSheet(t_checkStyle)
                    self.checkBoxList.append(ckbox)
                    #----------------------------------------------------------------------------
                    make_qty = make_qty + (int(D_rows[i]['QTY']) - int(D_rows[i]['LK_MAKE_QTY']))
                    #----------------------------------------------------------------------------
                    for count, j in enumerate(['CHECK', 'REG_SEQ', 'BUYER_NAME', 'ITEM_NAME', 'SPCL_NAME', 'KYU', 'QTY', 'MES_PRT_FLAG', 'MES_PRT_AUTO_FLAG', 'MES_PRT_AUTO_FLAG', 'LK_PRT_FLAG', '']):
                        if j == 'CHECK':
                            cellWidget = QWidget()
                            layoutCB = QHBoxLayout(cellWidget)
                            layoutCB.addWidget(self.checkBoxList[i])
                            layoutCB.setAlignment(QtCore.Qt.AlignCenter)
                            layoutCB.setContentsMargins(0, 0, 0, 0)
                            cellWidget.setLayout(layoutCB)
                            self.tableWidget.setCellWidget(i, count, cellWidget)
                            if i in self.set_check: self.checkBoxList[i].setChecked(True)
                        elif j == 'MES_PRT_FLAG':
                            states = QPushButton()
                            if D_rows[i]['MES_PRT_FLAG'] == '1':
                                states.setStyleSheet(stateBtnStyle + " background-image: url(./img/완료.png);")
                                prt_num += int(D_rows[i]['QTY'])
                            elif D_rows[i]['MES_PRT_FLAG'] == '2':
                                states.setStyleSheet(stateBtnStyle + "background-image: url(./img/진행.png);")
                                prt_num += int(D_rows[i]['QTY'])
                            else: states.setStyleSheet("background: none; border: none;")
                            self.select_group.addButton(states, i)
                            self.tableWidget.setCellWidget(i, count, states)
                        elif j == 'MES_PRT_AUTO_FLAG':
                            states = QPushButton()
                            if count == 8 and D_rows[i]['MES_PRT_AUTO_FLAG'] == '1':
                                states.setStyleSheet(stateBtnStyle + "background-image: url(./img/완료.png);")
                                auto_prt_num += int(D_rows[i]['QTY'])
                            elif count == 9 and D_rows[i]['MES_PRT_AUTO_FLAG'] == '2':
                                states.setStyleSheet(stateBtnStyle + "background-image: url(./img/완료.png);")
                                auto2_prt_num += int(D_rows[i]['QTY'])
                            else: states.setStyleSheet("background: none; border: none;")
                            self.select_group.addButton(states, i)
                            self.tableWidget.setCellWidget(i, count, states)
                        elif j == 'LK_PRT_FLAG':
                            states = QPushButton()
                            print_data = D_rows[i][j]
                            if print_data == '0' or print_data == '00': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/대기.png);")
                            elif print_data == '2' or print_data == '01': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/진행.png);")
                            elif print_data == '1' or print_data == '02': states.setStyleSheet(stateBtnStyle + "background-image: url(./img/완료.png);")
                            self.prt_group.addButton(states, i)
                            self.tableWidget.setCellWidget(i, count, states)
                        elif j == '': self.tableWidget.setItem(i, count, QTableWidgetItem(''))
                        else:
                            print_data = D_rows[i][j]
                            if print_data == None: print_data = ''
                            elif j == 'QTY':
                                print_data = int(print_data)
                                qty = qty + print_data
                            item_data = QTableWidgetItem(str(print_data))
                            if j == 'REG_SEQ': item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                            elif j == 'KYU' or j == 'QTY': item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                            self.tableWidget.setItem(i, count, item_data)
                DaesungFunctions.tableWidth(self, PROC_CODE, WC_CODE, len(D_rows))
                self.qty_label.setText('총 합계 {0}/{1}, 수동합계 {2}, 자동1합계 {3}, 자동2합계 {4}'.format(make_qty, qty, prt_num, auto_prt_num, auto2_prt_num))
                if self.reload_num == 0:
                    try:
                        self.th_rowCount = lotLoadThread(20)
                        self.th_rowCount.sig_data.connect(self.newData)
                        self.th_rowCount.start()
                        self.reload_num = 1
                    except: logging.debug("DBload : th_rowCount 실패")
            except: logging.debug("DBload : table 로드 실패")
    
    def changedPrtFlag(self, prt, auto_prt):
        try:
            PRT_DATE = time.strftime('%Y%m%d')
            checkArray = []
            for count, checkbox in enumerate(self.checkBoxList):
                if checkbox.isChecked() == True: checkArray.append(count)
            if checkArray != []:
                for i in checkArray:
                    REG_SEQ = self.tableWidget.item(i, 1).text()
                    PRT_TIME = time.strftime('%H%M%S')
                    #----------------------------------------------------------------------------
                    DaesungQuery.updateFrameDetailFlag(self, prt, auto_prt, self.REG_NO, REG_SEQ, PRT_DATE, PRT_TIME, EMPL_CODE)
                    #----------------------------------------------------------------------------
                result = DaesungQuery.DBCommit(self)    
                if result == 1:
                    logging.debug("changedPrtFlag : updatePrtFlag 성공")
                    #----------------------------------------------------------------------------
                    prt_flag = DaesungQuery.selectFramePrtFlag(self, self.REG_NO)
                    #----------------------------------------------------------------------------
                    result = DaesungQuery.updateFrameFlag(self, prt_flag['MES_PRT_FLAG'], PRT_DATE, PRT_TIME, EMPL_CODE, self.REG_NO)
                    #----------------------------------------------------------------------------
                    if result == 1: logging.debug("changedPrtFlag : updateFrameFlag 성공")
                    else: logging.debug("changedPrtFlag : updateFrameFlag 실패")
                else: logging.debug("changedPrtFlag : updatePrtFlag 실패")
                self.DBload()
        except: logging.debug("changedPrtFlag : failed")
    
    #라벨 발행
    def printLabel(self):
        checkArray = []
        if self.printer_flag == "success":
            for count, checkbox in enumerate(self.checkBoxList):
                if checkbox.isChecked() == True: checkArray.append(count)
            if checkArray != []:
                try:
                    self.mysocket = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
                    self.mysocket.settimeout(0.5)
                    self.mysocket.connect((self.printer_ip, int(self.printer_port)))
                    self.print_status.setStyleSheet("background-color: #55cba7;") #green
                    try: self.th_rowCount.terminate()
                    except: pass
                    self.fileName = f_label
                    try:
                        for i in checkArray:
                            REG_SEQ = self.tableWidget.item(i, 1).text()
                            f_name = open(self.fileName, 'r', encoding = 'utf-8')
                            m_textData = f_name.read()
                            f_name.close()
                            if self.set_win.printer_mode_check.isChecked() == True: m_textData = m_textData.replace("^CI28", "^CI28" + self.mode)
                            if self.set_win.printer_po_check.isChecked() == True: m_textData = m_textData.replace("^LS0", "^LS0^POI")
                            elif self.set_win.printer_po_check.isChecked() == False: m_textData = m_textData.replace("^LS0", "^LS0^PON")
                            #----------------------------------------------------------------------------
                            P_rows = DaesungQuery.selectFrameDetaiBarcodelList(self, self.REG_NO, REG_SEQ, self.s_date, PROC_CODE)
                            #----------------------------------------------------------------------------
                            if P_rows == 'failed': self.connectDBThread()
                            elif P_rows != []:
                                for row in P_rows:
                                    textData = m_textData
                                    try:
                                        for t in ['REG_NO', 'LOT_NUMB', 'DATE', 'ITEM_NAME', 'KYU', 'SPCL_NAME', 'BUYER_NAME', 'BUYER_CODE', 'LABEL_BIGO']:
                                            if t == 'DATE':
                                                print_data = row['REG_DATE']
                                                print_data = "{0}.{1}.{2}".format(print_data[2:4], print_data[4:6], print_data[6:8])
                                            else:
                                                print_data = row[t]
                                                if print_data == None: print_data = ""
                                            textData = textData.replace("{%s}"%t, str(print_data))
                                        self.mysocket.send(textData.encode())
                                        try:
                                            #----------------------------------------------------------------------------
                                            if row['PRT_FLAG'] != 1: DaesungQuery.LABEL_UPDATE_SQL(self, self.REG_NO, row['REG_SEQ'], row['SEQ_QTY']) #라벨발행 FLAG UPDATE
                                            #----------------------------------------------------------------------------
                                            M_rows = DaesungQuery.selectMakeData(self, row['PROC_CODE'], row['BAR_CODE'])
                                            if M_rows == (): DaesungQuery.PR_SAVE_MAKE_BAR_DETAIL(self, 'insert', '0', EMPL_CODE, self.REG_NO, P_rows[0]['REG_SEQ'], P_rows[0]['SORT_KEY'], P_rows[0]['BAR_CODE'], self.c_date, 1, 0) #실적등록
                                        except: pass
                                    except: logging.debug("printLabel : selectDetailList 실패")
                                DaesungQuery.DBCommit(self) #DB commit
                                time.sleep(0.3)
                            else: logging.debug("printLabel : 등록된 바코드 없음")
                        #----------------------------------------------------------------------------
                        DaesungQuery.SELECT_PR_PASS_JAKUP_MAKE(self)
                        #----------------------------------------------------------------------------
                    except: logging.debug("printLabel : select 실패")
                    self.mysocket.close()
                    self.DBload() #DB로드
                    self.th_rowCount.start()
                except:
                    self.print_status.setStyleSheet("background-color: #fd97a5;") #red
                    logging.debug("printLabel : 프린터 연결 실패")
            else: MessageWindow(self, "출력할 라벨을 선택해주세요.").showModal()
        else: MessageWindow(self, "프린터를 연결해주세요.").showModal()
    
    #---------------------------------------------------------------------------------------------------
    def connectTable(self, row):
        self.checkBoxList[row].setChecked(True)
    
    def clickedRow(self, row):
        s_index, s_row = -1, []
        for item in self.tableWidget.selectedIndexes():
            if s_index != item.row():
                    s_index = item.row()
                    s_row.append(s_index)
        for i in range(self.tableWidget.rowCount()):
            if i in s_row: self.checkBoxList[row].setChecked(True)
            else: self.checkBoxList[i].setChecked(False)
        self.checkBoxList[row].setChecked(True)
    
    def selectedAll(self, num):
        if num == 0 and self.check_flag == False:
            for checkbox in self.checkBoxList: checkbox.setChecked(True)
            self.check_flag = True
        elif num == 0 and self.check_flag == True:
            for checkbox in self.checkBoxList: checkbox.setChecked(False)
            self.check_flag = False
    
    #---------------------------------------------------------------------------------------------------
    @pyqtSlot(int)
    def newData(self, num):
        if num == 1: self.DBload() #DB로드
    
    #---------------------------------------------------------------------------------------------------
    def setData(self):
        try:
            SetWindow().showModal()
            self.hwConnect() #HW연결
        except: logging.debug("setData : 상세페이지 연결 실패")
    
    def logWindow(self):
        try:
            try: self.th_rowCount.terminate()
            except: pass
            MesLogWindow(TodayData).showModal()
            self.th_rowCount.start()
        except: logging.debug("logWindow : 로그 화면 연결 실패")
    
    #실시간 DB연결
    def connectDBThread(self):
        self.DB_btn.setStyleSheet("background-color: #fd97a5;") #red
        if self.DB_flag == 0:
            self.DB_flag = 1
            logging.debug("DBload : DB로드 실패")
            MessageWindow(self, "DB연결 실패").showModal()
            #----------------------------------------------------------------------------
            self.DB_th = ConnectDBThread()
            self.DB_th.sig_data.connect(self.DbThreadSlot)
            self.DB_th.start()
    
    @pyqtSlot(int)
    def DbThreadSlot(self, con):
        if con == 1:
            self.DB_th.terminate()
            self.DB_flag = 0
            #----------------------------------------------------------------------------
            result = DaesungQuery.connectDB(self, host, port, user, name)
            #----------------------------------------------------------------------------
            if result == 'success':
                self.connectPrinter() #바코드 프린터 연결
                self.DBload() #DB로드
    
    def back(self):
        try: self.p_con_th.terminate()
        except: pass
        try: self.th_rowCount.terminate()
        except: pass
        set_date = self.date_btn.text()
        widget.addWidget(MesFrameWindow(set_date))
        widget.setCurrentIndex(widget.currentIndex() + 1)
        self.deleteLater()

################################################################################################################
#문틀부 Window > 색상정보 Window
class SpclListWindow(QDialog):
    def __init__(self, date, widx, w_data):
        super(SpclListWindow, self).__init__()
        loadUi("ui\ListTable.ui", self)
        
        self.setFixedSize(800, 750)
        
        self.s_date, self.WIDX, self.W_DATA = date, widx, w_data
        self.WIDX_INT = self.WIDX.replace('AND MESF.WIDX = ', '')
        self.tableWidget.setStyleSheet(tableStyle)
        
        self.qtyChanged(1)
        
        self.cut_1_radio.clicked.connect(lambda: self.qtyChanged(1))
        self.cut_2_radio.clicked.connect(lambda: self.qtyChanged(2))
    
    def qtyChanged(self, qty):
        self.qty = qty
        if self.qty == 1: self.MESF = self.WIDX
        elif self.WIDX_INT == '': self.MESF = "AND MESF.WIDX IN (110, 130, 140) AND DJAKUP.SET_FLAG = 'Y'"
        elif int(self.WIDX_INT) > 140: self.MESF = 'X'
        else: self.MESF = self.WIDX + " AND DJAKUP.SET_FLAG = 'Y'"
        self.DBload()
    
    def AutoListData(self, S_rows):
        try:
            DaesungFunctions.dataListReset(self)
            for count, row in enumerate(S_rows):
                if row['SPCL_NAME'] == None: SPCL_NAME = ''
                else: SPCL_NAME = row['SPCL_NAME']
                #----------------------------------------------------------------------------
                ITEM_NAME = row['ITEM_NAME']
                lenx, widx, tikx, qty =  int(row['LENX']), int(row['WIDX']), int(re.findall(r'[0-9]+', ITEM_NAME)[0]), int(row['QTY']) #좌우, 상하, 두께, 수량
                jangbar = int(row['FLENX'])
                #----------------------------------------------------------------------------
                if row['SET_FLAG'] == 'Y': set34 = 4
                else: set34 = 3
                #----------------------------------------------------------------------------
                self.jangb.append(jangbar)
                self.tikxs.append(tikx)
                self.spcls.append(SPCL_NAME)
                if qty > 0:
                    c_qty, addnum = 0, 60 # 양방향 상하 컷팅시 카운트 초기화, 추가여분 사이즈
                    DaesungFunctions.kyulist(self, set34, lenx, widx, qty) # 규격*수량 정보
                    DaesungFunctions.setlist1(self, set34, lenx, widx) # 컷팅규격
                    DaesungFunctions.setlist2(self, self.set1) # 남은컷팅규격
                    if count != 0 and self.jangb[count - 1] != jangbar: self.cutbar = jangbar
                    if count != 0 and self.tikxs[count - 1] != tikx: self.cutbar = jangbar
                    if count != 0 and self.spcls[count - 1] != SPCL_NAME: self.cutbar = jangbar
                    if self.cutbar < min(self.kyus) + addnum: self.cutbar = jangbar #장바길이가 규격의 최소값보다 작은경우 새 장바 투입
                    for cut in range(len(self.kyus)):
                        if self.qty == 1: # --------------------------    1.한방향 컷팅     --------------------------
                            if cut % set34 == 0: c_qty += 1
                            if set34 == 3: setrow = (cut + 1) % 3
                            else: setrow = (cut + 1) % 4
                            if setrow == 0:
                                if self.cutbar < self.set2[0] + addnum: self.cutbar = jangbar
                                if self.cutbar >= self.set2[0] + addnum:
                                    cutkyu = self.set2[0]
                                    self.cutbar = self.cutbar - cutkyu
                                    self.set2.remove(cutkyu)
                                    #----------------------------------------------------------------------------
                                    if cutkyu == lenx: DaesungFunctions.datalist(self, self.cutbar + lenx, tikx, SPCL_NAME, jangbar)
                                    else: DaesungFunctions.datalist(self, self.cutbar + widx, tikx, SPCL_NAME, jangbar)
                                    #----------------------------------------------------------------------------
                                DaesungFunctions.setlist2(self, self.set1) # 한방향 모든요소 컷팅 완료후 컷팅규격으로 남은규격 정보를 가져온다.
                            else:
                                if self.cutbar < min(self.set2) + addnum: self.cutbar = jangbar
                                if self.cutbar >= max(self.set2) + addnum:
                                    cutkyu = max(self.set2)
                                    self.cutbar = self.cutbar - cutkyu
                                    self.set2.remove(cutkyu)
                                    #----------------------------------------------------------------------------
                                    if cutkyu == lenx: DaesungFunctions.datalist(self, self.cutbar + lenx, tikx, SPCL_NAME, jangbar)
                                    else: DaesungFunctions.datalist(self, self.cutbar + widx, tikx, SPCL_NAME, jangbar)
                                    #----------------------------------------------------------------------------
                                elif self.cutbar >= min(self.set2) + addnum:
                                    cutkyu = min(self.set2)
                                    self.cutbar = self.cutbar - cutkyu
                                    self.set2.remove(cutkyu)
                                    #----------------------------------------------------------------------------
                                    if cutkyu == lenx: DaesungFunctions.datalist(self, self.cutbar + lenx, tikx, SPCL_NAME, jangbar)
                                    else: DaesungFunctions.datalist(self, self.cutbar + widx, tikx, SPCL_NAME, jangbar)
                                    #----------------------------------------------------------------------------
                        else:
                            if cut % self.qty == 0: c_qty += 1
                            setrow = (cut + 1) % 2 # 컷팅시 셋트품 (1.단짝 0.완짝) # (cut +1) 단짝은 홀수 완짝은 짝수로 하기위해서 (편의상)
                            if setrow == 1: # 단짝인경우는 좌우, 상하 자를수 있는 규격값을 체크한다.
                                if self.cutbar >= max(self.kyus) + addnum: # 큰 규격부터 컷팅을 진행한다.
                                    self.cutbar, lwflag = self.cutbar - lenx, '1' #좌우
                                    self.set2.remove(lenx) # 사용한 규격 제거
                                    #----------------------------------------------------------------------------
                                    for qty in range(2): DaesungFunctions.datalist(self, self.cutbar + lenx, tikx, SPCL_NAME, jangbar)
                                    #----------------------------------------------------------------------------
                                    continue
                                if self.cutbar >= min(self.kyus) + addnum: # 작은 규격 컷팅을 진행한다.
                                    if set34 != 3: # 양방향 셋트3면 인경우 자른회수 체크 (수량이 많을 경우 홀수인경우 컷팅 짝수인경우 pass)
                                        self.cutbar, lwflag = self.cutbar - widx, '2' #상하
                                        self.set2.remove(widx)
                                        #----------------------------------------------------------------------------
                                        for qty in range(2): DaesungFunctions.datalist(self, self.cutbar + widx, tikx, SPCL_NAME, jangbar)
                                        #----------------------------------------------------------------------------
                                    continue
                                self.cutbar, lwflag = jangbar - lenx, '1' #자르는 규격이 작은경우 버림처리하고 재설정한다./좌우
                                self.set2.remove(lenx)
                                #----------------------------------------------------------------------------
                                for qty in range(2): DaesungFunctions.datalist(self, self.cutbar + lenx, tikx, SPCL_NAME, jangbar)
                                #----------------------------------------------------------------------------
                            else:
                                if lwflag == '1':
                                    if self.cutbar < widx + addnum: self.cutbar = jangbar
                                    self.cutbar, lwflag = self.cutbar - widx, '2' #상하
                                    self.set2.remove(widx)
                                    #----------------------------------------------------------------------------
                                    for qty in range(2): DaesungFunctions.datalist(self, self.cutbar + widx, tikx, SPCL_NAME, jangbar)
                                    #----------------------------------------------------------------------------
                                else: # 자르는 규격이 작은경우 버림처리하고 재설정한다.
                                    if self.cutbar < lenx: self.cutbar = jangbar
                                    self.cutbar, lwflag = self.cutbar - lenx, '1' #좌우
                                    self.set2.remove(lenx)
                                    #----------------------------------------------------------------------------
                                    for qty in range(2): DaesungFunctions.datalist(self, self.cutbar + lenx, tikx, SPCL_NAME, jangbar)
                                    #----------------------------------------------------------------------------
                                DaesungFunctions.setlist2(self, self.set1) # 한방향 모든요소 컷팅 완료후 컷팅규격으로 남은규격 정보를 가져온다.
                    self.kyus.clear()
                    self.set1.clear()
            df = pd.DataFrame({'장바L' : self.janglenx, '장바W' : self.jangwidx,'색상' : self.spclname, 'BOM' : self.jangbar})
        except: df = 0
        return df
    
    def DBload(self):
        self.tableWidget.clearSelection()
        self.tableWidget.setRowCount(0)
        if self.MESF == 'X':
            self.tableWidget.setRowCount(1)
            self.tableWidget.setRowHeight(0, 40)
            self.tableWidget.setSpan(0, 0, 1, 4)
            item_data = QTableWidgetItem("해당규격은 셋트절단이 불가합니다.")
            item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
            self.tableWidget.setItem(0, 0, item_data)
        else:
            #----------------------------------------------------------------------------
            S_rows = DaesungQuery.selectSpclList(self, self.s_date, self.W_DATA, self.MESF)
            #----------------------------------------------------------------------------
            if S_rows == 'failed': self.connectDBThread()
            else:
                if S_rows == ():
                    self.tableWidget.setRowCount(1)
                    self.tableWidget.setRowHeight(0, 40)
                    self.tableWidget.setSpan(0, 0, 1, 4)
                    item_data = QTableWidgetItem("해당일자의 데이터가 없습니다.")
                    item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                    self.tableWidget.setItem(0, 0, item_data)
                else:
                    df = self.AutoListData(S_rows)
                    self.a_count = len(df)
                    self.splc_array = []
                    t_count, jangl, splc, jangw = 0, '', '', ''
                    for i in range(len(df)):
                        jangbar = df.iloc[i, 3]
                        if df.iloc[i, 0] == jangbar and (self.qty == 1 or i % self.qty == 0):
                            t_count += 1
                            if df.iloc[i, 2] != splc or df.iloc[i, 1] != jangw or df.iloc[i, 0] != jangl:
                                splc = df.iloc[i, 2]
                                jangl = df.iloc[i, 0]
                                jangw = df.iloc[i, 1]
                                self.splc_array.append([jangl, jangw, splc, t_count])
                        if i + 1 == self.a_count: self.splc_array.append([jangl, jangw, splc, t_count + 1])
                    self.spclLoad()
        DaesungFunctions.tableWidth(self, 'SPCL', '', '')
    
    def spclLoad(self):
        if self.splc_array == []:
            self.tableWidget.setRowCount(1)
            self.tableWidget.setRowHeight(0, 40)
            self.tableWidget.setSpan(0, 0, 1, 4)
            item_data = QTableWidgetItem("해당일자의 색상정보가 없습니다.")
            item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
            self.tableWidget.setItem(0, 0, item_data)
        else:
            s_array, c_array = [], []
            for s in range(len(self.splc_array) - 1):
                d_array = []
                for count in range(4):
                    if count == 3: d_array.append(self.splc_array[s + 1][count] - self.splc_array[s][count])
                    else: d_array.append(self.splc_array[s][count])
                s_array.append(d_array)
            s_array.sort(key = lambda x:(x[2], x[0], x[1]))
            #-------------------------------------------------------------
            lenx, widx, name, count = 0, 0, '', -1
            for data in s_array:
                if lenx != int(data[0]) or widx != int(data[1]) or name != data[2]:
                    lenx, widx, name = int(data[0]), int(data[1]), data[2]
                    c_array.append([name, lenx, widx, int(data[3])])
                    count += 1
                else: c_array[count].append(int(data[3]))
            #-------------------------------------------------------------
            self.tableWidget.setRowCount(len(c_array))
            for s_count, s_data in enumerate(c_array):
                self.tableWidget.setRowHeight(s_count, 40)
                for count in range(4):
                    if count == 3 and len(s_data) > 4:
                        print_data = 0
                        for q in range(len(s_data) - count): print_data = print_data + s_data[count + q]
                    else: print_data = s_data[count]
                    item_data = QTableWidgetItem(str(print_data))
                    if count == 0: item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                    else: item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                    self.tableWidget.setItem(s_count, count, item_data)
    
    def showModal(self):
        return super().exec_()

################################################################################################################
#로그 Window
class MesLogWindow(QDialog):
    def __init__(self, date):
        super(MesLogWindow, self).__init__()
        loadUi("ui\DAESUNG_MES_Log.ui", self)
        
        self.date_btn.setText(date)
        self.tableWidget.setStyleSheet(tableStyle)
        self.tableWidget.horizontalHeader().setVisible(True)
        
        self.logLoad()
        
        # 화살표버튼 --------------------------------------------------
        self.top.clicked.connect(lambda: DaesungFunctions.topData(self))
        self.prev.clicked.connect(lambda: DaesungFunctions.prevData(self))
        self.next.clicked.connect(lambda: DaesungFunctions.nextData(self))
        self.bottom.clicked.connect(lambda: DaesungFunctions.bottomData(self))
        
        # 상단버튼 --------------------------------------------------
        self.exit_btn.clicked.connect(lambda: DaesungFunctions.closeWindow(self))
    
    def logLoad(self):
        LogFile = open(logFileName,'rt')
        fileLines = LogFile.readlines()
        LogFile.close()
        self.tableWidget.setRowCount(len(fileLines))
        for count, line in enumerate(fileLines):
            self.tableWidget.setRowHeight(count, 40)
            line = line.replace('\n', '')
            p_date = re.findall('\[[^)]*\]', line)
            c_data = line.replace(p_date[0], '')
            p_date = str(p_date[0].replace('[', '').replace(']', '')).split(',')
            for r_count, i in enumerate([count + 1, p_date[0], c_data]):
                item_data = QTableWidgetItem(str(i))
                if r_count == 2: item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                else: item_data.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                self.tableWidget.setItem(count, r_count, item_data)
        self.tableWidget.setColumnWidth(0, 150)
        self.tableWidget.setColumnWidth(1, 300)
        self.qty_label.setText(str(len(fileLines)))
    
    def showModal(self):
        return super().exec_()

################################################################################################################
#Message Box Window
class MessageWindow(QDialog):
    def __init__(self, parent, flag):
        super(MessageWindow, self).__init__(parent)
        loadUi("ui\MessageBox.ui", self)
        
        if flag == 2:
            self.setFixedSize(400, 270)
            self.stackedWidget.setCurrentWidget(self.close_page)
        else:
            self.setFixedSize(400, 200)
            self.stackedWidget.setCurrentWidget(self.ok_page)
            self.title2_label.setText(flag)
            
        self.ok_btn.clicked.connect(lambda: DaesungFunctions.closeWindow(self))
        self.yes_btn.clicked.connect(self.yesAction)
        self.no_btn.clicked.connect(lambda: DaesungFunctions.closeWindow(self))
        self.recon_btn.clicked.connect(self.reConnect)
    
    def yesAction(self):
        sys.exit(0)
    
    def reConnect(self):
        global RECONNECT_FLAG
        RECONNECT_FLAG = 1
        widget.addWidget(LoginWindow())
        widget.setCurrentIndex(widget.currentIndex() + 1)
        self.deleteLater()
        self.close()
    
    def showModal(self):
        return super().exec_()

################################################################################################################
class LoggerWriter:
    def __init__(self, level): self.level = level
    def write(self, message):
        if message != '\n': self.level(message)
    def flush(self): self.level(sys.stderr)

#---------------------------------------------------------
#Upgrade Window
class UpgradeWindow(QDialog):
    def __init__(self, t_flag, flag):
        super(UpgradeWindow, self).__init__()
        loadUi("ui\progressBar.ui", self)
        
        self.setFixedSize(400, 231)
        self.stackedWidget.setCurrentWidget(self.btns_page)
        self.t_flag = t_flag
        
        if flag == 1:
            self.title2_label.setText("현재 프로그램이 최신버전 입니다.")
            self.title3_label.setText("그래도 업데이트 하시겠습니까?")
            
        self.yes_btn.clicked.connect(self.doAction)
        self.no_btn.clicked.connect(self.noAction)
    
    def SetValue(self, data):
        self.progressBar.setValue(int(data))
    
    def doAction(self):
        sys.stdout = LoggerWriter(logging.debug)
        sys.stderr = LoggerWriter(logging.warning)
        #---------------------------------------------------------
        self.stackedWidget.setCurrentWidget(self.pro_page)
        self.title2_label.setText("실행파일 다운로드 중")
        self.title3_label.setText("loading...")
        time.sleep(0.5)
        try:
            ftp = ftplib.FTP()
            ftp.connect("ehandax.com", 2012)
            ftp.login("woodnsoft!", "woodnsoft.com")
            ftp.cwd('upgrade/WDNS_MES/DAESUNG_MES')
            fd = open("./" + "DAESUNG_MES_Setup.exe", 'wb')
            total = ftp.size("DAESUNG_MES_Setup.exe")
            pbar = tqdm(total = total)
            def bar(data):
                fd.write(data)
                pbar.update(len(data))
                p = str(pbar)
                self.SetValue(p[0:3])
            ftp.retrbinary("RETR " + "DAESUNG_MES_Setup.exe", bar, total)
            pbar.close()
            fd.close()
            os.popen("DAESUNG_MES_Setup.exe")
            time.sleep(1)
        except Exception as e:
            logging.debug(e)
            MessageWindow(self, "업데이트 서버 연결 실패").showModal()
        os.popen('taskkill /im DAESUNG_MES.exe /f')
    
    def noAction(self):
        self.close()
        if self.t_flag == 0: widget.addWidget(LoginWindow())
        widget.showMaximized()
    
    def showModal(self):
        return super().exec_()

################################################################################################################
if __name__ == "__main__":
    suppress_qt_warnings() #해상도별 글자크기 변경 함수
    freeze_support() #멀티프로세스 사용 시 지정함
    app = QApplication(sys.argv)
    #---------------------------------------------------------
    os.environ["NLS_LANG"] = ".UTF8"
    #---------------------------------------------------------
    TodayData = time.strftime('%Y-%m-%d ')
    logFileName = 'Log/%s.log'%time.strftime('%Y%m%d') #Log파일 생성
    dictConfig({'version': 1, 'formatters': {'default': {'format': '[%(asctime)s] %(message)s',} },
                'handlers': { 'file': {'level': 'DEBUG', 'class': 'logging.FileHandler', 'filename': logFileName, 'formatter': 'default',}, },
                'root': {'level': 'DEBUG', 'handlers': ['file']}})
    atexit.register(handleExit)
    logging.debug("프로그램 시작")
    #---------------------------------------------------------
    source, destination = r"./DAESUNG_ADDRESS.ini", r"./DAESUNG_ADDRESS2.ini"
    try:
        config = configparser.ConfigParser()
        config.read('DAESUNG_ADDRESS.ini')
        #---------------------------------------------------
        EMPL_CODE = config['PROCCODE']['EMPL']
        WC_CODE = config['PROCCODE']['WC']
        PROC_CODE = config['PROCCODE']['PROC']
        try: LINE_FLAG = config['PROCCODE']['LINE']
        except: LINE_FLAG = '1'
        try: pwd_flag, pwd_data = config['PROCCODE']['PWD/FLAG'], config['PROCCODE']['PWD/DATA']
        except: pwd_flag, pwd_data = 'f', ''
        #-----------------------------------------------------
        shutil.copyfile(source, destination) #ini파일 복사
        #-----------------------------------------------------
        widget = QtWidgets.QStackedWidget()
        widget.setWindowTitle("DAESUNG MES")
        widget.setWindowIcon(QIcon('./img/ICON.ico'))
        #widget.setFixedSize(1281,956)
        flag = checkVersion() #버전 체크
        if flag == 0: UpgradeWindow(0, 0).showModal()
        else:
            for f in [s_label, ss_label, c_label, cs_label, z_label, z300_label, zs_label, m_label, f_label, if_label]: downLoadFiles(f) #라벨 다운로드
            login_win = LoginWindow()
            widget.addWidget(login_win)
            widget.showMaximized()
    except:
        shutil.copyfile(destination, source)
        main()
    sys.exit(app.exec_())

